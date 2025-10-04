#!/usr/bin/env python3
"""
Original SMOKE Tests Reader
===========================
Read all smoke tests from the original sdm_test_suite.xlsx file.
"""

from openpyxl import load_workbook

def main():
    """Read all smoke tests from original Excel file"""
    excel_file = "sdm_test_suite.xlsx"
    
    print("🔍 ORIGINAL SMOKE TESTS ANALYSIS")
    print("=" * 50)
    print(f"Excel File: {excel_file}")
    print()
    
    try:
        workbook = load_workbook(excel_file)
        print(f"📋 Available sheets: {', '.join(workbook.sheetnames)}")
        print()
        
        # Check SMOKE sheet
        if "SMOKE" in workbook.sheetnames:
            ws = workbook["SMOKE"]
            print("📋 SMOKE Sheet Headers:")
            print("-" * 40)
            
            # Get headers
            headers = []
            for col in range(1, 15):  # Check up to 15 columns
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value))
                    print(f"Column {col}: '{cell_value}'")
            
            print()
            print(f"📊 Total rows in SMOKE sheet: {ws.max_row}")
            print(f"📊 Data rows (excluding header): {ws.max_row - 1}")
            print()
            
            print("📋 All SMOKE Tests:")
            print("-" * 40)
            
            # Read all test data
            all_tests = []
            for row in range(2, ws.max_row + 1):
                test_data = []
                for col in range(1, len(headers) + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    test_data.append(str(cell_value) if cell_value is not None else "")
                
                # Only add if there's a test ID (assuming column 2 is Test_Case_ID)
                if test_data and len(test_data) > 1 and test_data[1]:
                    all_tests.append(test_data)
                    print(f"Row {row}: {test_data[1]} - {test_data[2] if len(test_data) > 2 else 'No name'}")
            
            print()
            print(f"📊 Found {len(all_tests)} smoke tests")
            
            # Show sample test details
            print()
            print("📋 Sample Test Details:")
            print("-" * 40)
            for i, test in enumerate(all_tests[:5]):  # Show first 5 tests
                print(f"Test {i+1}:")
                for j, header in enumerate(headers):
                    if j < len(test):
                        print(f"  {header}: {test[j]}")
                print()
            
            return all_tests, headers
        
        else:
            print("❌ SMOKE sheet not found in the original file")
            return [], []
                
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return [], []

if __name__ == "__main__":
    tests, headers = main()
    if tests:
        print(f"✅ Successfully read {len(tests)} smoke tests from original file")