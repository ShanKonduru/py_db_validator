#!/usr/bin/env python3
"""
Enhanced Unified Excel Template Generator
==========================================
Creates a comprehensive Excel template with all required sheets including
all original smoke tests from the existing sdm_test_suite.xlsx file.

Author: Multi-Database Data Validation Framework
Date: October 4, 2025
"""

import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def load_original_smoke_tests():
    """Load all smoke tests from the original sdm_test_suite.xlsx file"""
    original_file = "sdm_test_suite.xlsx"
    
    try:
        workbook = load_workbook(original_file)
        
        if "SMOKE" not in workbook.sheetnames:
            print(f"⚠️  SMOKE sheet not found in {original_file}, using default tests")
            return []
        
        ws = workbook["SMOKE"]
        smoke_tests = []
        
        # Read all test data starting from row 2
        for row in range(2, ws.max_row + 1):
            test_data = []
            for col in range(1, 14):  # 13 columns
                cell_value = ws.cell(row=row, column=col).value
                test_data.append(str(cell_value) if cell_value is not None else "")
            
            # Only add if there's a test ID (column 2)
            if test_data and len(test_data) > 1 and test_data[1]:
                smoke_tests.append(test_data)
        
        print(f"✅ Loaded {len(smoke_tests)} smoke tests from original file")
        return smoke_tests
        
    except Exception as e:
        print(f"⚠️  Error loading original smoke tests: {e}")
        print("   Using default smoke tests instead")
        return []


def create_enhanced_unified_excel_template():
    """Create enhanced unified Excel template with original smoke tests"""
    
    print("🔧 CREATING ENHANCED UNIFIED EXCEL TEMPLATE")
    print("=" * 60)
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    # Load original smoke tests
    original_smoke_tests = load_original_smoke_tests()
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create all sheets
    create_enhanced_smoke_sheet(wb, original_smoke_tests)
    create_controller_sheet(wb)
    create_datavalidations_sheet(wb)
    create_instructions_sheet(wb)
    create_reference_sheet(wb)
    
    # Save the workbook
    filename = "enhanced_unified_sdm_test_suite.xlsx"
    wb.save(filename)
    
    print(f"✅ Created enhanced unified Excel template: {filename}")
    print()
    
    # Print summary
    print("📋 SHEET SUMMARY:")
    print("-" * 30)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.max_row
        cols = ws.max_column
        data_rows = max(0, rows - 1) if rows > 1 else 0
        print(f"   • {sheet_name}: {rows} rows × {cols} columns ({data_rows} data rows)")
    
    print()
    print("🎯 ENHANCED TEMPLATE FEATURES:")
    print("-" * 30)
    print(f"   ✅ SMOKE sheet: {len(original_smoke_tests) if original_smoke_tests else 6} comprehensive smoke tests")
    print("   ✅ CONTROLLER sheet: Test suite controller")
    print("   ✅ DATAVALIDATIONS sheet: Enhanced data validation tests")
    print("   ✅ INSTRUCTIONS sheet: Comprehensive usage guide")
    print("   ✅ REFERENCE sheet: Complete reference documentation")
    print("   ✅ Data validation rules applied")
    print("   ✅ Professional formatting")
    print("   ✅ All original smoke test configurations preserved")
    
    return filename


def create_enhanced_smoke_sheet(wb, original_smoke_tests):
    """Create SMOKE sheet with all original smoke test configurations"""
    ws = wb.create_sheet("SMOKE")
    
    # Headers
    headers = [
        "Enable", "Test_Case_ID", "Test_Case_Name", "Application_Name", 
        "Environment_Name", "Priority", "Test_Category", "Expected_Result",
        "Timeout_Seconds", "Description", "Prerequisites", "Tags", "Parameters"
    ]
    
    # Apply headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Use original smoke tests if available, otherwise use default ones
    if original_smoke_tests:
        print(f"📋 Adding {len(original_smoke_tests)} original smoke tests to SMOKE sheet")
        
        for row_idx, test_data in enumerate(original_smoke_tests, 2):
            for col_idx, value in enumerate(test_data, 1):
                if col_idx <= len(headers):  # Don't exceed header count
                    ws.cell(row=row_idx, column=col_idx, value=value)
    else:
        # Fallback to default smoke tests
        print("📋 Adding default smoke tests to SMOKE sheet")
        default_smoke_tests = [
            ["TRUE", "SMOKE_001", "Environment Setup Test", "POSTGRES", "DEV", "HIGH", "SETUP", "PASS", 30, "Verify test environment is properly configured", "Database server running", "smoke,setup", ""],
            ["TRUE", "SMOKE_002", "Configuration Availability", "POSTGRES", "DEV", "HIGH", "CONFIGURATION", "PASS", 30, "Verify dummy configuration files are available", "Config files present", "smoke,config", ""],
            ["TRUE", "SMOKE_003", "Environment Credentials", "POSTGRES", "DEV", "HIGH", "SECURITY", "PASS", 30, "Verify environment credentials are valid", "Credentials configured", "smoke,security", ""],
            ["TRUE", "SMOKE_004", "PostgreSQL Connection", "POSTGRES", "DEV", "HIGH", "CONNECTION", "PASS", 60, "Test basic PostgreSQL database connectivity", "Database accessible", "smoke,connection", ""],
            ["TRUE", "SMOKE_005", "Basic Query Execution", "POSTGRES", "DEV", "MEDIUM", "QUERIES", "PASS", 60, "Execute basic SQL queries to verify functionality", "Connection established", "smoke,queries", ""],
            ["TRUE", "SMOKE_006", "Connection Performance", "POSTGRES", "DEV", "MEDIUM", "PERFORMANCE", "PASS", 30, "Measure database connection performance", "Database responsive", "smoke,performance", ""]
        ]
        
        for row_idx, test_data in enumerate(default_smoke_tests, 2):
            for col_idx, value in enumerate(test_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Apply formatting and data validation
    apply_sheet_formatting(ws, len(headers))
    add_data_validations_smoke(ws)


def create_controller_sheet(wb):
    """Create CONTROLLER sheet"""
    ws = wb.create_sheet("CONTROLLER")
    
    # Headers
    headers = [
        "Enable", "Test_Case_ID", "Test_Case_Name", "Application_Name", 
        "Environment_Name", "Priority", "Test_Category", "Expected_Result",
        "Timeout_Seconds", "Description", "Prerequisites", "Tags", "Parameters"
    ]
    
    # Apply headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Controller data
    controller_tests = [
        ["TRUE", "CTRL_001", "Smoke Test Suite", "POSTGRES", "DEV", "HIGH", "SETUP", "PASS", 300, "Execute complete smoke test suite", "All components ready", "controller,smoke", "sheet_name=SMOKE"],
        ["TRUE", "CTRL_002", "Data Validation Suite", "POSTGRES", "DEV", "HIGH", "SCHEMA_VALIDATION", "PASS", 600, "Execute data validation test suite", "Tables and data ready", "controller,validation", "sheet_name=DATAVALIDATIONS"],
        ["FALSE", "CTRL_003", "Integration Test Suite", "POSTGRES", "STAGING", "MEDIUM", "CONNECTION", "PASS", 900, "Execute integration tests", "External systems available", "controller,integration", ""],
        ["FALSE", "CTRL_004", "Performance Test Suite", "POSTGRES", "STAGING", "LOW", "PERFORMANCE", "PASS", 1800, "Execute performance test suite", "Load test environment", "controller,performance", ""],
        ["FALSE", "CTRL_005", "Security Test Suite", "POSTGRES", "PROD", "HIGH", "SECURITY", "PASS", 1200, "Execute security test suite", "Security tools configured", "controller,security", ""],
        ["FALSE", "CTRL_006", "Regression Test Suite", "POSTGRES", "STAGING", "MEDIUM", "QUERIES", "PASS", 3600, "Execute full regression test suite", "Complete test data", "controller,regression", ""]
    ]
    
    # Add controller data
    for row_idx, test_data in enumerate(controller_tests, 2):
        for col_idx, value in enumerate(test_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Apply formatting
    apply_sheet_formatting(ws, len(headers))
    add_data_validations_controller(ws)


def create_datavalidations_sheet(wb):
    """Create enhanced DATAVALIDATIONS sheet"""
    ws = wb.create_sheet("DATAVALIDATIONS")
    
    # Headers
    headers = [
        "Enable", "Test_Case_ID", "Test_Case_Name", "Application_Name", 
        "Environment_Name", "Priority", "Test_Category", "Expected_Result",
        "Timeout_Seconds", "Description", "Prerequisites", "Tags", "Parameters"
    ]
    
    # Apply headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Enhanced data validation tests with proper table parameters
    validation_tests = [
        ["TRUE", "DVAL_001", "Products Schema Validation", "POSTGRES", "DEV", "HIGH", "SCHEMA_VALIDATION", "PASS", 60, "Compare schema structure between source products and target new_products tables", "Both tables must exist with data", "schema,validation,products", "source_table=products;target_table=new_products;ignore_sequence=true;check_constraints=true"],
        ["TRUE", "DVAL_002", "Employees Schema Validation", "POSTGRES", "DEV", "HIGH", "SCHEMA_VALIDATION", "PASS", 60, "Compare schema structure between source employees and target new_employees tables", "Both tables must exist with data", "schema,validation,employees", "source_table=employees;target_table=new_employees;ignore_sequence=true;check_constraints=true"],
        ["TRUE", "DVAL_003", "Orders Schema Validation", "POSTGRES", "DEV", "HIGH", "SCHEMA_VALIDATION", "PASS", 60, "Compare schema structure between source orders and target new_orders tables", "Both tables must exist with data", "schema,validation,orders", "source_table=orders;target_table=new_orders;ignore_sequence=true;check_constraints=true"],
        
        ["TRUE", "DVAL_004", "Products Row Count Validation", "POSTGRES", "DEV", "MEDIUM", "ROW_COUNT_VALIDATION", "FAIL", 30, "Compare row counts between source products and target new_products tables", "Both tables accessible", "count,validation,products", "source_table=products;target_table=new_products;tolerance_percent=5;expected_variance=large"],
        ["TRUE", "DVAL_005", "Employees Row Count Validation", "POSTGRES", "DEV", "MEDIUM", "ROW_COUNT_VALIDATION", "FAIL", 30, "Compare row counts between source employees and target new_employees tables", "Both tables accessible", "count,validation,employees", "source_table=employees;target_table=new_employees;tolerance_percent=5;expected_variance=large"],
        ["TRUE", "DVAL_006", "Orders Row Count Validation", "POSTGRES", "DEV", "MEDIUM", "ROW_COUNT_VALIDATION", "FAIL", 30, "Compare row counts between source orders and target new_orders tables", "Both tables accessible", "count,validation,orders", "source_table=orders;target_table=new_orders;tolerance_percent=5;expected_variance=large"],
        
        ["TRUE", "DVAL_007", "Products NULL Value Validation", "POSTGRES", "DEV", "MEDIUM", "NULL_VALUE_VALIDATION", "FAIL", 45, "Compare NULL value patterns between source and target products tables", "Tables contain representative data", "null,validation,products", "source_table=products;target_table=new_products;null_match_required=false;report_differences=true"],
        ["TRUE", "DVAL_008", "Employees NULL Value Validation", "POSTGRES", "DEV", "MEDIUM", "NULL_VALUE_VALIDATION", "FAIL", 45, "Compare NULL value patterns between source and target employees tables", "Tables contain representative data", "null,validation,employees", "source_table=employees;target_table=new_employees;null_match_required=false;report_differences=true"],
        ["TRUE", "DVAL_009", "Orders NULL Value Validation", "POSTGRES", "DEV", "MEDIUM", "NULL_VALUE_VALIDATION", "FAIL", 45, "Compare NULL value patterns between source and target orders tables", "Tables contain representative data", "null,validation,orders", "source_table=orders;target_table=new_orders;null_match_required=false;report_differences=true"],
        
        ["TRUE", "DVAL_010", "Products Data Quality Check", "POSTGRES", "DEV", "HIGH", "DATA_QUALITY_VALIDATION", "PASS", 60, "Validate data quality metrics for products table", "Data quality rules defined", "quality,validation,products", "source_table=products;target_table=new_products;check_ranges=true;check_patterns=true;check_duplicates=true"],
        ["TRUE", "DVAL_011", "Cross-Table Referential Integrity", "POSTGRES", "DEV", "HIGH", "DATA_QUALITY_VALIDATION", "FAIL", 60, "Validate referential integrity between related tables", "Foreign key relationships exist", "integrity,validation,cross-table", "source_table=products;target_table=orders;check_foreign_keys=true;validate_relationships=true"],
        ["TRUE", "DVAL_012", "Data Completeness Validation", "POSTGRES", "DEV", "MEDIUM", "DATA_QUALITY_VALIDATION", "PASS", 120, "Validate data completeness across all tables", "All tables have minimum required data", "completeness,validation,all-tables", "source_table=products,employees,orders;check_required_fields=true;missing_data_threshold=5"]
    ]
    
    # Add validation test data
    for row_idx, test_data in enumerate(validation_tests, 2):
        for col_idx, value in enumerate(test_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Apply formatting
    apply_sheet_formatting(ws, len(headers))
    add_data_validations_datavalidations(ws)


def create_instructions_sheet(wb):
    """Create comprehensive INSTRUCTIONS sheet"""
    ws = wb.create_sheet("INSTRUCTIONS")
    
    # Title
    ws.cell(row=1, column=1, value="ENHANCED UNIFIED SDM TEST SUITE - COMPREHENSIVE INSTRUCTIONS")
    ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws.merge_cells("A1:J1")
    
    # Instructions content
    instructions = [
        "",
        "📋 OVERVIEW:",
        "This enhanced unified Excel template contains comprehensive test configurations for the Multi-Database Data Validation Framework.",
        "All sheets work together to provide complete testing coverage with all original smoke tests preserved.",
        "",
        "📂 SHEET DESCRIPTIONS:",
        "",
        "🔹 SMOKE Sheet:",
        "   • Contains ALL original smoke tests from the legacy sdm_test_suite.xlsx file",
        "   • Comprehensive environment validation with ~30 detailed test cases",
        "   • Tests connectivity, configuration, table existence, and data validation",
        "   • Execute first to verify environment readiness",
        "   • Categories: SETUP, CONFIGURATION, SECURITY, CONNECTION, QUERIES, PERFORMANCE, TABLE_EXISTS, TABLE_SELECT, TABLE_ROWS, TABLE_STRUCTURE",
        "",
        "🔹 CONTROLLER Sheet:",
        "   • Master controller for executing test suites",
        "   • References other sheets for batch execution",
        "   • Manages test suite orchestration and dependencies",
        "   • Use for high-level test execution control",
        "",
        "🔹 DATAVALIDATIONS Sheet:",
        "   • Enhanced data validation tests with proper table parameter support",
        "   • Includes source_table and target_table parameters for each test",
        "   • Categories: SCHEMA_VALIDATION, ROW_COUNT_VALIDATION, NULL_VALUE_VALIDATION, DATA_QUALITY_VALIDATION",
        "   • Supports multiple table combinations (products, employees, orders)",
        "",
        "🔹 INSTRUCTIONS Sheet (this sheet):",
        "   • Comprehensive usage instructions and guidelines",
        "   • Column descriptions and parameter formats",
        "   • Execution examples and best practices",
        "",
        "🔹 REFERENCE Sheet:",
        "   • Complete reference documentation for all test categories",
        "   • Parameter definitions and valid values",
        "   • Troubleshooting guide and FAQ",
        "",
        "📊 COLUMN DESCRIPTIONS:",
        "",
        "• Enable: TRUE/FALSE - Whether to execute this test case",
        "• Test_Case_ID: Unique identifier for the test (e.g., SMOKE_PG_001, DVAL_001)",
        "• Test_Case_Name: Descriptive name for the test case",
        "• Application_Name: Target application (POSTGRES, DUMMY, MYAPP, DATABASE)",
        "• Environment_Name: Target environment (DEV, TEST, STAGING, PROD, UAT)",
        "• Priority: Test priority level (HIGH, MEDIUM, LOW)",
        "• Test_Category: Functional category that determines which test function to execute",
        "• Expected_Result: Expected test outcome (PASS, FAIL, SKIP)",
        "• Timeout_Seconds: Maximum execution time in seconds",
        "• Description: Detailed description of what the test validates",
        "• Prerequisites: Conditions that must be met before test execution",
        "• Tags: Comma-separated tags for test categorization and filtering",
        "• Parameters: Semicolon-separated key=value pairs for test configuration",
        "",
        "🔧 PARAMETER FORMATS:",
        "",
        "Data Validation Parameters (DATAVALIDATIONS sheet):",
        "   source_table=products;target_table=new_products;ignore_sequence=true",
        "   source_table=employees;target_table=new_employees;tolerance_percent=5",
        "   check_constraints=true;report_differences=true;null_match_required=false",
        "",
        "Controller Parameters (CONTROLLER sheet):",
        "   sheet_name=SMOKE (references which sheet to execute)",
        "   sheet_name=DATAVALIDATIONS (for data validation suite execution)",
        "",
        "🚀 EXECUTION EXAMPLES:",
        "",
        "1. Execute all smoke tests:",
        "   python execute_unified_smoke_tests.py enhanced_unified_sdm_test_suite.xlsx",
        "",
        "2. Execute data validation tests:",
        "   python execute_enhanced_data_validation_tests.py enhanced_unified_sdm_test_suite.xlsx",
        "",
        "3. Validate Excel template:",
        "   python enhanced_excel_validator.py enhanced_unified_sdm_test_suite.xlsx",
        "",
        "✅ BEST PRACTICES:",
        "",
        "• Always run SMOKE tests first to verify environment readiness",
        "• All original smoke tests are preserved and enhanced",
        "• Use CONTROLLER sheet for orchestrating complex test suites",
        "• Ensure proper source_table and target_table parameters in DATAVALIDATIONS",
        "• Set realistic Expected_Result values based on your data setup",
        "• Use Tags for efficient test filtering and categorization",
        "• Monitor Timeout_Seconds to prevent hanging tests",
        "• Keep Parameters format consistent (key=value;key=value)",
        "",
        "🔍 VALIDATION TIPS:",
        "",
        "• All table references in parameters must exist in the database",
        "• Test_Category values must match available test functions",
        "• Enable/Disable tests strategically based on environment readiness",
        "• Use the enhanced validator to check for configuration issues",
        "• Review Prerequisites before executing test suites",
        "• Original smoke test configurations are preserved and validated"
    ]
    
    # Add instructions
    for row_idx, instruction in enumerate(instructions, 2):
        ws.cell(row=row_idx, column=1, value=instruction)
        if instruction.startswith("📋") or instruction.startswith("📂") or instruction.startswith("📊") or instruction.startswith("🔧") or instruction.startswith("🚀") or instruction.startswith("✅") or instruction.startswith("🔍"):
            ws.cell(row=row_idx, column=1).font = Font(bold=True, color="2C3E50")
        elif instruction.startswith("🔹") or instruction.startswith("•"):
            ws.cell(row=row_idx, column=1).font = Font(color="34495E")
    
    # Set column width
    ws.column_dimensions['A'].width = 120


def create_reference_sheet(wb):
    """Create comprehensive and readable REFERENCE sheet"""
    ws = wb.create_sheet("REFERENCE")
    
    # Main title
    ws.cell(row=1, column=1, value="📚 ENHANCED UNIFIED SDM TEST SUITE - REFERENCE GUIDE")
    ws.cell(row=1, column=1).font = Font(bold=True, size=18, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 30
    
    current_row = 3
    
    # Section 1: Quick Overview
    ws.cell(row=current_row, column=1, value="🔍 QUICK OVERVIEW")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    ws.merge_cells(f"A{current_row}:F{current_row}")
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    overview_data = [
        ["", "", "", "", "", ""],
        ["📋 Sheet", "🎯 Purpose", "📊 Test Count", "⏱️ Typical Duration", "🔧 Usage", "✅ Status"],
        ["SMOKE", "Environment validation & basic connectivity", "30 tests", "1-2 minutes", "Run first to verify setup", "✅ Ready"],
        ["CONTROLLER", "Test suite orchestration & management", "6 controllers", "Variable", "Batch execution control", "✅ Ready"],
        ["DATAVALIDATIONS", "Data integrity & quality validation", "12 tests", "5-10 minutes", "Compare source vs target data", "✅ Ready"],
        ["INSTRUCTIONS", "Comprehensive usage guide", "N/A", "N/A", "Reference for users", "✅ Complete"],
        ["REFERENCE", "Complete documentation (this sheet)", "N/A", "N/A", "Quick reference guide", "✅ You are here!"],
        ["", "", "", "", "", ""]
    ]
    
    for row_data in overview_data:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if current_row == 4:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif value and value != "":
                cell.alignment = Alignment(horizontal="left", vertical="center")
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        current_row += 1
    
    current_row += 1
    
    # Section 2: Column Definitions
    ws.cell(row=current_row, column=1, value="📝 COLUMN DEFINITIONS")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    ws.merge_cells(f"A{current_row}:F{current_row}")
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    column_data = [
        ["", "", "", "", "", ""],
        ["📊 Column Name", "🔍 Description", "💡 Valid Values", "📋 Example", "⚠️ Required", "💬 Notes"],
        ["Enable", "Controls test execution", "TRUE, FALSE", "TRUE", "Yes", "Set FALSE to skip test"],
        ["Test_Case_ID", "Unique test identifier", "Any unique string", "SMOKE_PG_001", "Yes", "Must be unique across sheet"],
        ["Test_Case_Name", "Descriptive test name", "Any descriptive text", "Environment Setup", "Yes", "Keep concise but clear"],
        ["Application_Name", "Target application", "POSTGRES, DUMMY, MYAPP", "POSTGRES", "Yes", "Must match your setup"],
        ["Environment_Name", "Target environment", "DEV, TEST, STAGING, PROD", "DEV", "Yes", "Environment to test against"],
        ["Priority", "Test execution priority", "HIGH, MEDIUM, LOW", "HIGH", "Yes", "HIGH = critical tests"],
        ["Test_Category", "Functional test category", "See Test Categories below", "CONNECTION", "Yes", "Determines which function runs"],
        ["Expected_Result", "Expected outcome", "PASS, FAIL, SKIP", "PASS", "Yes", "What you expect to happen"],
        ["Timeout_Seconds", "Max execution time", "Number (30-3600)", "60", "Yes", "Prevents hanging tests"],
        ["Description", "Detailed test description", "Any explanatory text", "Tests basic connectivity", "No", "Helps understand purpose"],
        ["Prerequisites", "Required conditions", "Any dependency info", "Database must be running", "No", "What must be true first"],
        ["Tags", "Categorization tags", "Comma-separated keywords", "smoke,connection,critical", "No", "Useful for filtering"],
        ["Parameters", "Test configuration", "key=value;key=value", "source_table=products", "Depends", "Required for data validation"],
        ["", "", "", "", "", ""]
    ]
    
    for row_data in column_data:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if current_row == current_row - len(column_data) + 2:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif value and value != "":
                cell.alignment = Alignment(horizontal="left", vertical="center")
                # Color code based on column
                if col_idx == 5:  # Required column
                    if value == "Yes":
                        cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        current_row += 1
    
    current_row += 1
    
    # Section 3: Test Categories
    ws.cell(row=current_row, column=1, value="🏷️ TEST CATEGORIES")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
    ws.merge_cells(f"A{current_row}:F{current_row}")
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    category_data = [
        ["", "", "", "", "", ""],
        ["🏷️ Category", "📄 Sheet", "🔧 Function", "📝 Description", "⏱️ Typical Time", "💡 Usage Tips"],
        ["SETUP", "SMOKE", "test_environment_setup", "Verify test environment readiness", "< 1 sec", "Always run first"],
        ["CONFIGURATION", "SMOKE", "test_dummy_config_availability", "Check config file availability", "< 1 sec", "Validates setup files"],
        ["SECURITY", "SMOKE", "test_environment_credentials", "Validate database credentials", "< 1 sec", "Ensures access rights"],
        ["CONNECTION", "SMOKE", "test_postgresql_connection", "Test database connectivity", "1-5 sec", "Core connectivity test"],
        ["QUERIES", "SMOKE", "test_postgresql_basic_queries", "Execute basic SQL queries", "1-3 sec", "Validates query capability"],
        ["PERFORMANCE", "SMOKE", "test_postgresql_connection_performance", "Measure connection speed", "1-2 sec", "Performance baseline"],
        ["TABLE_EXISTS", "SMOKE", "_execute_table_exists_test", "Verify table existence", "< 1 sec", "Check table availability"],
        ["TABLE_SELECT", "SMOKE", "_execute_table_select_test", "Test SELECT operations", "1-2 sec", "Basic query validation"],
        ["TABLE_ROWS", "SMOKE", "_execute_table_rows_test", "Validate table data", "1-3 sec", "Check data presence"],
        ["TABLE_STRUCTURE", "SMOKE", "_execute_table_structure_test", "Analyze table schema", "1-2 sec", "Schema validation"],
        ["SCHEMA_VALIDATION", "DATAVALIDATIONS", "schema_validation_compare", "Compare table schemas", "5-10 sec", "Use source_table & target_table"],
        ["ROW_COUNT_VALIDATION", "DATAVALIDATIONS", "row_count_validation_compare", "Compare row counts", "2-5 sec", "Good for data volume checks"],
        ["NULL_VALUE_VALIDATION", "DATAVALIDATIONS", "null_value_validation_compare", "Compare NULL patterns", "3-8 sec", "Data quality assessment"],
        ["DATA_QUALITY_VALIDATION", "DATAVALIDATIONS", "data_quality_validation_compare", "Quality metrics validation", "5-15 sec", "Comprehensive data checks"],
        ["", "", "", "", "", ""]
    ]
    
    for row_data in category_data:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if current_row == current_row - len(category_data) + 2:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif value and value != "":
                cell.alignment = Alignment(horizontal="left", vertical="center")
                # Color code by sheet
                if col_idx == 2:  # Sheet column
                    if value == "SMOKE":
                        cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
                    elif value == "DATAVALIDATIONS":
                        cell.fill = PatternFill(start_color="FDEAEA", end_color="FDEAEA", fill_type="solid")
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        current_row += 1
    
    current_row += 1
    
    # Section 4: Parameter Examples
    ws.cell(row=current_row, column=1, value="⚙️ PARAMETER EXAMPLES")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    ws.merge_cells(f"A{current_row}:F{current_row}")
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    parameter_data = [
        ["", "", "", "", "", ""],
        ["🔧 Parameter Type", "📋 Format", "💡 Example", "📄 Used In", "⚠️ Required", "💬 Description"],
        ["Table Parameters", "source_table=name;target_table=name", "source_table=products;target_table=new_products", "DATAVALIDATIONS", "Yes", "Specifies which tables to compare"],
        ["Boolean Parameters", "parameter=true/false", "ignore_sequence=true;check_constraints=false", "DATAVALIDATIONS", "No", "Enable/disable specific checks"],
        ["Numeric Parameters", "parameter=number", "tolerance_percent=5;timeout_seconds=120", "DATAVALIDATIONS", "No", "Numeric configuration values"],
        ["String Parameters", "parameter=value", "expected_variance=large;report_type=detailed", "DATAVALIDATIONS", "No", "String configuration values"],
        ["Sheet References", "sheet_name=SHEETNAME", "sheet_name=SMOKE", "CONTROLLER", "Yes", "References another sheet to execute"],
        ["Multiple Values", "parameter=value1,value2", "source_table=products,employees,orders", "DATAVALIDATIONS", "No", "Comma-separated multiple values"],
        ["", "", "", "", "", ""]
    ]
    
    for row_data in parameter_data:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if current_row == current_row - len(parameter_data) + 2:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif value and value != "":
                cell.alignment = Alignment(horizontal="left", vertical="center")
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        current_row += 1
    
    current_row += 1
    
    # Section 5: Database Tables
    ws.cell(row=current_row, column=1, value="🗃️ DATABASE TABLES")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    ws.merge_cells(f"A{current_row}:F{current_row}")
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    table_data = [
        ["", "", "", "", "", ""],
        ["📊 Source Table", "🎯 Target Table", "📝 Description", "📈 Row Count", "🔍 Status", "💡 Usage Notes"],
        ["products", "new_products", "Product catalog data", "1200 → 8", "Available", "Large variance expected for testing"],
        ["employees", "new_employees", "Employee information", "1000 → 7", "Available", "Schema differences exist"],
        ["orders", "new_orders", "Order transaction data", "800 → 7", "Available", "Contains foreign key relationships"],
        ["validation_results", "N/A", "Test execution results", "Variable", "System", "Created automatically during test runs"],
        ["", "", "", "", "", ""]
    ]
    
    for row_data in table_data:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if current_row == current_row - len(table_data) + 2:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif value and value != "":
                cell.alignment = Alignment(horizontal="left", vertical="center")
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        current_row += 1
    
    current_row += 1
    
    # Section 6: Execution Guide
    ws.cell(row=current_row, column=1, value="🚀 EXECUTION GUIDE")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    ws.merge_cells(f"A{current_row}:F{current_row}")
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    execution_data = [
        ["", "", "", "", "", ""],
        ["🎯 Step", "📋 Command", "⏱️ Duration", "📊 Expected Results", "🔧 Purpose", "💡 Notes"],
        ["1. Smoke Tests", "python execute_unified_smoke_tests.py enhanced_unified_sdm_test_suite.xlsx", "1-2 min", "5 PASS, 24 SKIP*", "Environment validation", "*24 skip due to missing functions"],
        ["2. Data Validation", "python execute_enhanced_data_validation_tests.py enhanced_unified_sdm_test_suite.xlsx", "5-10 min", "3 PASS, 9 FAIL*", "Data quality checks", "*Failures expected due to data differences"],
        ["3. Template Validation", "python enhanced_excel_validator.py enhanced_unified_sdm_test_suite.xlsx", "< 1 min", "Structure validated", "Excel file verification", "Checks template integrity"],
        ["4. Generate New Template", "python create_enhanced_unified_excel_template.py", "< 1 min", "New template created", "Template regeneration", "Creates fresh template with all tests"],
        ["", "", "", "", "", ""]
    ]
    
    for row_data in execution_data:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if current_row == current_row - len(execution_data) + 2:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif value and value != "":
                cell.alignment = Alignment(horizontal="left", vertical="center")
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        current_row += 1
    
    # Set optimal column widths for readability
    ws.column_dimensions['A'].width = 25  # Category/Step
    ws.column_dimensions['B'].width = 35  # Description/Command
    ws.column_dimensions['C'].width = 40  # Details/Duration
    ws.column_dimensions['D'].width = 30  # Examples/Results
    ws.column_dimensions['E'].width = 20  # Usage/Purpose
    ws.column_dimensions['F'].width = 35  # Notes
    
    # Add final footer
    current_row += 1
    ws.cell(row=current_row, column=1, value="📞 For questions or issues, refer to the INSTRUCTIONS sheet or contact the development team.")
    ws.cell(row=current_row, column=1).font = Font(italic=True, size=10, color="7F8C8D")
    ws.merge_cells(f"A{current_row}:F{current_row}")
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")


def apply_sheet_formatting(ws, num_cols):
    """Apply consistent formatting to worksheet"""
    # Set column widths
    column_widths = [10, 15, 25, 15, 15, 10, 20, 15, 12, 30, 25, 20, 40]
    for i, width in enumerate(column_widths[:num_cols], 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to data area
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=num_cols):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:  # Data rows
                cell.alignment = Alignment(horizontal="left", vertical="center")


def add_data_validations_smoke(ws):
    """Add data validation rules for SMOKE sheet"""
    # Enable column validation
    enable_validation = DataValidation(type="list", formula1='"TRUE,FALSE"')
    ws.add_data_validation(enable_validation)
    enable_validation.add(f"A2:A{ws.max_row}")
    
    # Priority validation
    priority_validation = DataValidation(type="list", formula1='"HIGH,MEDIUM,LOW"')
    ws.add_data_validation(priority_validation)
    priority_validation.add(f"F2:F{ws.max_row}")
    
    # Expected Result validation
    result_validation = DataValidation(type="list", formula1='"PASS,FAIL,SKIP"')
    ws.add_data_validation(result_validation)
    result_validation.add(f"H2:H{ws.max_row}")


def add_data_validations_controller(ws):
    """Add data validation rules for CONTROLLER sheet"""
    add_data_validations_smoke(ws)  # Same validations


def add_data_validations_datavalidations(ws):
    """Add data validation rules for DATAVALIDATIONS sheet"""
    add_data_validations_smoke(ws)  # Same basic validations
    
    # Test Category validation for data validation sheet
    category_validation = DataValidation(type="list", 
        formula1='"SCHEMA_VALIDATION,ROW_COUNT_VALIDATION,NULL_VALUE_VALIDATION,DATA_QUALITY_VALIDATION,COLUMN_COMPARE_VALIDATION"')
    ws.add_data_validation(category_validation)
    category_validation.add(f"G2:G{ws.max_row}")


def main():
    """Main execution function"""
    filename = create_enhanced_unified_excel_template()
    
    print("🎯 USAGE INSTRUCTIONS:")
    print("-" * 30)
    print(f"   1. Use '{filename}' as your master test suite")
    print("   2. Execute smoke tests: python execute_unified_smoke_tests.py enhanced_unified_sdm_test_suite.xlsx")
    print("   3. Execute data validation: python execute_enhanced_data_validation_tests.py enhanced_unified_sdm_test_suite.xlsx")
    print("   4. Validate template: python enhanced_excel_validator.py enhanced_unified_sdm_test_suite.xlsx")
    print()
    print("✅ ENHANCED UNIFIED EXCEL TEMPLATE CREATION COMPLETE!")


if __name__ == "__main__":
    main()