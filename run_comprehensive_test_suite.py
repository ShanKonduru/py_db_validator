#!/usr/bin/env python3
"""
Comprehensive Test Suite Runner
==============================
Runs all enabled tests from the enhanced_unified_sdm_test_suite.xlsx workbook
based on the Enable flag in each sheet.

This script will execute:
1. SMOKE tests (if enabled)
2. CONTROLLER tests (if enabled) 
3. DATAVALIDATIONS tests (if enabled)

Author: Multi-Database Data Validation Framework
Date: October 4, 2025
"""

import sys
import os
import subprocess
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

def load_excel_file(excel_file):
    """Load and validate Excel file"""
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return None
    
    try:
        workbook = load_workbook(excel_file)
        return workbook
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return None

def count_enabled_tests(workbook, sheet_name):
    """Count enabled tests in a specific sheet"""
    if sheet_name not in workbook.sheetnames:
        return 0, 0
    
    ws = workbook[sheet_name]
    total_tests = 0
    enabled_tests = 0
    
    # Skip header row (row 1)
    for row in range(2, ws.max_row + 1):
        test_id = ws.cell(row=row, column=2).value  # Test_Case_ID column
        enable_flag = ws.cell(row=row, column=1).value  # Enable column
        
        if test_id and str(test_id).strip():  # Valid test ID
            total_tests += 1
            if str(enable_flag).upper() == 'TRUE':
                enabled_tests += 1
    
    return enabled_tests, total_tests

def run_command(command, description):
    """Execute a command and return the result"""
    print(f"\n>> {description}")
    print(f"Command: {command}")
    print("-" * 60)
    
    try:
        # Set encoding environment variable for Windows
        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        
        # Run the command and capture output
        result = subprocess.run(
            command, 
            shell=True, 
            capture_output=True, 
            text=True, 
            cwd=os.getcwd(),
            env=env
        )
        
        # Print the output
        if result.stdout:
            print(result.stdout)
        if result.stderr:
            print("STDERR:", result.stderr)
        
        return result.returncode == 0
    
    except Exception as e:
        print(f"Error executing command: {e}")
        return False

def run_comprehensive_test_suite(excel_file):
    """Run comprehensive test suite based on Enable flags"""
    
    print("🚀 COMPREHENSIVE TEST SUITE EXECUTION")
    print("=" * 70)
    print(f"📅 Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"📊 Excel File: {excel_file}")
    print()
    
    # Load Excel file
    workbook = load_excel_file(excel_file)
    if not workbook:
        return False
    
    # Analyze test counts
    print("📋 TEST SUITE ANALYSIS:")
    print("-" * 30)
    
    smoke_enabled, smoke_total = count_enabled_tests(workbook, "SMOKE")
    controller_enabled, controller_total = count_enabled_tests(workbook, "CONTROLLER")
    dataval_enabled, dataval_total = count_enabled_tests(workbook, "DATAVALIDATIONS")
    
    print(f"🔥 SMOKE Tests: {smoke_enabled}/{smoke_total} enabled")
    print(f"🎛️  CONTROLLER Tests: {controller_enabled}/{controller_total} enabled")
    print(f"📊 DATAVALIDATIONS Tests: {dataval_enabled}/{dataval_total} enabled")
    print()
    
    total_enabled = smoke_enabled + controller_enabled + dataval_enabled
    total_tests = smoke_total + controller_total + dataval_total
    
    print(f"🎯 TOTAL: {total_enabled}/{total_tests} tests enabled ({total_enabled/total_tests*100:.1f}%)")
    print()
    
    if total_enabled == 0:
        print("⚠️  No tests are enabled! Please check Enable flags in Excel sheets.")
        return False
    
    # Execution plan
    print("📋 EXECUTION PLAN:")
    print("-" * 25)
    
    execution_steps = []
    
    if smoke_enabled > 0:
        execution_steps.append({
            'name': 'SMOKE Tests',
            'command': f'python execute_unified_smoke_tests.py {excel_file}',
            'description': f'Environment validation ({smoke_enabled} tests)',
            'enabled_count': smoke_enabled
        })
    
    if controller_enabled > 0:
        execution_steps.append({
            'name': 'CONTROLLER Tests', 
            'command': f'python execute_unified_smoke_tests.py {excel_file} --sheet CONTROLLER',
            'description': f'Test suite orchestration ({controller_enabled} tests)',
            'enabled_count': controller_enabled
        })
    
    if dataval_enabled > 0:
        execution_steps.append({
            'name': 'DATAVALIDATIONS Tests',
            'command': f'python execute_enhanced_data_validation_tests.py {excel_file}',
            'description': f'Data quality validation ({dataval_enabled} tests)',
            'enabled_count': dataval_enabled
        })
    
    for i, step in enumerate(execution_steps, 1):
        print(f"   {i}. {step['name']}: {step['description']}")
    
    print()
    
    # Execute test suites
    print("🎬 STARTING TEST EXECUTION:")
    print("=" * 40)
    
    results = []
    
    for i, step in enumerate(execution_steps, 1):
        print(f"\n🎯 STEP {i}/{len(execution_steps)}: {step['name']}")
        print(f"📊 Expected to run {step['enabled_count']} enabled tests")
        
        success = run_command(step['command'], step['description'])
        
        results.append({
            'name': step['name'],
            'success': success,
            'enabled_count': step['enabled_count']
        })
        
        if success:
            print(f"✅ {step['name']} completed successfully")
        else:
            print(f"⚠️  {step['name']} completed with issues")
    
    # Final summary
    print("\n" + "=" * 70)
    print("📊 COMPREHENSIVE TEST SUITE SUMMARY")
    print("=" * 70)
    print(f"📅 Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    successful_suites = sum(1 for r in results if r['success'])
    total_suites = len(results)
    
    print("🎯 EXECUTION RESULTS:")
    print("-" * 25)
    
    for result in results:
        status = "✅ SUCCESS" if result['success'] else "⚠️  ISSUES"
        print(f"   {result['name']}: {status} ({result['enabled_count']} tests)")
    
    print()
    print(f"📈 Suite Success Rate: {successful_suites}/{total_suites} ({successful_suites/total_suites*100:.1f}%)")
    print(f"🎯 Total Enabled Tests Executed: {sum(r['enabled_count'] for r in results)}")
    print()
    
    if successful_suites == total_suites:
        print("🎉 ALL TEST SUITES COMPLETED SUCCESSFULLY!")
        return True
    else:
        print("⚠️  SOME TEST SUITES HAD ISSUES - Check individual results above")
        return False

def main():
    """Main execution function"""
    
    # Default Excel file
    default_excel = "enhanced_unified_sdm_test_suite.xlsx"
    
    # Get Excel file from command line or use default
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = default_excel
    
    print("🎯 COMPREHENSIVE TEST SUITE RUNNER")
    print("=" * 50)
    print("Executes all enabled tests from Excel workbook")
    print()
    print("📋 USAGE:")
    print(f"   python {Path(__file__).name}")
    print(f"   python {Path(__file__).name} <excel_file>")
    print()
    print("🔍 FEATURES:")
    print("   • Analyzes Enable flags in all sheets")
    print("   • Runs only enabled tests")
    print("   • Provides comprehensive reporting")
    print("   • Executes SMOKE → CONTROLLER → DATAVALIDATIONS")
    print()
    
    # Run comprehensive test suite
    success = run_comprehensive_test_suite(excel_file)
    
    print("\n🎯 NEXT STEPS:")
    print("-" * 15)
    if success:
        print("   • Review individual test results above")
        print("   • Check any failed tests for configuration issues")
        print("   • Use results to improve data quality")
    else:
        print("   • Check Excel file Enable flags")
        print("   • Verify test configurations and parameters")
        print("   • Review error messages above")
    
    print()
    print("📖 For detailed guidance, check the REFERENCE sheet in Excel!")
    
    return 0 if success else 1

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)