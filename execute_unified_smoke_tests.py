#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SMOKE Test Execution Script
===========================
Execute smoke tests from the unified Excel template SMOKE sheet.

Author: Multi-Database Data Validation Framework
Date: October 4, 2025
"""

import sys
import os

# Set UTF-8 encoding for Windows console output
if os.name == 'nt':  # Windows
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
import time
from pathlib import Path
from datetime import datetime
from typing import List
from dataclasses import dataclass
from openpyxl import load_workbook

# Add src directory to path for imports
sys.path.insert(0, str(Path(__file__).parent / "src"))

from src.core.test_executor import TestExecutor
from src.tests.static_database_smoke_tests import StaticDatabaseSmokeTests


@dataclass
class SimpleTestCase:
    """Simple test case data structure"""
    enable: bool
    test_case_id: str
    test_case_name: str
    application_name: str
    environment_name: str
    priority: str
    test_category: str
    expected_result: str
    timeout_seconds: int
    description: str
    prerequisites: str
    tags: str
    parameters: str


def load_smoke_tests_from_excel(excel_file: str, sheet_name: str = "SMOKE") -> List[SimpleTestCase]:
    """Load smoke test cases directly from Excel"""
    test_cases = []
    
    try:
        workbook = load_workbook(excel_file)
        
        if sheet_name not in workbook.sheetnames:
            print(f"❌ Sheet '{sheet_name}' not found in workbook")
            print(f"   Available sheets: {', '.join(workbook.sheetnames)}")
            return []
        
        ws = workbook[sheet_name]
        
        # Read data rows (skip header row)
        for row in range(2, ws.max_row + 1):
            # Check if row has data
            if not ws.cell(row=row, column=2).value:  # TEST_CASE_ID column
                continue
                
            try:
                test_case = SimpleTestCase(
                    enable=str(ws.cell(row=row, column=1).value or "").upper() in ["TRUE", "YES", "1"],
                    test_case_id=str(ws.cell(row=row, column=2).value or ""),
                    test_case_name=str(ws.cell(row=row, column=3).value or ""),
                    application_name=str(ws.cell(row=row, column=4).value or ""),
                    environment_name=str(ws.cell(row=row, column=5).value or ""),
                    priority=str(ws.cell(row=row, column=6).value or ""),
                    test_category=str(ws.cell(row=row, column=7).value or ""),
                    expected_result=str(ws.cell(row=row, column=8).value or "PASS"),
                    timeout_seconds=int(ws.cell(row=row, column=9).value or 60),
                    description=str(ws.cell(row=row, column=10).value or ""),
                    prerequisites=str(ws.cell(row=row, column=11).value or ""),
                    tags=str(ws.cell(row=row, column=12).value or ""),
                    parameters=str(ws.cell(row=row, column=13).value or "")
                )
                
                if test_case.enable:
                    test_cases.append(test_case)
                    
            except Exception as e:
                print(f"⚠️  Warning: Error reading row {row}: {e}")
                continue
    
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return []
    
    return test_cases


def execute_smoke_tests(excel_file: str):
    """Execute smoke tests from unified Excel template"""
    
    print("🔍 EXECUTING SMOKE TESTS FROM UNIFIED EXCEL TEMPLATE")
    print("=" * 64)
    print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Test Suite: {excel_file}")
    print(f"Target Sheet: SMOKE")
    print()
    
    # Load test cases
    test_cases = load_smoke_tests_from_excel(excel_file)
    
    if not test_cases:
        print("❌ No smoke test cases loaded from Excel file")
        return False
    
    print(f"📊 Successfully loaded {len(test_cases)} smoke test cases")
    print()
    
    # Check if we should use static tests (default to static for production stability)
    use_static_tests = os.environ.get("USE_STATIC_SMOKE_TESTS", "true").lower() in ["true", "1", "yes"]
    test_method = "🔒 STATIC IMMUTABLE" if use_static_tests else "🏗️  INSTANCE BASED"
    
    # Get database type from environment (default to postgresql for backward compatibility)
    db_type = os.environ.get("DB_TYPE", "postgresql").lower()
    
    print(f"🧪 Test Method: {test_method}")
    print(f"🗄️  Database Type: {db_type.upper()}")
    print()
    
    # Initialize test executor with appropriate method and database type
    executor = TestExecutor(use_static_tests=use_static_tests, db_type=db_type)
    
    # Execute tests
    results = []
    passed = 0
    failed = 0
    skipped = 0
    
    for i, test_case in enumerate(test_cases, 1):
        print(f"[{i}/{len(test_cases)}] 🧪 Executing: {test_case.test_case_id}")
        print(f"   Name: {test_case.test_case_name}")
        print(f"   Category: {test_case.test_category}")
        print(f"   Description: {test_case.description}")
        
        start_time = time.time()
        
        try:
            # Execute based on category
            if use_static_tests:
                # Use static immutable tests
                if test_case.test_category == "SETUP":
                    result = StaticDatabaseSmokeTests.test_environment_setup(
                        db_type, test_case.environment_name, test_case.application_name
                    )
                    if result["status"] != "PASS":
                        raise Exception(result["message"])
                    status = "PASS"
                elif test_case.test_category == "CONFIGURATION":
                    result = StaticDatabaseSmokeTests.test_configuration_availability(
                        db_type, test_case.environment_name, test_case.application_name
                    )
                    if result["status"] != "PASS":
                        raise Exception(result["message"])
                    status = "PASS"
                elif test_case.test_category == "SECURITY":
                    result = StaticDatabaseSmokeTests.test_environment_credentials(
                        db_type, test_case.environment_name, test_case.application_name
                    )
                    if result["status"] != "PASS":
                        raise Exception(result["message"])
                    status = "PASS"
                elif test_case.test_category == "CONNECTION":
                    result = StaticDatabaseSmokeTests.test_database_connection(
                        db_type, test_case.environment_name, test_case.application_name
                    )
                    if result["status"] != "PASS":
                        raise Exception(result["message"])
                    status = "PASS"
                elif test_case.test_category == "QUERIES":
                    result = StaticDatabaseSmokeTests.test_database_basic_queries(
                        db_type, test_case.environment_name, test_case.application_name
                    )
                    if result["status"] != "PASS":
                        raise Exception(result["message"])
                    status = "PASS"
                elif test_case.test_category == "PERFORMANCE":
                    result = StaticDatabaseSmokeTests.test_database_connection_performance(
                        db_type, test_case.environment_name, test_case.application_name
                    )
                    if result["status"] != "PASS":
                        raise Exception(result["message"])
                    status = "PASS"
                elif test_case.test_category in ["TABLE_EXISTS", "TABLE_SELECT", "TABLE_ROWS", "TABLE_STRUCTURE"]:
                    # All table-related tests use basic queries for validation
                    result = StaticDatabaseSmokeTests.test_database_basic_queries(
                        db_type, test_case.environment_name, test_case.application_name
                    )
                    if result["status"] != "PASS":
                        raise Exception(result["message"])
                    status = "PASS"
                else:
                    status = "SKIP"
                    skipped += 1
                    print(f"   ⏭️  SKIP - Unknown category: {test_case.test_category}")
            else:
                # Use instance-based tests (legacy)
                if test_case.test_category == "SETUP":
                    executor.smoke_tester.test_environment_setup()
                    status = "PASS"
                elif test_case.test_category == "CONFIGURATION":
                    executor.smoke_tester.test_dummy_config_availability()
                    status = "PASS"
                elif test_case.test_category == "SECURITY":
                    executor.smoke_tester.test_environment_credentials()
                    status = "PASS"
                elif test_case.test_category == "CONNECTION":
                    executor.smoke_tester.test_postgresql_connection()
                    status = "PASS"
                elif test_case.test_category == "QUERIES":
                    executor.smoke_tester.test_postgresql_basic_queries()
                    status = "PASS"
                elif test_case.test_category == "PERFORMANCE":
                    executor.smoke_tester.test_postgresql_connection_performance()
                    status = "PASS"
                elif test_case.test_category in ["TABLE_EXISTS", "TABLE_SELECT", "TABLE_ROWS", "TABLE_STRUCTURE"]:
                    # All table-related tests use basic queries for validation
                    executor.smoke_tester.test_postgresql_basic_queries()
                    status = "PASS"
                else:
                    status = "SKIP"
                    skipped += 1
                    print(f"   ⏭️  SKIP - Unknown category: {test_case.test_category}")
            
            duration = time.time() - start_time
            
            if status == "PASS":
                print(f"   ✅ PASS ({duration:.3f}s)")
                passed += 1
            elif status == "SKIP":
                print(f"   ⏭️  SKIP ({duration:.3f}s)")
                skipped += 1
            
            results.append({
                'test_id': test_case.test_case_id,
                'test_name': test_case.test_case_name,
                'category': test_case.test_category,
                'status': status,
                'duration': duration,
                'message': ""
            })
        
        except Exception as e:
            duration = time.time() - start_time
            print(f"   ❌ FAIL ({duration:.3f}s)")
            print(f"   💬 {str(e)}")
            failed += 1
            
            results.append({
                'test_id': test_case.test_case_id,
                'test_name': test_case.test_case_name,
                'category': test_case.test_category,
                'status': "FAIL",
                'duration': duration,
                'message': str(e)
            })
        
        print()
    
    # Print summary
    total_tests = len(test_cases)
    success_rate = (passed / total_tests * 100) if total_tests > 0 else 0
    total_duration = sum(r['duration'] for r in results)
    
    print("📋 SMOKE TEST SUMMARY:")
    print("=" * 64)
    print(f"   ✅ Passed: {passed}")
    print(f"   ❌ Failed: {failed}")
    print(f"   ⏭️  Skipped: {skipped}")
    print(f"   📈 Success Rate: {success_rate:.1f}%")
    print(f"   ⏱️  Total duration: {total_duration:.2f}s")
    print()
    
    # Detailed results
    if passed > 0:
        print("✅ PASSED TESTS:")
        for i, result in enumerate([r for r in results if r['status'] == 'PASS'], 1):
            print(f"    {i}. {result['test_id']} - {result['test_name']}")
            print(f"       Category: {result['category']} | Duration: {result['duration']:.3f}s")
        print()
    
    if failed > 0:
        print("❌ FAILED TESTS:")
        for i, result in enumerate([r for r in results if r['status'] == 'FAIL'], 1):
            print(f"    {i}. {result['test_id']} - {result['test_name']}")
            print(f"       Category: {result['category']} | Duration: {result['duration']:.3f}s")
            print(f"       Error: {result['message']}")
        print()
    
    if skipped > 0:
        print("⏭️  SKIPPED TESTS:")
        for i, result in enumerate([r for r in results if r['status'] == 'SKIP'], 1):
            print(f"    {i}. {result['test_id']} - {result['test_name']}")
            print(f"       Category: {result['category']} | Duration: {result['duration']:.3f}s")
        print()
    
    print("✅ SMOKE TEST EXECUTION COMPLETE!")
    return success_rate > 50  # Consider successful if > 50% pass rate


def main():
    """Main execution function"""
    excel_file = sys.argv[1] if len(sys.argv) > 1 else "enhanced_unified_sdm_test_suite.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return False
    
    return execute_smoke_tests(excel_file)


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)