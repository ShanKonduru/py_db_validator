#!/usr/bin/env python
"""
Examine SMOKE sheet structure
"""
import openpyxl

def examine_smoke_sheet():
    wb = openpyxl.load_workbook('sdm_test_suite.xlsx')
    ws = wb['SMOKE']
    
    print('Current SMOKE sheet structure:')
    print('Headers:')
    headers = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            headers.append(header)
            print(f'  Column {col}: {header}')
    
    print('\nSample data rows:')
    for row in range(2, min(6, ws.max_row + 1)):
        row_data = []
        for col in range(1, len(headers) + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                row_data.append(str(value)[:20])
        if row_data:
            print(f'  Row {row}: {" | ".join(row_data)}')

if __name__ == '__main__':
    examine_smoke_sheet()