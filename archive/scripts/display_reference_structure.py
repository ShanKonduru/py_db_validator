#!/usr/bin/env python3
"""
Display Reference Sheet Structure
================================
Shows the structure and content of the improved REFERENCE sheet
"""

from pathlib import Path
import sys

# Add project root to path
project_root = Path(__file__).resolve().parent
sys.path.insert(0, str(project_root))

from openpyxl import load_workbook

def display_reference_sheet():
    """Display the structure of the REFERENCE sheet"""
    
    print("📚 ENHANCED REFERENCE SHEET STRUCTURE")
    print("=" * 60)
    
    try:
        # Load the workbook
        wb = load_workbook("enhanced_unified_sdm_test_suite.xlsx")
        
        if "REFERENCE" not in wb.sheetnames:
            print("❌ REFERENCE sheet not found!")
            return
        
        ws = wb["REFERENCE"]
        
        print(f"📊 Sheet Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        print()
        
        # Show section headers and structure
        section_headers = []
        for row in range(1, min(ws.max_row + 1, 80)):  # First 80 rows
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and str(cell_value).strip():
                cell_str = str(cell_value).strip()
                
                # Check if it's a section header (contains emoji or is formatted as header)
                if any(char in cell_str for char in ['📚', '🔍', '📝', '🏷️', '⚙️', '🗃️', '🚀']):
                    section_headers.append((row, cell_str))
        
        print("📋 SECTION STRUCTURE:")
        print("-" * 40)
        for row, header in section_headers:
            print(f"Row {row:2d}: {header}")
        
        print()
        print("✅ IMPROVED FEATURES:")
        print("-" * 30)
        print("   🎨 Professional formatting with color-coded sections")
        print("   📊 Structured tables with clear headers and borders")
        print("   🔍 Six-column layout for maximum information")
        print("   💡 Usage tips and practical examples")
        print("   ⚙️ Comprehensive parameter documentation")
        print("   🗃️ Database table reference with row counts")
        print("   🚀 Step-by-step execution guide")
        print("   📞 Contact information and support guidance")
        
        print()
        print("🎯 COLUMN LAYOUT:")
        print("-" * 20)
        print("   Column A (25 chars): Categories/Steps")
        print("   Column B (35 chars): Descriptions/Commands") 
        print("   Column C (40 chars): Details/Examples")
        print("   Column D (30 chars): Usage/Results")
        print("   Column E (20 chars): Status/Purpose")
        print("   Column F (35 chars): Notes/Tips")
        
        # Show sample content from each major section
        print()
        print("📖 SAMPLE CONTENT PREVIEW:")
        print("-" * 30)
        
        # Find and show a few sample rows from different sections
        sample_rows = [3, 10, 20, 30, 40, 50]
        for sample_row in sample_rows:
            if sample_row <= ws.max_row:
                row_data = []
                for col in range(1, 7):  # 6 columns
                    cell_value = ws.cell(row=sample_row, column=col).value
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        if len(cell_str) > 30:
                            cell_str = cell_str[:27] + "..."
                        row_data.append(cell_str)
                    else:
                        row_data.append("")
                
                if any(row_data):  # Only show if there's content
                    print(f"Row {sample_row}: {' | '.join(row_data[:3])}...")
        
    except Exception as e:
        print(f"❌ Error reading reference sheet: {e}")

if __name__ == "__main__":
    display_reference_sheet()