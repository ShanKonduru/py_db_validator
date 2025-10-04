#!/usr/bin/env python3
"""
Direct Excel Parameter Test
===========================
Directly test parameter parsing from the enhanced Excel template.
"""

from openpyxl import load_workbook

def parse_test_params(parameters: str) -> dict:
    """Parse test parameters from parameters string - same logic as test executor"""
    params = {}
    if not parameters:
        return params
        
    # Simple parameter parsing from parameters
    # Expected format: "source_table=products;target_table=new_products;column_name=product_name"
    for param in parameters.split(';'):
        if '=' in param:
            key, value = param.strip().split('=', 1)
            params[key.strip()] = value.strip()
    
    return params

def main():
    """Test parameter parsing directly from Excel"""
    excel_file = "enhanced_sdm_test_suite.xlsx"
    
    print("🔍 DIRECT EXCEL PARAMETER TEST")
    print("=" * 50)
    print(f"Excel File: {excel_file}")
    print()
    
    workbook = load_workbook(excel_file)
    ws = workbook["DATAVALIDATIONS"]
    
    print("📋 Testing parameter parsing for each test case:")
    print()
    
    for row in range(2, 14):  # Rows 2-13 (data rows)
        test_id = ws.cell(row=row, column=2).value  # TEST_CASE_ID
        test_name = ws.cell(row=row, column=3).value  # TEST_CASE_NAME
        test_category = ws.cell(row=row, column=7).value  # TEST_CATEGORY
        parameters = ws.cell(row=row, column=13).value  # PARAMETERS
        
        if test_id and test_category:
            print(f"[{row-1}] {test_id} - {test_name}")
            print(f"    Category: {test_category}")
            print(f"    Parameters: '{parameters}'")
            
            # Parse parameters using the same logic as test executor
            params = parse_test_params(parameters or "")
            
            source_table = params.get('source_table', 'DEFAULT: products')
            target_table = params.get('target_table', 'DEFAULT: new_products')
            
            print(f"    Parsed source_table: {source_table}")
            print(f"    Parsed target_table: {target_table}")
            
            # Show if we're using defaults or actual parameters
            if 'source_table' in params and 'target_table' in params:
                print(f"    ✅ Using PROPER table parameters!")
            else:
                print(f"    ❌ Using DEFAULT parameters!")
            print()

if __name__ == "__main__":
    main()