#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Simple Test Suite Runner
========================
Runs all enabled tests from enhanced_unified_sdm_test_suite.xlsx
based on Enable flags. Simple version without Unicode characters.

Author: Multi-Database Data Validation Framework  
Date: October 4, 2025
"""

import sys
import os
import subprocess
from pathlib import Path

# Set UTF-8 encoding for Windows console output
if os.name == 'nt':  # Windows
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
from datetime import datetime
from openpyxl import load_workbook

def count_enabled_tests(workbook, sheet_name):
    """Count enabled tests in a specific sheet"""
    if sheet_name not in workbook.sheetnames:
        return 0, 0
    
    ws = workbook[sheet_name]
    total_tests = 0
    enabled_tests = 0
    
    # Skip header row (row 1)
    for row in range(2, ws.max_row + 1):
        test_id = ws.cell(row=row, column=2).value  # Test_Case_ID column
        enable_flag = ws.cell(row=row, column=1).value  # Enable column
        
        if test_id and str(test_id).strip():  # Valid test ID
            total_tests += 1
            if str(enable_flag).upper() == 'TRUE':
                enabled_tests += 1
    
    return enabled_tests, total_tests

def run_test_suite(excel_file):
    """Run test suite based on Enable flags"""
    
    print("=" * 60)
    print("COMPREHENSIVE TEST SUITE RUNNER")
    print("=" * 60)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Excel File: {excel_file}")
    print()
    
    # Load Excel file
    if not os.path.exists(excel_file):
        print(f"ERROR: Excel file not found: {excel_file}")
        return False
    
    try:
        workbook = load_workbook(excel_file)
    except Exception as e:
        print(f"ERROR: Could not load Excel file: {e}")
        return False
    
    # Count enabled tests
    smoke_enabled, smoke_total = count_enabled_tests(workbook, "SMOKE")
    controller_enabled, controller_total = count_enabled_tests(workbook, "CONTROLLER")
    dataval_enabled, dataval_total = count_enabled_tests(workbook, "DATAVALIDATIONS")
    
    print("TEST ANALYSIS:")
    print("-" * 20)
    print(f"SMOKE Tests: {smoke_enabled}/{smoke_total} enabled")
    print(f"CONTROLLER Tests: {controller_enabled}/{controller_total} enabled") 
    print(f"DATAVALIDATIONS Tests: {dataval_enabled}/{dataval_total} enabled")
    print()
    
    total_enabled = smoke_enabled + controller_enabled + dataval_enabled
    print(f"TOTAL ENABLED TESTS: {total_enabled}")
    print()
    
    if total_enabled == 0:
        print("WARNING: No tests are enabled!")
        print("Please check Enable flags in Excel sheets.")
        return False
    
    # Run test suites
    results = []
    
    if smoke_enabled > 0:
        print("=" * 60)
        print(f"RUNNING SMOKE TESTS ({smoke_enabled} enabled)")
        print("=" * 60)
        result = subprocess.run(
            f"python execute_unified_smoke_tests.py {excel_file}",
            shell=True
        )
        results.append(("SMOKE", result.returncode == 0, smoke_enabled))
        print()
    
    if dataval_enabled > 0:
        print("=" * 60)
        print(f"RUNNING DATA VALIDATION TESTS ({dataval_enabled} enabled)")
        print("=" * 60)
        result = subprocess.run(
            f"python execute_enhanced_data_validation_tests.py {excel_file}",
            shell=True
        )
        results.append(("DATAVALIDATIONS", result.returncode == 0, dataval_enabled))
        print()
    
    # Summary
    print("=" * 60)
    print("COMPREHENSIVE TEST SUITE SUMMARY")
    print("=" * 60)
    print(f"Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    print("RESULTS:")
    print("-" * 15)
    for suite_name, success, count in results:
        status = "SUCCESS" if success else "ISSUES"
        print(f"{suite_name}: {status} ({count} tests)")
    
    successful = sum(1 for _, success, _ in results if success)
    total_suites = len(results)
    
    print()
    print(f"SUITE SUCCESS RATE: {successful}/{total_suites}")
    print(f"TOTAL TESTS EXECUTED: {sum(count for _, _, count in results)}")
    
    return successful == total_suites

def main():
    """Main function"""
    
    # Get Excel file from command line or use default
    excel_file = sys.argv[1] if len(sys.argv) > 1 else "enhanced_unified_sdm_test_suite.xlsx"
    
    print("USAGE:")
    print(f"  python {Path(__file__).name}")
    print(f"  python {Path(__file__).name} <excel_file>")
    print()
    print("FEATURES:")
    print("  - Analyzes Enable flags in all test sheets")
    print("  - Runs only tests where Enable=TRUE")
    print("  - Executes SMOKE and DATAVALIDATIONS test suites")
    print("  - Provides summary of all results")
    print()
    
    success = run_test_suite(excel_file)
    
    print()
    print("NEXT STEPS:")
    if success:
        print("  - Review test results above")
        print("  - Check any failed individual tests")
    else:
        print("  - Check Enable flags in Excel sheets")
        print("  - Verify test configurations")
    
    return 0 if success else 1

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)