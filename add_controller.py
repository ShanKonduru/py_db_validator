#!/usr/bin/env python
"""
Add CONTROLLER sheet to default Excel file
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def add_controller_sheet():
    """Add CONTROLLER sheet to sdm_test_suite.xlsx"""
    # Load workbook
    wb = openpyxl.load_workbook('sdm_test_suite.xlsx')
    
    # Create CONTROLLER sheet if not exists
    if 'CONTROLLER' not in wb.sheetnames:
        ws = wb.create_sheet('CONTROLLER')
        
        # Headers
        headers = ['Enable', 'Sheet_Name', 'Description', 'Priority']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='D63384', end_color='D63384', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sample data
        data = [
            ['TRUE', 'SMOKE', 'PostgreSQL smoke tests - basic connectivity and functionality', 'HIGH'],
            ['FALSE', 'INTEGRATION', 'Integration tests with external systems', 'MEDIUM'],
            ['FALSE', 'PERFORMANCE', 'Performance and load testing suite', 'LOW'],
            ['FALSE', 'SECURITY', 'Security and penetration testing', 'HIGH']
        ]
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx).value = value
        
        # Add dropdowns
        enable_dv = DataValidation(type='list', formula1='"TRUE,FALSE"', allow_blank=False)
        enable_dv.add('A2:A1000')
        ws.add_data_validation(enable_dv)
        
        priority_dv = DataValidation(type='list', formula1='"HIGH,MEDIUM,LOW"', allow_blank=True)
        priority_dv.add('D2:D1000')
        ws.add_data_validation(priority_dv)
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        
        ws.freeze_panes = 'A2'
    
    # Save
    wb.save('sdm_test_suite.xlsx')
    print('✅ CONTROLLER sheet added to sdm_test_suite.xlsx')

if __name__ == '__main__':
    add_controller_sheet()