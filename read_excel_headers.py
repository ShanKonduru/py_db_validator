#!/usr/bin/env python3
"""
Simple Excel Header Reader
=========================
Read the actual headers from the enhanced Excel template to debug the case sensitivity issue.
"""

from openpyxl import load_workbook

def main():
    """Read headers from Excel file"""
    excel_file = "unified_sdm_test_suite.xlsx"
    
    print("🔍 EXCEL HEADER ANALYSIS")
    print("=" * 50)
    print(f"Excel File: {excel_file}")
    print()
    
    try:
        workbook = load_workbook(excel_file)
        print(f"📋 Available sheets: {', '.join(workbook.sheetnames)}")
        print()
        
        # Check DATAVALIDATIONS sheet
        if "DATAVALIDATIONS" in workbook.sheetnames:
            ws = workbook["DATAVALIDATIONS"]
            print("📋 DATAVALIDATIONS Sheet Headers:")
            print("-" * 40)
            
            for col in range(1, 14):  # First 13 columns
                cell_value = ws.cell(row=1, column=col).value
                print(f"Column {col}: '{cell_value}'")
            
            print()
            print("📋 First few data rows:")
            print("-" * 40)
            
            for row in range(2, 5):  # First 3 data rows
                row_data = []
                for col in range(1, 14):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                print(f"Row {row}: {row_data}")
                
                # Check specifically for parameters column (column 13)
                params = ws.cell(row=row, column=13).value
                print(f"  Parameters (Column 13): '{params}'")
        
        # Also check SMOKE sheet
        if "SMOKE" in workbook.sheetnames:
            ws = workbook["SMOKE"]
            print()
            print("📋 SMOKE Sheet Headers:")
            print("-" * 40)
            
            for col in range(1, 14):  # First 13 columns
                cell_value = ws.cell(row=1, column=col).value
                print(f"Column {col}: '{cell_value}'")
            
            print()
            print("📋 SMOKE Sheet - First few data rows:")
            print("-" * 40)
            
            for row in range(2, 4):  # First 2 data rows
                row_data = []
                for col in range(1, 14):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                print(f"Row {row}: {row_data}")
                
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")


if __name__ == "__main__":
    main()