#!/usr/bin/env python
"""
Comprehensive unit tests for Multi Sheet Controller
"""
import unittest
import pytest
import tempfile
import os
from unittest.mock import patch, Mock, MagicMock
from openpyxl import Workbook
from src.core.multi_sheet_controller import MultiSheetTestController, SheetController


class TestMultiSheetController(unittest.TestCase):
    """Test cases for Multi Sheet Controller"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, 'multi_sheet_test.xlsx')
        self.create_test_excel_file()
        
    def tearDown(self):
        """Clean up test fixtures"""
        if os.path.exists(self.test_file):
            os.remove(self.test_file)
        if os.path.exists(self.temp_dir):
            os.rmdir(self.temp_dir)
        
    def create_test_excel_file(self):
        """Create a test Excel file with multiple sheets and CONTROLLER sheet"""
        workbook = Workbook()
        
        # Create CONTROLLER sheet
        controller_sheet = workbook.active
        controller_sheet.title = 'CONTROLLER'
        
        # Controller headers
        controller_headers = ['Enable', 'Sheet_Name', 'Description', 'Priority']
        for idx, header in enumerate(controller_headers, 1):
            controller_sheet.cell(row=1, column=idx, value=header)
        
        # Controller data
        controller_data = [
            ['TRUE', 'SMOKE', 'Smoke tests for database connectivity', 'HIGH'],
            ['TRUE', 'INTEGRATION', 'Integration tests for data flow', 'MEDIUM'],
            ['FALSE', 'PERFORMANCE', 'Performance tests for load testing', 'LOW']
        ]
        
        for row_idx, row_data in enumerate(controller_data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                controller_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Create SMOKE sheet
        smoke_sheet = workbook.create_sheet('SMOKE')
        smoke_headers = [
            'Enable', 'Test_Case_ID', 'Test_Case_Name', 'Application_Name',
            'Environment_Name', 'Priority', 'Test_Category', 'Expected_Result',
            'Timeout_Seconds', 'Description', 'Prerequisites', 'Tags', 'Parameters'
        ]
        
        for idx, header in enumerate(smoke_headers, 1):
            smoke_sheet.cell(row=1, column=idx, value=header)
        
        smoke_data = [
            ['TRUE', 'SMOKE_001', 'Connection Test', 'POSTGRES', 'DEV', 'HIGH',
             'CONNECTION', 'PASS', '30', 'Test connection', 'DB available', 'smoke,connection', ''],
            ['TRUE', 'SMOKE_002', 'Table Test', 'POSTGRES', 'DEV', 'MEDIUM',
             'TABLE_EXISTS', 'PASS', '30', 'Test table', 'DB available', 'smoke,table', 'table_name=test']
        ]
        
        for row_idx, row_data in enumerate(smoke_data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                smoke_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Create INTEGRATION sheet
        integration_sheet = workbook.create_sheet('INTEGRATION')
        for idx, header in enumerate(smoke_headers, 1):
            integration_sheet.cell(row=1, column=idx, value=header)
        
        integration_data = [
            ['TRUE', 'INT_001', 'Data Flow Test', 'POSTGRES', 'DEV', 'HIGH',
             'QUERIES', 'PASS', '60', 'Test data flow', 'DB available', 'integration,data', '']
        ]
        
        for row_idx, row_data in enumerate(integration_data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                integration_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Create PERFORMANCE sheet (will be disabled)
        performance_sheet = workbook.create_sheet('PERFORMANCE')
        for idx, header in enumerate(smoke_headers, 1):
            performance_sheet.cell(row=1, column=idx, value=header)
        
        performance_data = [
            ['TRUE', 'PERF_001', 'Load Test', 'POSTGRES', 'DEV', 'LOW',
             'PERFORMANCE', 'PASS', '300', 'Test performance', 'DB available', 'performance,load', '']
        ]
        
        for row_idx, row_data in enumerate(performance_data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                performance_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        workbook.save(self.test_file)

    @pytest.mark.positive
    @pytest.mark.multi_sheet
    @pytest.mark.controller
    def test_controller_initialization(self):
        """Test MultiSheetTestController initialization"""
        controller = MultiSheetTestController(self.test_file)
        self.assertIsNotNone(controller)
        self.assertEqual(str(controller.excel_file), self.test_file)
        self.assertIsNone(controller.workbook)  # Not loaded yet
        self.assertEqual(len(controller.sheet_controllers), 0)  # Not loaded yet

    @pytest.mark.positive
    @pytest.mark.multi_sheet
    @pytest.mark.excel_processing
    def test_load_workbook_success(self):
        """Test successful workbook loading"""
        controller = MultiSheetTestController(self.test_file)
        result = controller.load_workbook()
        
        self.assertTrue(result)
        self.assertIsNotNone(controller.workbook)
        self.assertIn('CONTROLLER', controller.workbook.sheetnames)
        self.assertIn('SMOKE', controller.workbook.sheetnames)

    @pytest.mark.negative
    @pytest.mark.multi_sheet
    @pytest.mark.edge_case
    def test_load_workbook_invalid_file(self):
        """Test workbook loading with invalid file"""
        invalid_file = os.path.join(self.temp_dir, 'nonexistent.xlsx')
        controller = MultiSheetTestController(invalid_file)
        result = controller.load_workbook()
        
        self.assertFalse(result)

    @pytest.mark.positive
    @pytest.mark.multi_sheet
    @pytest.mark.validation
    def test_validate_controller_sheet_success(self):
        """Test successful CONTROLLER sheet validation"""
        controller = MultiSheetTestController(self.test_file)
        controller.load_workbook()
        
        result = controller.validate_controller_sheet()
        self.assertTrue(result)

    @pytest.mark.negative
    @pytest.mark.multi_sheet
    @pytest.mark.validation
    def test_validate_controller_sheet_missing(self):
        """Test CONTROLLER sheet validation when sheet is missing"""
        # Create file without CONTROLLER sheet
        no_controller_file = os.path.join(self.temp_dir, 'no_controller.xlsx')
        workbook = Workbook()
        workbook.save(no_controller_file)
        
        controller = MultiSheetTestController(no_controller_file)
        controller.load_workbook()
        
        result = controller.validate_controller_sheet()
        self.assertFalse(result)
        
        # Clean up
        os.remove(no_controller_file)

    @pytest.mark.positive
    @pytest.mark.multi_sheet
    @pytest.mark.controller
    def test_load_controller_data_success(self):
        """Test successful controller data loading"""
        controller = MultiSheetTestController(self.test_file)
        controller.load_workbook()
        
        result = controller.load_controller_data()
        self.assertTrue(result)
        
        # Should have 3 sheet controllers
        self.assertEqual(len(controller.sheet_controllers), 3)
        
        # Check specific controllers
        sheet_names = [sc.sheet_name for sc in controller.sheet_controllers]
        self.assertIn('SMOKE', sheet_names)
        self.assertIn('INTEGRATION', sheet_names)
        self.assertIn('PERFORMANCE', sheet_names)

    @pytest.mark.positive
    @pytest.mark.multi_sheet
    @pytest.mark.controller
    def test_get_enabled_sheets(self):
        """Test getting enabled sheets"""
        controller = MultiSheetTestController(self.test_file)
        controller.load_workbook()
        controller.load_controller_data()
        
        enabled_sheets = controller.get_enabled_sheets()
        
        # Should have 2 enabled sheets (SMOKE and INTEGRATION)
        self.assertEqual(len(enabled_sheets), 2)
        
        enabled_names = [sheet.sheet_name for sheet in enabled_sheets]
        self.assertIn('SMOKE', enabled_names)
        self.assertIn('INTEGRATION', enabled_names)
        self.assertNotIn('PERFORMANCE', enabled_names)

    @pytest.mark.positive
    @pytest.mark.multi_sheet
    @pytest.mark.controller
    def test_get_disabled_sheets(self):
        """Test getting disabled sheets"""
        controller = MultiSheetTestController(self.test_file)
        controller.load_workbook()
        controller.load_controller_data()
        
        disabled_sheets = controller.get_disabled_sheets()
        
        # Should have 1 disabled sheet (PERFORMANCE)
        self.assertEqual(len(disabled_sheets), 1)
        self.assertEqual(disabled_sheets[0].sheet_name, 'PERFORMANCE')

    @pytest.mark.positive
    @pytest.mark.utility
    @pytest.mark.edge_case
    def test_convert_bool_method(self):
        """Test _convert_bool method"""
        controller = MultiSheetTestController(self.test_file)
        
        # Test various boolean representations
        self.assertTrue(controller._convert_bool(True))
        self.assertTrue(controller._convert_bool('TRUE'))
        self.assertTrue(controller._convert_bool('YES'))
        self.assertTrue(controller._convert_bool('Y'))
        self.assertTrue(controller._convert_bool('1'))
        self.assertTrue(controller._convert_bool(1))
        
        self.assertFalse(controller._convert_bool(False))
        self.assertFalse(controller._convert_bool('FALSE'))
        self.assertFalse(controller._convert_bool('NO'))
        self.assertFalse(controller._convert_bool('N'))
        self.assertFalse(controller._convert_bool('0'))
        self.assertFalse(controller._convert_bool(0))
        self.assertFalse(controller._convert_bool(''))
        self.assertFalse(controller._convert_bool(None))

    @pytest.mark.positive
    @pytest.mark.multi_sheet
    @pytest.mark.statistical
    def test_get_sheet_test_counts(self):
        """Test _get_sheet_test_counts method"""
        controller = MultiSheetTestController(self.test_file)
        controller.load_workbook()
        
        # Test SMOKE sheet (has 2 tests, both enabled)
        total, enabled = controller._get_sheet_test_counts('SMOKE')
        self.assertEqual(total, 2)
        self.assertEqual(enabled, 2)
        
        # Test INTEGRATION sheet (has 1 test, enabled)
        total, enabled = controller._get_sheet_test_counts('INTEGRATION')
        self.assertEqual(total, 1)
        self.assertEqual(enabled, 1)
        
        # Test non-existent sheet
        total, enabled = controller._get_sheet_test_counts('NONEXISTENT')
        self.assertEqual(total, 0)
        self.assertEqual(enabled, 0)

    @pytest.mark.positive
    @pytest.mark.dataclass
    @pytest.mark.data_structures
    def test_sheet_controller_dataclass(self):
        """Test SheetController dataclass"""
        sheet_controller = SheetController(
            enable=True,
            sheet_name='TEST_SHEET',
            description='Test description',
            priority='HIGH',
            test_count=5,
            enabled_test_count=3
        )
        
        self.assertTrue(sheet_controller.enable)
        self.assertEqual(sheet_controller.sheet_name, 'TEST_SHEET')
        self.assertEqual(sheet_controller.description, 'Test description')
        self.assertEqual(sheet_controller.priority, 'HIGH')
        self.assertEqual(sheet_controller.test_count, 5)
        self.assertEqual(sheet_controller.enabled_test_count, 3)

    @pytest.mark.positive
    @pytest.mark.multi_sheet
    @pytest.mark.output
    def test_print_controller_summary(self):
        """Test print_controller_summary method"""
        controller = MultiSheetTestController(self.test_file)
        controller.load_workbook()
        controller.load_controller_data()
        
        # Should not raise an exception
        try:
            controller.print_controller_summary()
        except Exception as e:
            self.fail(f"print_controller_summary raised an exception: {e}")

    @pytest.mark.positive
    @pytest.mark.multi_sheet
    @pytest.mark.functional
    @pytest.mark.integration
    def test_execute_controlled_tests(self):
        """Test execute_controlled_tests method"""
        controller = MultiSheetTestController(self.test_file)
        controller.load_workbook()
        controller.load_controller_data()
        
        # Mock the test execution to avoid actual database connections
        with patch('src.core.test_executor.TestExecutor') as mock_executor_class:
            mock_executor = mock_executor_class.return_value
            mock_executor.execute_test_case.return_value = Mock()
            
            # Execute tests
            results = controller.execute_controlled_tests()
            
            # Should return a dictionary of results
            self.assertIsInstance(results, dict)
            
            # Should have results for enabled sheets
            self.assertIn('SMOKE', results)
            self.assertIn('INTEGRATION', results)
            self.assertNotIn('PERFORMANCE', results)  # Disabled

    @pytest.mark.positive
    @pytest.mark.multi_sheet
    @pytest.mark.filtering
    def test_apply_filters_method(self):
        """Test _apply_filters method"""
        controller = MultiSheetTestController(self.test_file)
        controller.load_workbook()
        
        # Create mock test cases
        from src.utils.excel_test_suite_reader import TestCase
        test_cases = [
            TestCase(
                test_case_id='TEST_001',
                test_case_name='Test 1',
                application_name='POSTGRES',
                environment_name='DEV',
                priority='HIGH',
                test_category='CONNECTION',
                expected_result='PASS',
                timeout_seconds=30,
                description='Test 1',
                prerequisites='None',
                tags='smoke,high',
                parameters='',
                enable=True
            ),
            TestCase(
                test_case_id='TEST_002',
                test_case_name='Test 2',
                application_name='POSTGRES',
                environment_name='DEV',
                priority='MEDIUM',
                test_category='TABLE_EXISTS',
                expected_result='PASS',
                timeout_seconds=30,
                description='Test 2',
                prerequisites='None',
                tags='smoke,medium',
                parameters='',
                enable=True
            )
        ]
        
        # Test filtering by priority
        filtered = controller._apply_filters(test_cases, priority='HIGH')
        self.assertEqual(len(filtered), 1)
        self.assertEqual(filtered[0].priority, 'HIGH')
        
        # Test filtering by category
        filtered = controller._apply_filters(test_cases, category='CONNECTION')
        self.assertEqual(len(filtered), 1)
        self.assertEqual(filtered[0].test_category, 'CONNECTION')

    @pytest.mark.negative
    @pytest.mark.multi_sheet
    @pytest.mark.validation
    @pytest.mark.edge_case
    def test_error_handling_invalid_controller_sheet(self):
        """Test error handling with invalid CONTROLLER sheet structure"""
        # Create file with wrong CONTROLLER headers
        invalid_file = os.path.join(self.temp_dir, 'invalid_controller.xlsx')
        workbook = Workbook()
        controller_sheet = workbook.active
        controller_sheet.title = 'CONTROLLER'
        
        # Wrong headers
        wrong_headers = ['Wrong', 'Headers', 'Here']
        for idx, header in enumerate(wrong_headers, 1):
            controller_sheet.cell(row=1, column=idx, value=header)
        
        workbook.save(invalid_file)
        
        controller = MultiSheetTestController(invalid_file)
        controller.load_workbook()
        
        result = controller.validate_controller_sheet()
        self.assertFalse(result)
        
        # Clean up
        os.remove(invalid_file)

    @pytest.mark.negative
    @pytest.mark.multi_sheet
    @pytest.mark.edge_case
    def test_missing_referenced_sheets(self):
        """Test handling of missing referenced sheets"""
        # Create file where CONTROLLER references non-existent sheets
        missing_sheet_file = os.path.join(self.temp_dir, 'missing_sheets.xlsx')
        workbook = Workbook()
        controller_sheet = workbook.active
        controller_sheet.title = 'CONTROLLER'
        
        # Controller headers
        controller_headers = ['Enable', 'Sheet_Name', 'Description', 'Priority']
        for idx, header in enumerate(controller_headers, 1):
            controller_sheet.cell(row=1, column=idx, value=header)
        
        # Reference non-existent sheet
        controller_data = [
            ['TRUE', 'NONEXISTENT_SHEET', 'Missing sheet', 'HIGH']
        ]
        
        for row_idx, row_data in enumerate(controller_data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                controller_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        workbook.save(missing_sheet_file)
        
        controller = MultiSheetTestController(missing_sheet_file)
        controller.load_workbook()
        
        # Should handle gracefully and still load controller data
        result = controller.load_controller_data()
        self.assertTrue(result)  # Should succeed but warn about missing sheet
        
        # Clean up
        os.remove(missing_sheet_file)

    @pytest.mark.edge_case
    @pytest.mark.multi_sheet
    @pytest.mark.validation
    def test_empty_controller_sheet(self):
        """Test handling of empty CONTROLLER sheet"""
        empty_file = os.path.join(self.temp_dir, 'empty_controller.xlsx')
        workbook = Workbook()
        controller_sheet = workbook.active
        controller_sheet.title = 'CONTROLLER'
        
        # Only headers, no data
        controller_headers = ['Enable', 'Sheet_Name', 'Description', 'Priority']
        for idx, header in enumerate(controller_headers, 1):
            controller_sheet.cell(row=1, column=idx, value=header)
        
        workbook.save(empty_file)
        
        controller = MultiSheetTestController(empty_file)
        controller.load_workbook()
        
        result = controller.load_controller_data()
        self.assertTrue(result)
        self.assertEqual(len(controller.sheet_controllers), 0)
        
        # Clean up
        os.remove(empty_file)


if __name__ == '__main__':
    unittest.main()