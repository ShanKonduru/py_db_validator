#!/usr/bin/env python
"""
Comprehensive unit tests for Excel validator
"""
import unittest
import tempfile
import os
from unittest.mock import patch, Mock
from openpyxl import Workbook
from src.validation.excel_validator import ExcelTestSuiteValidator


class TestExcelValidator(unittest.TestCase):
    """Test cases for Excel validator"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.temp_dir = tempfile.mkdtemp()
        self.valid_file = os.path.join(self.temp_dir, 'valid_test.xlsx')
        self.invalid_file = os.path.join(self.temp_dir, 'invalid_test.xlsx')
        self.create_test_files()
        
    def tearDown(self):
        """Clean up test fixtures"""
        for file_path in [self.valid_file, self.invalid_file]:
            if os.path.exists(file_path):
                os.remove(file_path)
        os.rmdir(self.temp_dir)
        
    def create_test_files(self):
        """Create test Excel files"""
        self.create_valid_excel_file()
        self.create_invalid_excel_file()
        
    def create_valid_excel_file(self):
        """Create a valid Excel test file"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'SMOKE'
        
        # Valid headers
        headers = [
            'Enable', 'Test Case ID', 'Test Case Name', 'Application Name',
            'Environment Name', 'Priority', 'Test Category', 'Expected Result',
            'Timeout (seconds)', 'Description', 'Prerequisites', 'Tags', 'Parameters'
        ]
        
        for idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=idx, value=header)
        
        # Valid test data
        test_data = [
            ['TRUE', 'SMOKE_001', 'Valid Connection Test', 'TestApp', 'DEV', 'HIGH',
             'CONNECTION_TEST', 'PASS', '30', 'Valid test description', 'Database available',
             'smoke,connection', ''],
            ['TRUE', 'SMOKE_002', 'Valid Table Test', 'TestApp', 'DEV', 'MEDIUM',
             'TABLE_EXISTS', 'PASS', '30', 'Valid table test', 'Database available',
             'smoke,table', 'table_name=public.products'],
            ['FALSE', 'SMOKE_003', 'Disabled Test', 'TestApp', 'DEV', 'LOW',
             'SMOKE_TEST', 'PASS', '30', 'Disabled test', 'None',
             'smoke,disabled', '']
        ]
        
        for row_idx, row_data in enumerate(test_data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        workbook.save(self.valid_file)
        
    def create_invalid_excel_file(self):
        """Create an invalid Excel test file"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'SMOKE'
        
        # Missing some required headers
        headers = [
            'Enable', 'Test Case ID', 'Invalid Name', 'Application Name',
            'Environment Name', 'Priority'
            # Missing several required headers
        ]
        
        for idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=idx, value=header)
        
        # Invalid test data
        test_data = [
            ['INVALID', 'INVALID_ID', '', 'TestApp', 'DEV', 'INVALID_PRIORITY'],
            ['TRUE', '', 'Test without ID', 'TestApp', 'DEV', 'HIGH']
        ]
        
        for row_idx, row_data in enumerate(test_data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        workbook.save(self.invalid_file)

    def test_validator_initialization(self):
        """Test ExcelTestSuiteValidator initialization"""
        validator = ExcelTestSuiteValidator()
        self.assertIsNotNone(validator)

    def test_validate_file_path(self):
        """Test file path validation"""
        validator = ExcelTestSuiteValidator()
        
        # Valid file
        result = validator.validate_file_path(self.valid_file)
        self.assertTrue(result)
        
        # Non-existent file
        result = validator.validate_file_path('nonexistent.xlsx')
        self.assertFalse(result)
        
        # None file
        result = validator.validate_file_path(None)
        self.assertFalse(result)

    def test_validate_excel_structure(self):
        """Test Excel structure validation"""
        validator = ExcelTestSuiteValidator()
        
        # Valid file
        result = validator.validate_excel_structure(self.valid_file)
        self.assertTrue(result)
        
        # Invalid file
        result = validator.validate_excel_structure(self.invalid_file)
        self.assertFalse(result)

    def test_validate_smoke_sheet_exists(self):
        """Test SMOKE sheet existence validation"""
        validator = ExcelTestSuiteValidator()
        
        # Valid file with SMOKE sheet
        result = validator.validate_smoke_sheet_exists(self.valid_file)
        self.assertTrue(result)
        
        # Create file without SMOKE sheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'OTHER'
        
        no_smoke_file = os.path.join(self.temp_dir, 'no_smoke.xlsx')
        workbook.save(no_smoke_file)
        
        result = validator.validate_smoke_sheet_exists(no_smoke_file)
        self.assertFalse(result)
        
        # Clean up
        os.remove(no_smoke_file)

    def test_validate_headers(self):
        """Test header validation"""
        validator = ExcelTestSuiteValidator()
        
        # Valid headers
        valid_headers = [
            'Enable', 'Test Case ID', 'Test Case Name', 'Application Name',
            'Environment Name', 'Priority', 'Test Category', 'Expected Result',
            'Timeout (seconds)', 'Description', 'Prerequisites', 'Tags', 'Parameters'
        ]
        
        result = validator.validate_headers(valid_headers)
        self.assertTrue(result)
        
        # Invalid headers (missing required)
        invalid_headers = ['Enable', 'Test Case ID', 'Wrong Header']
        result = validator.validate_headers(invalid_headers)
        self.assertFalse(result)
        
        # Empty headers
        result = validator.validate_headers([])
        self.assertFalse(result)

    def test_validate_test_data(self):
        """Test test data validation"""
        validator = ExcelTestSuiteValidator()
        
        # Valid test row
        valid_row = [
            'TRUE', 'TEST_001', 'Valid Test', 'App', 'DEV', 'HIGH',
            'CONNECTION_TEST', 'PASS', '30', 'Description', 'Prerequisites',
            'tags', 'parameters'
        ]
        
        result = validator.validate_test_data_row(valid_row, 2)
        self.assertTrue(result)
        
        # Invalid test row (missing required fields)
        invalid_row = ['TRUE', '', 'Test Name', 'App', 'DEV', 'INVALID_PRIORITY']
        result = validator.validate_test_data_row(invalid_row, 2)
        self.assertFalse(result)

    def test_validate_enable_values(self):
        """Test Enable column validation"""
        validator = ExcelTestSuiteValidator()
        
        # Valid values
        self.assertTrue(validator.is_valid_enable_value('TRUE'))
        self.assertTrue(validator.is_valid_enable_value('FALSE'))
        
        # Invalid values
        self.assertFalse(validator.is_valid_enable_value('INVALID'))
        self.assertFalse(validator.is_valid_enable_value(''))
        self.assertFalse(validator.is_valid_enable_value(None))

    def test_validate_priority_values(self):
        """Test Priority column validation"""
        validator = ExcelTestSuiteValidator()
        
        # Valid values
        self.assertTrue(validator.is_valid_priority('HIGH'))
        self.assertTrue(validator.is_valid_priority('MEDIUM'))
        self.assertTrue(validator.is_valid_priority('LOW'))
        
        # Invalid values
        self.assertFalse(validator.is_valid_priority('INVALID'))
        self.assertFalse(validator.is_valid_priority(''))
        self.assertFalse(validator.is_valid_priority(None))

    def test_validate_test_category_values(self):
        """Test Test Category column validation"""
        validator = ExcelTestSuiteValidator()
        
        # Valid values
        self.assertTrue(validator.is_valid_test_category('CONNECTION_TEST'))
        self.assertTrue(validator.is_valid_test_category('SMOKE_TEST'))
        self.assertTrue(validator.is_valid_test_category('TABLE_EXISTS'))
        self.assertTrue(validator.is_valid_test_category('TABLE_SELECT'))
        self.assertTrue(validator.is_valid_test_category('TABLE_ROWS'))
        self.assertTrue(validator.is_valid_test_category('TABLE_STRUCTURE'))
        
        # Invalid values
        self.assertFalse(validator.is_valid_test_category('INVALID_CATEGORY'))
        self.assertFalse(validator.is_valid_test_category(''))
        self.assertFalse(validator.is_valid_test_category(None))

    def test_validate_expected_result_values(self):
        """Test Expected Result column validation"""
        validator = ExcelTestSuiteValidator()
        
        # Valid values
        self.assertTrue(validator.is_valid_expected_result('PASS'))
        self.assertTrue(validator.is_valid_expected_result('FAIL'))
        
        # Invalid values
        self.assertFalse(validator.is_valid_expected_result('INVALID'))
        self.assertFalse(validator.is_valid_expected_result(''))
        self.assertFalse(validator.is_valid_expected_result(None))

    def test_validate_timeout_values(self):
        """Test Timeout column validation"""
        validator = ExcelTestSuiteValidator()
        
        # Valid values
        self.assertTrue(validator.is_valid_timeout('30'))
        self.assertTrue(validator.is_valid_timeout('60'))
        self.assertTrue(validator.is_valid_timeout(30))
        
        # Invalid values
        self.assertFalse(validator.is_valid_timeout('invalid'))
        self.assertFalse(validator.is_valid_timeout('-10'))
        self.assertFalse(validator.is_valid_timeout(''))
        self.assertFalse(validator.is_valid_timeout(None))

    def test_validate_required_fields(self):
        """Test required field validation"""
        validator = ExcelTestSuiteValidator()
        
        # Valid required field
        self.assertTrue(validator.is_required_field_valid('Required Value'))
        
        # Invalid required fields
        self.assertFalse(validator.is_required_field_valid(''))
        self.assertFalse(validator.is_required_field_valid(None))
        self.assertFalse(validator.is_required_field_valid('   '))  # Only whitespace

    def test_get_validation_errors(self):
        """Test getting validation errors"""
        validator = ExcelTestSuiteValidator()
        
        # Validate invalid file and check errors
        result = validator.validate_excel_structure(self.invalid_file)
        self.assertFalse(result)
        
        errors = validator.get_validation_errors()
        self.assertIsInstance(errors, list)
        self.assertGreater(len(errors), 0)

    def test_clear_validation_errors(self):
        """Test clearing validation errors"""
        validator = ExcelTestSuiteValidator()
        
        # Generate some errors
        validator.validate_excel_structure(self.invalid_file)
        errors = validator.get_validation_errors()
        self.assertGreater(len(errors), 0)
        
        # Clear errors
        validator.clear_validation_errors()
        errors = validator.get_validation_errors()
        self.assertEqual(len(errors), 0)

    def test_validate_complete_file(self):
        """Test complete file validation"""
        validator = ExcelTestSuiteValidator()
        
        # Valid file
        result = validator.validate_complete_file(self.valid_file)
        self.assertTrue(result)
        
        # Invalid file
        result = validator.validate_complete_file(self.invalid_file)
        self.assertFalse(result)

    def test_get_validation_summary(self):
        """Test getting validation summary"""
        validator = ExcelTestSuiteValidator()
        
        # Validate a file
        validator.validate_complete_file(self.invalid_file)
        summary = validator.get_validation_summary()
        
        self.assertIsInstance(summary, str)
        self.assertIn('validation', summary.lower())

    def test_validate_parameters_column(self):
        """Test Parameters column validation"""
        validator = ExcelTestSuiteValidator()
        
        # Valid parameter formats
        valid_params = [
            '',  # Empty is valid
            'table_name=public.products',  # Single parameter
            'table_name=public.employees,min_rows=5',  # Multiple parameters
            'simple_table_name'  # Simple value
        ]
        
        for param in valid_params:
            result = validator.is_valid_parameters(param)
            self.assertTrue(result, f"Failed for parameter: {param}")

    def test_edge_cases(self):
        """Test edge cases and error conditions"""
        validator = ExcelTestSuiteValidator()
        
        # Empty file path
        result = validator.validate_file_path('')
        self.assertFalse(result)
        
        # None as input
        result = validator.validate_headers(None)
        self.assertFalse(result)
        
        # Very long test case ID
        long_id = 'A' * 1000
        result = validator.is_required_field_valid(long_id)
        self.assertTrue(result)  # Should still be valid

    def test_case_sensitivity(self):
        """Test case sensitivity in validation"""
        validator = ExcelTestSuiteValidator()
        
        # Test case sensitivity for enable values
        self.assertTrue(validator.is_valid_enable_value('TRUE'))
        self.assertFalse(validator.is_valid_enable_value('true'))  # Should be case-sensitive
        
        # Test case sensitivity for priority
        self.assertTrue(validator.is_valid_priority('HIGH'))
        self.assertFalse(validator.is_valid_priority('high'))

    def test_whitespace_handling(self):
        """Test whitespace handling in validation"""
        validator = ExcelTestSuiteValidator()
        
        # Test whitespace trimming in required fields
        self.assertTrue(validator.is_required_field_valid('  Valid Value  '))
        self.assertFalse(validator.is_required_field_valid('   '))  # Only whitespace

    def test_numeric_validation(self):
        """Test numeric validation for timeout"""
        validator = ExcelTestSuiteValidator()
        
        # Valid numeric values
        self.assertTrue(validator.is_valid_timeout(30))
        self.assertTrue(validator.is_valid_timeout('30'))
        self.assertTrue(validator.is_valid_timeout('0'))
        
        # Invalid numeric values
        self.assertFalse(validator.is_valid_timeout(-1))
        self.assertFalse(validator.is_valid_timeout('-1'))
        self.assertFalse(validator.is_valid_timeout('abc'))

    def test_file_format_validation(self):
        """Test file format validation"""
        validator = ExcelTestSuiteValidator()
        
        # Create a non-Excel file
        text_file = os.path.join(self.temp_dir, 'not_excel.txt')
        with open(text_file, 'w') as f:
            f.write('This is not an Excel file')
        
        result = validator.validate_file_path(text_file)
        # Should still return True for file existence, but Excel validation should fail
        self.assertTrue(result)  # File exists
        
        result = validator.validate_excel_structure(text_file)
        self.assertFalse(result)  # Not a valid Excel file
        
        # Clean up
        os.remove(text_file)


if __name__ == '__main__':
    unittest.main()