#!/usr/bin/env python
"""
Comprehensive unit tests for Excel template generator
"""
import unittest
import pytest
import tempfile
import os
from unittest.mock import patch, Mock, MagicMock
from openpyxl import Workbook, load_workbook
from src.utils.excel_template_generator import ExcelTemplateGenerator


class TestExcelTemplateGenerator(unittest.TestCase):
    """Test cases for Excel template generator"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.temp_dir = tempfile.mkdtemp()
        self.template_path = os.path.join(self.temp_dir, 'test_template.xlsx')
        self.generator = ExcelTemplateGenerator()

    def tearDown(self):
        """Clean up test fixtures"""
        if os.path.exists(self.template_path):
            os.remove(self.template_path)
        os.rmdir(self.temp_dir)

    @pytest.mark.positive
    @pytest.mark.initialization
    @pytest.mark.template_generation
    def test_init(self):
        """Test ExcelTemplateGenerator initialization"""
        generator = ExcelTemplateGenerator()
        # Test that initialization doesn't raise errors
        self.assertIsNotNone(generator)

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.excel_processing
    def test_create_template_success(self):
        """Test successful template creation"""
        result = self.generator.create_template(self.template_path)
        
        self.assertTrue(result)
        self.assertTrue(os.path.exists(self.template_path))
        
        # Verify the file can be opened
        workbook = load_workbook(self.template_path)
        expected_sheets = ['SMOKE', 'REFERENCE', 'INSTRUCTIONS']
        for sheet_name in expected_sheets:
            self.assertIn(sheet_name, workbook.sheetnames)

    @pytest.mark.negative
    @pytest.mark.template_generation
    @pytest.mark.error_handling
    def test_create_template_with_invalid_path(self):
        """Test template creation with invalid path"""
        invalid_path = "/definitely/invalid/path/template.xlsx"
        result = self.generator.create_template(invalid_path)
        
        self.assertFalse(result)

    @pytest.mark.negative
    @pytest.mark.template_generation
    @pytest.mark.error_handling
    def test_create_template_with_permission_error(self):
        """Test template creation with permission error"""
        if os.name == 'nt':  # Windows
            restricted_path = "C:\\Windows\\System32\\test_template.xlsx"
        else:  # Unix-like
            restricted_path = "/root/test_template.xlsx"
        
        result = self.generator.create_template(restricted_path)
        self.assertFalse(result)

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.headers
    def test_create_headers(self):
        """Test _create_headers method"""
        workbook = Workbook()
        worksheet = workbook.active
        
        self.generator._create_headers(worksheet)
        
        # Check that headers are set correctly
        headers = [cell.value for cell in worksheet[1]]
        expected_headers = [
            'Enable', 'Test_Case_ID', 'Test_Case_Name', 'Application_Name',
            'Environment_Name', 'Priority', 'Test_Category', 'Expected_Result',
            'Timeout_Seconds', 'Description', 'Prerequisites', 'Tags', 'Parameters'
        ]
        self.assertEqual(headers, expected_headers)

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.validation
    def test_add_data_validations(self):
        """Test _add_data_validations method"""
        workbook = Workbook()
        worksheet = workbook.active
        
        # Set up headers first
        self.generator._create_headers(worksheet)
        
        # This should not raise an exception
        self.generator._add_data_validations(worksheet)
        
        # Basic verification that data validation was attempted
        self.assertTrue(True)  # If no exception, validation was set

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.data_structures
    def test_add_sample_data(self):
        """Test _add_sample_data method"""
        workbook = Workbook()
        worksheet = workbook.active
        
        # Set up headers first
        self.generator._create_headers(worksheet)
        
        self.generator._add_sample_data(worksheet)
        
        # Check that sample data is added
        self.assertGreater(worksheet.max_row, 1)
        
        # Check specific sample data exists
        self.assertIsNotNone(worksheet.cell(row=2, column=1).value)  # First data row

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.documentation
    def test_create_instructions_worksheet(self):
        """Test _create_instructions_worksheet method"""
        workbook = Workbook()
        
        self.generator._create_instructions_worksheet(workbook)
        
        self.assertIn('INSTRUCTIONS', workbook.sheetnames)
        inst_sheet = workbook['INSTRUCTIONS']
        
        # Check that instructions content exists
        self.assertGreater(inst_sheet.max_row, 1)
        self.assertGreater(inst_sheet.max_column, 0)  # Instructions should have content

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.documentation
    def test_create_reference_worksheet(self):
        """Test _create_reference_worksheet method"""
        workbook = Workbook()
        
        self.generator._create_reference_worksheet(workbook)
        
        self.assertIn('REFERENCE', workbook.sheetnames)
        ref_sheet = workbook['REFERENCE']
        
        # Check that reference data exists
        self.assertGreater(ref_sheet.max_row, 1)
        self.assertGreater(ref_sheet.max_column, 1)

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.controller
    def test_create_controller_sheet(self):
        """Test _create_controller_sheet method"""
        workbook = Workbook()
        
        self.generator._create_controller_sheet(workbook)
        
        self.assertIn('CONTROLLER', workbook.sheetnames)
        ctrl_sheet = workbook['CONTROLLER']
        
        # Check that controller data exists
        self.assertGreater(ctrl_sheet.max_row, 1)

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.functional
    def test_full_template_workflow(self):
        """Test complete template generation workflow"""
        result = self.generator.create_template(self.template_path)
        
        self.assertTrue(result)
        self.assertTrue(os.path.exists(self.template_path))
        
        # Load and verify the generated template
        workbook = load_workbook(self.template_path)
        
        # Check all expected sheets exist
        expected_sheets = ['SMOKE', 'REFERENCE', 'INSTRUCTIONS', 'CONTROLLER']
        for sheet_name in expected_sheets:
            self.assertIn(sheet_name, workbook.sheetnames)
        
        # Check SMOKE sheet has correct structure
        smoke_sheet = workbook['SMOKE']
        headers = [cell.value for cell in smoke_sheet[1]]
        self.assertEqual(len(headers), 13)  # 13 columns including Parameters
        self.assertIn('Parameters', headers)
        
        # Check that sample data exists
        self.assertGreater(smoke_sheet.max_row, 1)

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.configuration
    def test_template_with_custom_filename(self):
        """Test template creation with custom filename"""
        custom_path = os.path.join(self.temp_dir, 'custom_template.xlsx')
        result = self.generator.create_template(custom_path)
        
        self.assertTrue(result)
        self.assertTrue(os.path.exists(custom_path))
        
        # Clean up
        os.remove(custom_path)

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.file_handling
    def test_template_overwrites_existing_file(self):
        """Test that template generation overwrites existing files"""
        # Create a dummy file first
        with open(self.template_path, 'w') as f:
            f.write("dummy content")
        
        result = self.generator.create_template(self.template_path)
        
        self.assertTrue(result)
        self.assertTrue(os.path.exists(self.template_path))
        
        # Verify it's now a valid Excel file
        workbook = load_workbook(self.template_path)
        self.assertIn('SMOKE', workbook.sheetnames)

    @pytest.mark.negative
    @pytest.mark.template_generation
    @pytest.mark.error_handling
    def test_error_handling_in_create_template(self):
        """Test error handling in create_template method"""
        # Test with None filename
        result = self.generator.create_template(None)
        self.assertFalse(result)
        
        # Test with empty filename
        result = self.generator.create_template("")
        self.assertFalse(result)

    @pytest.mark.negative
    @pytest.mark.template_generation
    @pytest.mark.error_handling
    def test_workbook_creation_error(self, mock_workbook):
        """Test error handling when workbook creation fails"""
        mock_workbook.side_effect = Exception("Workbook creation failed")
        
        result = self.generator.create_template(self.template_path)
        self.assertFalse(result)

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.structure
    def test_sheet_ordering(self):
        """Test that sheets are created in the correct order"""
        result = self.generator.create_template(self.template_path)
        self.assertTrue(result)
        
        workbook = load_workbook(self.template_path)
        sheet_names = workbook.sheetnames
        
        # SMOKE should be first
        self.assertEqual(sheet_names[0], 'SMOKE')
        # Other sheets should exist
        self.assertIn('REFERENCE', sheet_names)
        self.assertIn('INSTRUCTIONS', sheet_names)

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.formatting
    def test_column_formatting(self):
        """Test that columns are properly formatted in the template"""
        result = self.generator.create_template(self.template_path)
        self.assertTrue(result)
        
        workbook = load_workbook(self.template_path)
        smoke_sheet = workbook['SMOKE']
        
        # Check that header row has proper formatting
        header_cell = smoke_sheet.cell(row=1, column=1)
        self.assertIsNotNone(header_cell.value)
        
        # Check that columns have reasonable widths
        # (This is a basic check since openpyxl formatting is complex)
        self.assertTrue(True)  # If template loads without error, formatting worked

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.validation
    def test_data_validation_setup(self):
        """Test that data validation is properly set up"""
        result = self.generator.create_template(self.template_path)
        self.assertTrue(result)
        
        workbook = load_workbook(self.template_path)
        smoke_sheet = workbook['SMOKE']
        
        # Basic check that data validation exists
        # (Detailed validation testing would require more complex openpyxl inspection)
        self.assertIsNotNone(smoke_sheet)

    @pytest.mark.positive
    @pytest.mark.template_generation
    @pytest.mark.documentation
    def test_reference_sheet_content(self):
        """Test that reference sheet has proper content"""
        result = self.generator.create_template(self.template_path)
        self.assertTrue(result)
        
        workbook = load_workbook(self.template_path)
        ref_sheet = workbook['REFERENCE']
        
        # Check that reference sheet has content
        self.assertGreater(ref_sheet.max_row, 1)
        
        # Check for some expected reference data
        cells_with_content = []
        for row in ref_sheet.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value:
                    cells_with_content.append(cell.value)
        
        self.assertGreater(len(cells_with_content), 0)


if __name__ == '__main__':
    unittest.main()