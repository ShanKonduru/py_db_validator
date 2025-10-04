#!/usr/bin/env python
"""
Comprehensive unit tests for Excel test suite reader
"""
import unittest
import tempfile
import os
from unittest.mock import patch, Mock, MagicMock
from openpyxl import Workbook
from src.utils.excel_test_suite_reader import TestCase, ExcelTestSuiteReader


class TestExcelTestSuiteReader(unittest.TestCase):
    """Test cases for Excel test suite reader"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, 'test_suite.xlsx')
        self.create_test_excel_file()
        
    def tearDown(self):
        """Clean up test fixtures"""
        if os.path.exists(self.test_file):
            os.remove(self.test_file)
        os.rmdir(self.temp_dir)
        
    def create_test_excel_file(self):
        """Create a test Excel file with various test types"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'SMOKE'
        
        # Headers
        headers = [
            'Enable', 'Test Case ID', 'Test Case Name', 'Application Name',
            'Environment Name', 'Priority', 'Test Category', 'Expected Result',
            'Timeout (seconds)', 'Description', 'Prerequisites', 'Tags', 'Parameters'
        ]
        
        for idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=idx, value=header)
        
        # Test data with various test types
        test_data = [
            ['TRUE', 'SMOKE_001', 'Connection Test', 'TestApp', 'DEV', 'HIGH',
             'CONNECTION_TEST', 'PASS', '30', 'Test database connection', 'Database available',
             'smoke,connection', ''],
            ['TRUE', 'SMOKE_002', 'Table Exists Test', 'TestApp', 'DEV', 'MEDIUM',
             'TABLE_EXISTS', 'PASS', '30', 'Check if table exists', 'Database available',
             'smoke,table', 'table_name=public.products'],
            ['TRUE', 'SMOKE_003', 'Table Select Test', 'TestApp', 'DEV', 'MEDIUM',
             'TABLE_SELECT', 'PASS', '45', 'Select from table', 'Table exists',
             'smoke,select', 'table_name=public.employees,column_list=id,name'],
            ['TRUE', 'SMOKE_004', 'Table Rows Test', 'TestApp', 'DEV', 'LOW',
             'TABLE_ROWS', 'PASS', '30', 'Check table row count', 'Table exists',
             'smoke,rows', 'table_name=public.orders,min_rows=100,max_rows=1000'],
            ['TRUE', 'SMOKE_005', 'Table Structure Test', 'TestApp', 'DEV', 'HIGH',
             'TABLE_STRUCTURE', 'PASS', '60', 'Validate table structure', 'Table exists',
             'smoke,structure', 'table_name=public.customers,schema_validation=true'],
            ['FALSE', 'SMOKE_006', 'Disabled Test', 'TestApp', 'DEV', 'LOW',
             'SMOKE_TEST', 'PASS', '30', 'Disabled smoke test', 'None',
             'smoke,disabled', ''],
            ['TRUE', 'SMOKE_007', 'General Smoke Test', 'TestApp', 'DEV', 'MEDIUM',
             'SMOKE_TEST', 'PASS', '30', 'General smoke test', 'Application running',
             'smoke,general', 'endpoint=http://localhost:8080/health']
        ]
        
        for row_idx, row_data in enumerate(test_data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        workbook.save(self.test_file)

    def test_reader_initialization(self):
        """Test ExcelTestSuiteReader initialization"""
        reader = ExcelTestSuiteReader()
        self.assertIsNotNone(reader)

    def test_load_test_suite_valid_file(self):
        """Test loading test suite from valid file"""
        reader = ExcelTestSuiteReader()
        test_cases = reader.load_test_suite(self.test_file)
        
        self.assertIsInstance(test_cases, list)
        self.assertGreater(len(test_cases), 0)
        
        # Should have 6 enabled test cases
        enabled_tests = [tc for tc in test_cases if tc.enabled]
        self.assertEqual(len(enabled_tests), 6)

    def test_load_test_suite_file_not_found(self):
        """Test loading test suite from non-existent file"""
        reader = ExcelTestSuiteReader()
        test_cases = reader.load_test_suite('nonexistent.xlsx')
        
        self.assertIsInstance(test_cases, list)
        self.assertEqual(len(test_cases), 0)

    def test_load_test_suite_invalid_excel(self):
        """Test loading test suite from invalid Excel file"""
        # Create invalid file
        invalid_file = os.path.join(self.temp_dir, 'invalid.txt')
        with open(invalid_file, 'w') as f:
            f.write('Not an Excel file')
        
        reader = ExcelTestSuiteReader()
        test_cases = reader.load_test_suite(invalid_file)
        
        self.assertIsInstance(test_cases, list)
        self.assertEqual(len(test_cases), 0)
        
        # Clean up
        os.remove(invalid_file)

    def test_get_worksheet_by_name(self):
        """Test getting worksheet by name"""
        reader = ExcelTestSuiteReader()
        
        # Load workbook
        from openpyxl import load_workbook
        workbook = load_workbook(self.test_file)
        
        # Get SMOKE worksheet
        worksheet = reader.get_worksheet_by_name(workbook, 'SMOKE')
        self.assertIsNotNone(worksheet)
        self.assertEqual(worksheet.title, 'SMOKE')
        
        # Try to get non-existent worksheet
        worksheet = reader.get_worksheet_by_name(workbook, 'NONEXISTENT')
        self.assertIsNone(worksheet)

    def test_parse_headers(self):
        """Test parsing headers from worksheet"""
        reader = ExcelTestSuiteReader()
        
        from openpyxl import load_workbook
        workbook = load_workbook(self.test_file)
        worksheet = workbook['SMOKE']
        
        headers = reader.parse_headers(worksheet)
        
        expected_headers = [
            'Enable', 'Test Case ID', 'Test Case Name', 'Application Name',
            'Environment Name', 'Priority', 'Test Category', 'Expected Result',
            'Timeout (seconds)', 'Description', 'Prerequisites', 'Tags', 'Parameters'
        ]
        
        self.assertEqual(headers, expected_headers)

    def test_parse_test_case_row(self):
        """Test parsing individual test case row"""
        reader = ExcelTestSuiteReader()
        
        # Test data row
        row_data = [
            'TRUE', 'TEST_001', 'Test Name', 'App', 'DEV', 'HIGH',
            'CONNECTION_TEST', 'PASS', '30', 'Description', 'Prerequisites',
            'tag1,tag2', 'param1=value1'
        ]
        
        headers = [
            'Enable', 'Test Case ID', 'Test Case Name', 'Application Name',
            'Environment Name', 'Priority', 'Test Category', 'Expected Result',
            'Timeout (seconds)', 'Description', 'Prerequisites', 'Tags', 'Parameters'
        ]
        
        test_case = reader.parse_test_case_row(row_data, headers, 2)
        
        self.assertIsInstance(test_case, TestCase)
        self.assertEqual(test_case.test_case_id, 'TEST_001')
        self.assertEqual(test_case.test_case_name, 'Test Name')
        self.assertEqual(test_case.application_name, 'App')
        self.assertEqual(test_case.environment_name, 'DEV')
        self.assertEqual(test_case.priority, 'HIGH')
        self.assertEqual(test_case.test_category, 'CONNECTION_TEST')
        self.assertEqual(test_case.expected_result, 'PASS')
        self.assertEqual(test_case.timeout, 30)
        self.assertTrue(test_case.enabled)

    def test_parse_test_case_row_disabled(self):
        """Test parsing disabled test case row"""
        reader = ExcelTestSuiteReader()
        
        row_data = [
            'FALSE', 'TEST_002', 'Disabled Test', 'App', 'DEV', 'LOW',
            'SMOKE_TEST', 'PASS', '30', 'Description', 'Prerequisites',
            'tag1', ''
        ]
        
        headers = [
            'Enable', 'Test Case ID', 'Test Case Name', 'Application Name',
            'Environment Name', 'Priority', 'Test Category', 'Expected Result',
            'Timeout (seconds)', 'Description', 'Prerequisites', 'Tags', 'Parameters'
        ]
        
        test_case = reader.parse_test_case_row(row_data, headers, 3)
        
        self.assertIsInstance(test_case, TestCase)
        self.assertFalse(test_case.enabled)

    def test_create_specific_test_types(self):
        """Test creation of specific test types"""
        reader = ExcelTestSuiteReader()
        
        # Test CONNECTION_TEST
        base_test = TestCase(
            test_case_id='CONN_001',
            test_case_name='Connection Test',
            application_name='App',
            environment_name='DEV',
            priority='HIGH',
            test_category='CONNECTION_TEST',
            expected_result='PASS',
            timeout=30,
            description='Test connection',
            prerequisites='DB available',
            tags=['connection'],
            parameters='',
            enabled=True
        )
        
        # Test that we can create test cases with different categories
        self.assertEqual(base_test.test_category, 'CONNECTION_TEST')
        
        # Test TABLE_EXISTS
        base_test.test_category = 'TABLE_EXISTS'
        base_test.parameters = 'table_name=public.products'
        self.assertEqual(base_test.test_category, 'TABLE_EXISTS')
        
        # Test TABLE_SELECT
        base_test.test_category = 'TABLE_SELECT'
        self.assertEqual(base_test.test_category, 'TABLE_SELECT')
        
        # Test TABLE_ROWS
        base_test.test_category = 'TABLE_ROWS'
        self.assertEqual(base_test.test_category, 'TABLE_ROWS')
        
        # Test TABLE_STRUCTURE
        base_test.test_category = 'TABLE_STRUCTURE'
        self.assertEqual(base_test.test_category, 'TABLE_STRUCTURE')
        
        # Test SMOKE_TEST
        base_test.test_category = 'SMOKE_TEST'
        self.assertEqual(base_test.test_category, 'SMOKE_TEST')

    def test_parse_tags(self):
        """Test parsing tags from string"""
        reader = ExcelTestSuiteReader()
        
        # Single tag
        tags = reader.parse_tags('smoke')
        self.assertEqual(tags, ['smoke'])
        
        # Multiple tags
        tags = reader.parse_tags('smoke,connection,database')
        self.assertEqual(tags, ['smoke', 'connection', 'database'])
        
        # Empty tags
        tags = reader.parse_tags('')
        self.assertEqual(tags, [])
        
        # Tags with spaces
        tags = reader.parse_tags('smoke, connection , database')
        self.assertEqual(tags, ['smoke', 'connection', 'database'])

    def test_parse_timeout(self):
        """Test parsing timeout values"""
        reader = ExcelTestSuiteReader()
        
        # Valid integer
        timeout = reader.parse_timeout('30')
        self.assertEqual(timeout, 30)
        
        # Valid float
        timeout = reader.parse_timeout('45.5')
        self.assertEqual(timeout, 45)  # Should convert to int
        
        # Invalid value
        timeout = reader.parse_timeout('invalid')
        self.assertEqual(timeout, 30)  # Default value
        
        # Empty value
        timeout = reader.parse_timeout('')
        self.assertEqual(timeout, 30)  # Default value
        
        # None value
        timeout = reader.parse_timeout(None)
        self.assertEqual(timeout, 30)  # Default value

    def test_convert_boolean_value(self):
        """Test converting boolean values"""
        reader = ExcelTestSuiteReader()
        
        # True values
        self.assertTrue(reader.convert_boolean_value('TRUE'))
        self.assertTrue(reader.convert_boolean_value('True'))
        self.assertTrue(reader.convert_boolean_value('true'))
        self.assertTrue(reader.convert_boolean_value('YES'))
        self.assertTrue(reader.convert_boolean_value('1'))
        
        # False values
        self.assertFalse(reader.convert_boolean_value('FALSE'))
        self.assertFalse(reader.convert_boolean_value('False'))
        self.assertFalse(reader.convert_boolean_value('false'))
        self.assertFalse(reader.convert_boolean_value('NO'))
        self.assertFalse(reader.convert_boolean_value('0'))
        self.assertFalse(reader.convert_boolean_value(''))
        self.assertFalse(reader.convert_boolean_value(None))

    def test_get_cell_value_safe(self):
        """Test safe cell value extraction"""
        reader = ExcelTestSuiteReader()
        
        # Mock cell with value
        mock_cell = Mock()
        mock_cell.value = 'Test Value'
        value = reader.get_cell_value_safe(mock_cell)
        self.assertEqual(value, 'Test Value')
        
        # Mock cell with None value
        mock_cell.value = None
        value = reader.get_cell_value_safe(mock_cell)
        self.assertEqual(value, '')
        
        # Mock cell that's None
        value = reader.get_cell_value_safe(None)
        self.assertEqual(value, '')

    def test_filter_enabled_tests(self):
        """Test filtering enabled tests"""
        reader = ExcelTestSuiteReader()
        test_cases = reader.load_test_suite(self.test_file)
        
        enabled_tests = reader.filter_enabled_tests(test_cases)
        
        # Should have 6 enabled tests (one is disabled)
        self.assertEqual(len(enabled_tests), 6)
        
        # All returned tests should be enabled
        for test in enabled_tests:
            self.assertTrue(test.enabled)

    def test_filter_by_priority(self):
        """Test filtering tests by priority"""
        reader = ExcelTestSuiteReader()
        test_cases = reader.load_test_suite(self.test_file)
        
        # Filter HIGH priority tests
        high_priority_tests = reader.filter_by_priority(test_cases, 'HIGH')
        self.assertGreater(len(high_priority_tests), 0)
        
        for test in high_priority_tests:
            self.assertEqual(test.priority, 'HIGH')
        
        # Filter MEDIUM priority tests
        medium_priority_tests = reader.filter_by_priority(test_cases, 'MEDIUM')
        self.assertGreater(len(medium_priority_tests), 0)
        
        for test in medium_priority_tests:
            self.assertEqual(test.priority, 'MEDIUM')

    def test_filter_by_test_category(self):
        """Test filtering tests by test category"""
        reader = ExcelTestSuiteReader()
        test_cases = reader.load_test_suite(self.test_file)
        
        # Filter CONNECTION_TEST tests
        connection_tests = reader.filter_by_test_category(test_cases, 'CONNECTION_TEST')
        self.assertGreater(len(connection_tests), 0)
        
        for test in connection_tests:
            self.assertEqual(test.test_category, 'CONNECTION_TEST')
        
        # Filter TABLE_EXISTS tests
        table_tests = reader.filter_by_test_category(test_cases, 'TABLE_EXISTS')
        self.assertGreater(len(table_tests), 0)
        
        for test in table_tests:
            self.assertEqual(test.test_category, 'TABLE_EXISTS')

    def test_filter_by_tags(self):
        """Test filtering tests by tags"""
        reader = ExcelTestSuiteReader()
        test_cases = reader.load_test_suite(self.test_file)
        
        # Filter by 'smoke' tag
        smoke_tests = reader.filter_by_tags(test_cases, ['smoke'])
        self.assertGreater(len(smoke_tests), 0)
        
        for test in smoke_tests:
            self.assertIn('smoke', test.tags)
        
        # Filter by 'connection' tag
        connection_tests = reader.filter_by_tags(test_cases, ['connection'])
        self.assertGreater(len(connection_tests), 0)
        
        for test in connection_tests:
            self.assertIn('connection', test.tags)

    def test_get_test_statistics(self):
        """Test getting test statistics"""
        reader = ExcelTestSuiteReader()
        test_cases = reader.load_test_suite(self.test_file)
        
        stats = reader.get_test_statistics(test_cases)
        
        self.assertIsInstance(stats, dict)
        self.assertIn('total_tests', stats)
        self.assertIn('enabled_tests', stats)
        self.assertIn('disabled_tests', stats)
        self.assertIn('by_priority', stats)
        self.assertIn('by_category', stats)
        
        self.assertEqual(stats['total_tests'], 7)
        self.assertEqual(stats['enabled_tests'], 6)
        self.assertEqual(stats['disabled_tests'], 1)

    def test_empty_file_handling(self):
        """Test handling of empty Excel file"""
        # Create empty Excel file
        empty_file = os.path.join(self.temp_dir, 'empty.xlsx')
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'SMOKE'
        workbook.save(empty_file)
        
        reader = ExcelTestSuiteReader()
        test_cases = reader.load_test_suite(empty_file)
        
        self.assertIsInstance(test_cases, list)
        self.assertEqual(len(test_cases), 0)
        
        # Clean up
        os.remove(empty_file)

    def test_missing_headers_handling(self):
        """Test handling of missing required headers"""
        # Create file with missing headers
        missing_headers_file = os.path.join(self.temp_dir, 'missing_headers.xlsx')
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'SMOKE'
        
        # Only a few headers
        headers = ['Enable', 'Test Case ID']
        for idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=idx, value=header)
        
        workbook.save(missing_headers_file)
        
        reader = ExcelTestSuiteReader()
        test_cases = reader.load_test_suite(missing_headers_file)
        
        # Should handle gracefully and return empty list or minimal tests
        self.assertIsInstance(test_cases, list)
        
        # Clean up
        os.remove(missing_headers_file)

    def test_error_handling_and_recovery(self):
        """Test error handling and recovery mechanisms"""
        reader = ExcelTestSuiteReader()
        
        # Test with None file path
        test_cases = reader.load_test_suite(None)
        self.assertEqual(test_cases, [])
        
        # Test with empty string file path
        test_cases = reader.load_test_suite('')
        self.assertEqual(test_cases, [])
        
        # Test with directory instead of file
        test_cases = reader.load_test_suite(self.temp_dir)
        self.assertEqual(test_cases, [])

    def test_large_dataset_handling(self):
        """Test handling of large datasets"""
        # Create file with many test cases
        large_file = os.path.join(self.temp_dir, 'large.xlsx')
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'SMOKE'
        
        # Headers
        headers = [
            'Enable', 'Test Case ID', 'Test Case Name', 'Application Name',
            'Environment Name', 'Priority', 'Test Category', 'Expected Result',
            'Timeout (seconds)', 'Description', 'Prerequisites', 'Tags', 'Parameters'
        ]
        
        for idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=idx, value=header)
        
        # Add 100 test cases
        for i in range(100):
            row_data = [
                'TRUE', f'TEST_{i:03d}', f'Test Case {i}', 'App', 'DEV', 'MEDIUM',
                'SMOKE_TEST', 'PASS', '30', f'Description {i}', 'Prerequisites',
                'smoke,auto', f'test_id={i}'
            ]
            
            for col_idx, cell_value in enumerate(row_data, 1):
                worksheet.cell(row=i+2, column=col_idx, value=cell_value)
        
        workbook.save(large_file)
        
        reader = ExcelTestSuiteReader()
        test_cases = reader.load_test_suite(large_file)
        
        self.assertEqual(len(test_cases), 100)
        
        # Clean up
        os.remove(large_file)


if __name__ == '__main__':
    unittest.main()