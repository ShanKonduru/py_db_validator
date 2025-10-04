#!/usr/bin/env python
"""
Excel Test Suite Reader
Reads and validates sdm_test_suite.xlsx file for test execution
"""
import sys
from pathlib import Path
from typing import List, Dict, Optional, Any
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent / "src"))

# Import validation module
try:
    from src.validation.excel_validator import ExcelTestSuiteValidator, ValidationSeverity
except ImportError:
    # Fallback if validation module not available
    ExcelTestSuiteValidator = None
    ValidationSeverity = None


@dataclass
class TestCase:
    """Data class representing a test case from Excel"""
    __test__ = False  # Tell pytest this is not a test class

    enable: bool
    test_case_id: str
    test_case_name: str
    application_name: str
    environment_name: str
    priority: str
    test_category: str
    expected_result: str
    timeout_seconds: int
    description: str
    prerequisites: str
    tags: str
    parameters: str = ""  # New field for test parameters

    def get_tags_list(self) -> List[str]:
        """Get tags as a list of strings"""
        if not self.tags:
            return []
        return [tag.strip() for tag in self.tags.split(",")]

    def has_tag(self, tag: str) -> bool:
        """Check if test case has a specific tag"""
        return tag.lower() in [t.lower() for t in self.get_tags_list()]

    def get_parameters_dict(self) -> Dict[str, str]:
        """Get parameters as a dictionary"""
        if not self.parameters:
            return {}
        
        params = {}
        for param in self.parameters.split(","):
            if "=" in param:
                key, value = param.split("=", 1)
                params[key.strip()] = value.strip()
            else:
                # For simple values like table names, use a default key
                params["table_name"] = param.strip()
        return params

    def get_parameter(self, key: str, default: str = "") -> str:
        """Get a specific parameter value"""
        return self.get_parameters_dict().get(key, default)

    def is_enabled(self) -> bool:
        """Check if test case is enabled"""
        return self.enable

    def matches_filter(
        self,
        environment: Optional[str] = None,
        application: Optional[str] = None,
        priority: Optional[str] = None,
        category: Optional[str] = None,
        tags: Optional[List[str]] = None,
    ) -> bool:
        """Check if test case matches the given filters"""

        if environment and self.environment_name.upper() != environment.upper():
            return False

        if application and self.application_name.upper() != application.upper():
            return False

        if priority and self.priority.upper() != priority.upper():
            return False

        if category and self.test_category.upper() != category.upper():
            return False

        if tags:
            test_tags = [t.lower() for t in self.get_tags_list()]
            for required_tag in tags:
                if required_tag.lower() not in test_tags:
                    return False

        return True


class ExcelTestSuiteReader:
    """Reads and validates Excel test suite files"""

    def __init__(self, excel_file: str, sheet_name: str = "SMOKE"):
        """Initialize with Excel file path and sheet name"""
        self.excel_file = Path(excel_file)
        self.sheet_name = sheet_name
        self.workbook: Optional[Workbook] = None
        self.test_cases: List[TestCase] = []
        self.validator = ExcelTestSuiteValidator() if ExcelTestSuiteValidator else None
        self.validation_passed = True
        self.validation_report = ""

    def validate_file_exists(self) -> bool:
        """Validate that Excel file exists"""
        return self.excel_file.exists()

    def load_workbook(self) -> bool:
        """Load the Excel workbook with validation"""
        try:
            self.workbook = load_workbook(self.excel_file)
            
            # Check if sheet exists
            if self.sheet_name not in self.workbook.sheetnames:
                print(f"❌ Sheet '{self.sheet_name}' not found in workbook")
                print(f"   Available sheets: {', '.join(self.workbook.sheetnames)}")
                return False
            
            # Run validation if validator is available
            if self.validator:
                self.validation_passed, validation_messages = self.validator.validate_test_suite(
                    self.workbook, self.sheet_name
                )
                self.validation_report = self.validator.generate_validation_report()
                
                # Print validation report
                print(f"🔍 EXCEL VALIDATION RESULTS (Sheet: {self.sheet_name}):")
                print(self.validation_report)
                
                # Count errors
                errors = [msg for msg in validation_messages if msg.severity == ValidationSeverity.ERROR]
                if errors:
                    print(f"\n❌ Found {len(errors)} critical errors that must be fixed before execution!")
                    return False
                else:
                    print(f"\n✅ Validation passed! File is ready for test execution.")
            
            return True
        except Exception as e:
            print(f"❌ Error loading Excel file: {e}")
            return False

    def get_validation_report(self) -> str:
        """Get the validation report"""
        return self.validation_report

    def is_validation_passed(self) -> bool:
        """Check if validation passed"""
        return self.validation_passed

    def validate_structure(self) -> bool:
        """Validate Excel file structure"""
        if not self.workbook:
            return False

        # Check if specified sheet exists
        if self.sheet_name not in self.workbook.sheetnames:
            print(f"❌ Missing required '{self.sheet_name}' sheet in Excel file")
            print(f"   Available sheets: {', '.join(self.workbook.sheetnames)}")
            return False

        ws = self.workbook[self.sheet_name]

        # Expected headers
        expected_headers = [
            "Enable",
            "Test_Case_ID",
            "Test_Case_Name",
            "Application_Name",
            "Environment_Name",
            "Priority",
            "Test_Category",
            "Expected_Result",
            "Timeout_Seconds",
            "Description",
            "Prerequisites",
            "Tags",
            "Parameters",  # New column for test parameters
        ]

        # Get actual headers from first row
        actual_headers = []
        for col in range(1, len(expected_headers) + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                actual_headers.append(str(cell_value).strip())

        # Validate headers
        missing_headers = set(expected_headers) - set(actual_headers)
        if missing_headers:
            print(f"❌ Missing required headers: {missing_headers}")
            return False

        return True

    def read_test_cases(self) -> bool:
        """Read test cases from Excel file"""
        if not self.workbook:
            return False

        ws = self.workbook[self.sheet_name]
        self.test_cases = []

        # Get headers mapping
        headers = {}
        for col in range(1, 14):  # 13 columns expected (including Parameters)
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers[col] = str(cell_value).strip()

        # Read data rows
        row_num = 2
        while True:
            # Check if we've reached the end (empty Test_Case_ID)
            test_id_cell = None
            for col, header in headers.items():
                if header == "Test_Case_ID":
                    test_id_cell = ws.cell(row=row_num, column=col).value
                    break

            if not test_id_cell:
                break

            try:
                # Read row data
                row_data = {}
                for col, header in headers.items():
                    cell_value = ws.cell(row=row_num, column=col).value
                    row_data[header] = cell_value

                # Convert and validate data
                test_case = TestCase(
                    enable=self._convert_bool(row_data.get("Enable", False)),
                    test_case_id=str(row_data.get("Test_Case_ID", "")),
                    test_case_name=str(row_data.get("Test_Case_Name", "")),
                    application_name=str(row_data.get("Application_Name", "")),
                    environment_name=str(row_data.get("Environment_Name", "")),
                    priority=str(row_data.get("Priority", "MEDIUM")),
                    test_category=str(row_data.get("Test_Category", "")),
                    expected_result=str(row_data.get("Expected_Result", "PASS")),
                    timeout_seconds=self._convert_int(
                        row_data.get("Timeout_Seconds", 60)
                    ),
                    description=str(row_data.get("Description", "")),
                    prerequisites=str(row_data.get("Prerequisites", "")),
                    tags=str(row_data.get("Tags", "")),
                    parameters=str(row_data.get("Parameters", "")),  # New parameters field
                )

                # Basic validation
                if test_case.test_case_id and test_case.test_case_name:
                    self.test_cases.append(test_case)
                else:
                    print(f"⚠️  Skipping invalid row {row_num}: missing required fields")

            except Exception as e:
                print(f"⚠️  Error reading row {row_num}: {e}")

            row_num += 1

        return len(self.test_cases) > 0

    def _convert_bool(self, value: Any) -> bool:
        """Convert various boolean representations to bool"""
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.upper() in ("TRUE", "YES", "1", "ON", "ENABLED")
        if isinstance(value, (int, float)):
            return value != 0
        return False

    def _convert_int(self, value: Any) -> int:
        """Convert value to integer with fallback"""
        try:
            return int(value)
        except (ValueError, TypeError):
            return 60  # Default timeout

    def get_all_test_cases(self) -> List[TestCase]:
        """Get all test cases"""
        # Load data if not already loaded
        if not self.test_cases and not self.workbook:
            if not self.load_workbook():
                return []
            if not self.validate_structure():
                return []
            if not self.read_test_cases():
                return []
        
        return self.test_cases

    def get_enabled_test_cases(self) -> List[TestCase]:
        """Get only enabled test cases"""
        return [tc for tc in self.test_cases if tc.is_enabled()]

    def get_filtered_test_cases(
        self,
        environment: Optional[str] = None,
        application: Optional[str] = None,
        priority: Optional[str] = None,
        category: Optional[str] = None,
        tags: Optional[List[str]] = None,
        enabled_only: bool = True,
    ) -> List[TestCase]:
        """Get test cases matching the given filters"""

        test_cases = (
            self.get_enabled_test_cases() if enabled_only else self.get_all_test_cases()
        )

        filtered_cases = []
        for test_case in test_cases:
            if test_case.matches_filter(
                environment, application, priority, category, tags
            ):
                filtered_cases.append(test_case)

        return filtered_cases

    def get_test_case_by_id(self, test_id: str) -> Optional[TestCase]:
        """Get a specific test case by ID"""
        for test_case in self.test_cases:
            if test_case.test_case_id == test_id:
                return test_case
        return None

    def get_statistics(self) -> Dict[str, Any]:
        """Get statistics about loaded test cases"""
        if not self.test_cases:
            return {}

        total_tests = len(self.test_cases)
        enabled_tests = len(self.get_enabled_test_cases())
        disabled_tests = total_tests - enabled_tests

        # Count by priority
        priority_counts = {}
        for tc in self.test_cases:
            priority = tc.priority
            priority_counts[priority] = priority_counts.get(priority, 0) + 1

        # Count by category
        category_counts = {}
        for tc in self.test_cases:
            category = tc.test_category
            category_counts[category] = category_counts.get(category, 0) + 1

        # Count by environment
        env_counts = {}
        for tc in self.test_cases:
            env = tc.environment_name
            env_counts[env] = env_counts.get(env, 0) + 1

        # Count by application
        app_counts = {}
        for tc in self.test_cases:
            app = tc.application_name
            app_counts[app] = app_counts.get(app, 0) + 1

        return {
            "total_tests": total_tests,
            "enabled_tests": enabled_tests,
            "disabled_tests": disabled_tests,
            "priority_distribution": priority_counts,
            "category_distribution": category_counts,
            "environment_distribution": env_counts,
            "application_distribution": app_counts,
        }

    def load_and_validate(self) -> bool:
        """Complete loading and validation process"""
        print(f"📊 Loading Excel test suite: {self.excel_file}")

        if not self.validate_file_exists():
            print(f"❌ Excel file not found: {self.excel_file}")
            return False

        if not self.load_workbook():
            return False

        if not self.validate_structure():
            return False

        if not self.read_test_cases():
            print("❌ No valid test cases found in Excel file")
            return False

        # Print statistics
        stats = self.get_statistics()
        print(f"✅ Successfully loaded {stats['total_tests']} test cases")
        print(f"   - Enabled: {stats['enabled_tests']}")
        print(f"   - Disabled: {stats['disabled_tests']}")
        print(f"   - Priorities: {stats['priority_distribution']}")
        print(f"   - Categories: {stats['category_distribution']}")
        print(f"   - Environments: {stats['environment_distribution']}")
        print(f"   - Applications: {stats['application_distribution']}")

        return True


def main():
    """Test the Excel reader with sample file"""
    reader = ExcelTestSuiteReader("sdm_test_suite.xlsx")

    if reader.load_and_validate():
        print("\n🧪 Example Filters:")

        # Get all enabled tests
        enabled_tests = reader.get_enabled_test_cases()
        print(f"   - Enabled tests: {len(enabled_tests)}")

        # Get high priority tests
        high_priority = reader.get_filtered_test_cases(priority="HIGH")
        print(f"   - High priority tests: {len(high_priority)}")

        # Get connection tests
        connection_tests = reader.get_filtered_test_cases(category="CONNECTION")
        print(f"   - Connection tests: {len(connection_tests)}")

        # Get tests with 'smoke' tag
        smoke_tests = reader.get_filtered_test_cases(tags=["smoke"])
        print(f"   - Smoke tagged tests: {len(smoke_tests)}")

        return 0
    else:
        return 1


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
