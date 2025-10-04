"""
Excel Template Generator with Data Validation Dropdowns

This module creates Excel templates with data validation dropdowns to prevent
user input errors and ensure data consistency.
"""
import sys
from pathlib import Path
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add project root to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from src.validation.excel_validator import ExcelTestSuiteValidator


class ExcelTemplateGenerator:
    """Generates Excel templates with data validation dropdowns"""
    
    def __init__(self):
        """Initialize the template generator"""
        self.validator = ExcelTestSuiteValidator()
        
        # Define dropdown options for each column
        self.dropdown_options = {
            "Enable": ["TRUE", "FALSE"],
            "Priority": list(self.validator.VALID_PRIORITIES),
            "Test_Category": list(self.validator.VALID_TEST_CATEGORIES.keys()),
            "Expected_Result": list(self.validator.VALID_EXPECTED_RESULTS),
            "Environment_Name": list(self.validator.VALID_ENVIRONMENTS),
            "Application_Name": list(self.validator.VALID_APPLICATIONS),
        }
        
        # Column definitions with their properties
        self.column_definitions = [
            {"name": "Enable", "width": 12, "dropdown": True, "required": True},
            {"name": "Test_Case_ID", "width": 18, "dropdown": False, "required": True},
            {"name": "Test_Case_Name", "width": 30, "dropdown": False, "required": True},
            {"name": "Application_Name", "width": 18, "dropdown": True, "required": True},
            {"name": "Environment_Name", "width": 18, "dropdown": True, "required": True},
            {"name": "Priority", "width": 12, "dropdown": True, "required": True},
            {"name": "Test_Category", "width": 18, "dropdown": True, "required": True},
            {"name": "Expected_Result", "width": 16, "dropdown": True, "required": True},
            {"name": "Timeout_Seconds", "width": 16, "dropdown": False, "required": True},
            {"name": "Description", "width": 40, "dropdown": False, "required": False},
            {"name": "Prerequisites", "width": 30, "dropdown": False, "required": False},
            {"name": "Tags", "width": 20, "dropdown": False, "required": False},
            {"name": "Parameters", "width": 25, "dropdown": False, "required": False},  # New column
        ]
    
    def create_template(self, filename: str = "test_suite_template.xlsx", 
                       include_sample_data: bool = True,
                       include_controller: bool = True) -> bool:
        """
        Create Excel template with data validation dropdowns
        
        Args:
            filename: Output Excel file name
            include_sample_data: Whether to include sample test data
            include_controller: Whether to include CONTROLLER sheet
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "SMOKE"
            
            # Set up headers with styling
            self._create_headers(ws)
            
            # Add data validation dropdowns
            self._add_data_validations(ws)
            
            # Add sample data if requested
            if include_sample_data:
                self._add_sample_data(ws)
            
            # Create controller sheet if requested
            if include_controller:
                self._create_controller_sheet(wb)
            
            # Create data validations sheet for advanced testing
            self._create_data_validations_sheet(wb)
            
            # Create instructions worksheet
            self._create_instructions_worksheet(wb)
            
            # Create validation reference worksheet
            self._create_reference_worksheet(wb)
            
            # Save the file
            wb.save(filename)
            print(f"✅ Excel template created: {filename}")
            return True
            
        except Exception as e:
            print(f"❌ Error creating Excel template: {e}")
            return False
    
    def _create_headers(self, ws):
        """Create and style headers"""
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers
        for col_idx, col_def in enumerate(self.column_definitions, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_def["name"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            
            # Set column width
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = col_def["width"]
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def _add_data_validations(self, ws):
        """Add data validation dropdowns to appropriate columns"""
        # Define the range for data validation (rows 2 to 1000)
        start_row = 2
        end_row = 1000
        
        for col_idx, col_def in enumerate(self.column_definitions, 1):
            column_letter = get_column_letter(col_idx)
            
            if col_def["dropdown"] and col_def["name"] in self.dropdown_options:
                # Create dropdown validation
                options = self.dropdown_options[col_def["name"]]
                formula = f'"{",".join(options)}"'
                
                # Create data validation
                dv = DataValidation(
                    type="list",
                    formula1=formula,
                    allow_blank=not col_def["required"]
                )
                
                # Set error messages
                dv.error = f"Invalid value for {col_def['name']}"
                dv.errorTitle = "Invalid Input"
                dv.prompt = f"Please select from: {', '.join(options)}"
                dv.promptTitle = f"Select {col_def['name']}"
                
                # Apply to range
                range_string = f"{column_letter}{start_row}:{column_letter}{end_row}"
                dv.add(range_string)
                ws.add_data_validation(dv)
            
            elif col_def["name"] == "Timeout_Seconds":
                # Numeric validation for timeout
                dv = DataValidation(
                    type="whole",
                    operator="between",
                    formula1=5,
                    formula2=3600,
                    allow_blank=False
                )
                dv.error = "Timeout must be between 5 and 3600 seconds"
                dv.errorTitle = "Invalid Timeout"
                dv.prompt = "Enter timeout in seconds (5-3600)"
                dv.promptTitle = "Timeout Seconds"
                
                range_string = f"{column_letter}{start_row}:{column_letter}{end_row}"
                dv.add(range_string)
                ws.add_data_validation(dv)
    
    def _add_sample_data(self, ws):
        """Add sample test data"""
        sample_data = [
            [
                "TRUE", "SMOKE_PG_001", "Environment Setup Validation", "DUMMY", "DEV", 
                "HIGH", "SETUP", "PASS", 30, 
                "Validates that environment and configuration are properly set up",
                "Configuration file or environment variables must be available",
                "smoke,db,setup", ""
            ],
            [
                "TRUE", "SMOKE_PG_002", "Configuration Availability", "DUMMY", "DEV",
                "HIGH", "CONFIGURATION", "PASS", 30,
                "Checks if database configuration is available and accessible",
                "Environment setup must be completed",
                "smoke,db,config", ""
            ],
            [
                "TRUE", "SMOKE_PG_003", "Credentials Validation", "DUMMY", "DEV",
                "HIGH", "SECURITY", "PASS", 30,
                "Validates database credentials and authentication",
                "Valid credentials must be configured",
                "smoke,db,security", ""
            ],
            [
                "TRUE", "SMOKE_PG_004", "Database Connectivity", "DUMMY", "DEV",
                "HIGH", "CONNECTION", "PASS", 60,
                "Tests basic database connection functionality",
                "Database server must be running and accessible",
                "smoke,db,connection", ""
            ],
            [
                "TRUE", "SMOKE_PG_005", "Basic Database Queries", "DUMMY", "DEV",
                "HIGH", "QUERIES", "PASS", 60,
                "Tests basic SQL query execution capabilities",
                "Database connection must be established",
                "smoke,db,queries", ""
            ],
            [
                "FALSE", "SMOKE_PG_006", "Connection Performance", "DUMMY", "DEV",
                "MEDIUM", "PERFORMANCE", "PASS", 30,
                "Tests database connection performance metrics",
                "Database must be optimized for performance testing",
                "smoke,db,performance", ""
            ],
            [
                "TRUE", "SMOKE_PG_007", "Products Table Exists", "DUMMY", "DEV",
                "HIGH", "TABLE_EXISTS", "PASS", 30,
                "Verifies that the products table exists in the database",
                "Database connection must be established",
                "smoke,table,structure", "table_name=public.products"
            ],
            [
                "TRUE", "SMOKE_PG_008", "Employees Table Exists", "DUMMY", "DEV",
                "HIGH", "TABLE_EXISTS", "PASS", 30,
                "Verifies that the employees table exists in the database",
                "Database connection must be established",
                "smoke,table,structure", "table_name=public.employees"
            ],
            [
                "TRUE", "SMOKE_PG_009", "Orders Table Exists", "DUMMY", "DEV",
                "HIGH", "TABLE_EXISTS", "PASS", 30,
                "Verifies that the orders table exists in the database",
                "Database connection must be established",
                "smoke,table,structure", "table_name=public.orders"
            ],
            [
                "TRUE", "SMOKE_PG_010", "Products Table Select Test", "DUMMY", "DEV",
                "HIGH", "TABLE_SELECT", "PASS", 30,
                "Tests SELECT operations on the products table",
                "Products table must exist and be accessible",
                "smoke,table,queries", "table_name=public.products"
            ],
            [
                "TRUE", "SMOKE_PG_011", "Employees Table Select Test", "DUMMY", "DEV",
                "HIGH", "TABLE_SELECT", "PASS", 30,
                "Tests SELECT operations on the employees table",
                "Employees table must exist and be accessible",
                "smoke,table,queries", "table_name=public.employees"
            ],
            [
                "TRUE", "SMOKE_PG_012", "Orders Table Select Test", "DUMMY", "DEV",
                "HIGH", "TABLE_SELECT", "PASS", 30,
                "Tests SELECT operations on the orders table",
                "Orders table must exist and be accessible",
                "smoke,table,queries", "table_name=public.orders"
            ],
            [
                "TRUE", "SMOKE_PG_013", "Products Table Has Data", "DUMMY", "DEV",
                "MEDIUM", "TABLE_ROWS", "PASS", 30,
                "Verifies that the products table contains data",
                "Products table must exist and be populated",
                "smoke,table,data", "table_name=public.products,min_rows=1"
            ],
            [
                "TRUE", "SMOKE_PG_014", "Employees Table Has Data", "DUMMY", "DEV",
                "MEDIUM", "TABLE_ROWS", "PASS", 30,
                "Verifies that the employees table contains data",
                "Employees table must exist and be populated",
                "smoke,table,data", "table_name=public.employees,min_rows=1"
            ],
            [
                "TRUE", "SMOKE_PG_015", "Orders Table Has Data", "DUMMY", "DEV",
                "MEDIUM", "TABLE_ROWS", "PASS", 30,
                "Verifies that the orders table contains data",
                "Orders table must exist and be populated",
                "smoke,table,data", "table_name=public.orders,min_rows=1"
            ],
            [
                "TRUE", "SMOKE_PG_016", "Products Table Structure", "DUMMY", "DEV",
                "LOW", "TABLE_STRUCTURE", "PASS", 30,
                "Validates products table structure and columns",
                "Products table must exist",
                "smoke,table,structure", "table_name=public.products"
            ],
            [
                "TRUE", "SMOKE_PG_017", "Employees Table Structure", "DUMMY", "DEV",
                "LOW", "TABLE_STRUCTURE", "PASS", 30,
                "Validates employees table structure and columns",
                "Employees table must exist",
                "smoke,table,structure", "table_name=public.employees"
            ],
            [
                "TRUE", "SMOKE_PG_018", "Orders Table Structure", "DUMMY", "DEV",
                "LOW", "TABLE_STRUCTURE", "PASS", 30,
                "Validates orders table structure and columns",
                "Orders table must exist",
                "smoke,table,structure", "table_name=public.orders"
            ]
        ]
        
        # Add sample data rows
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                # Add subtle styling to data rows
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    def _create_instructions_worksheet(self, wb):
        """Create instructions worksheet"""
        ws = wb.create_sheet("INSTRUCTIONS")
        
        # Title
        ws['A1'] = "📋 Excel Test Suite Instructions"
        ws['A1'].font = Font(size=16, bold=True, color="366092")
        
        instructions = [
            "",
            "🎯 HOW TO USE THIS TEMPLATE:",
            "",
            "🎛️ CONTROLLER SHEET - MULTI-SHEET EXECUTION:",
            "",
            "The CONTROLLER sheet allows you to enable/disable entire test sheets for execution:",
            "• Enable=TRUE: Include sheet in test execution",
            "• Enable=FALSE: Skip sheet during execution", 
            "• Sheet_Name: Name of the test sheet (SMOKE, INTEGRATION, etc.)",
            "• Description: Brief description of what the sheet tests",
            "• Priority: Execution priority (HIGH, MEDIUM, LOW)",
            "",
            "Multi-sheet execution: python excel_test_driver.py --multi-sheet --reports",
            "Single sheet execution: python excel_test_driver.py --sheet SMOKE --reports",
            "",
            "1. Use the 'CONTROLLER' sheet to enable/disable test sheet execution",
            "2. Use individual sheets (SMOKE, etc.) to define your test cases",
            "3. Fill in each row with test case details",
            "4. Use DROPDOWNS for columns with validation (they have colored headers)",
            "5. Required fields are marked with * in the reference sheet",
            "",
            "�️ PARAMETERIZED TABLE TESTING:",
            "",
            "The Parameters column enables parameterized testing for table validation:",
            "• TABLE_EXISTS: table_name=public.users",
            "• TABLE_SELECT: table_name=public.products", 
            "• TABLE_ROWS: table_name=public.orders,min_rows=10",
            "• TABLE_STRUCTURE: table_name=public.employees",
            "",
            "Parameter format: key=value or key1=value1,key2=value2",
            "Example: table_name=public.products,min_rows=5",
            "",
            "�📝 IMPORTANT NOTES:",
            "",
            "• Test_Case_ID must be unique within each sheet",
            "• Test_Category determines which test function will be executed",
            "• Use dropdowns to prevent data entry errors",
            "• Timeout_Seconds must be between 5-3600 seconds",
            "• Tags should be comma-separated without spaces",
            "• Parameters column is optional for non-table tests",
            "",
            "🔍 VALIDATION:",
            "",
            "• Run 'python validate_excel.py' to check for errors",
            "• Fix all validation errors before running tests",
            "• Use 'python validate_excel.py --fix-suggestions' for help",
            "",
            "🚀 EXECUTION:",
            "",
            "• Run 'python excel_test_driver.py --multi-sheet --reports' for multi-sheet execution",
            "• Use 'python excel_test_driver.py --reports' for single sheet (SMOKE) execution",
            "• Use filters like --priority HIGH or --category CONNECTION",
            "• Generated reports will be saved in test_reports/ directory",
            "",
            "💡 TIPS:",
            "",
            "• Copy existing rows and modify instead of typing from scratch",
            "• Always validate before sharing the file with others",
            "• Check the REFERENCE sheet for valid values and mappings",
            "• Use CONTROLLER sheet to manage execution of multiple test suites",
            "• For table tests, always include schema.table_name format",
            "",
        ]
        
        for i, instruction in enumerate(instructions, 2):
            ws[f'A{i}'] = instruction
            if instruction.startswith(('🎯', '📝', '🔍', '🚀', '💡')):
                ws[f'A{i}'].font = Font(bold=True, color="D63384")
        
        # Set column width
        ws.column_dimensions['A'].width = 80
    
    def _create_reference_worksheet(self, wb):
        """Create reference worksheet with valid values"""
        ws = wb.create_sheet("REFERENCE")
        
        # Title
        ws['A1'] = "📚 Test Category Reference & Valid Values"
        ws['A1'].font = Font(size=16, bold=True, color="366092")
        
        # Test Category Mapping Table
        ws['A3'] = "🎯 TEST CATEGORY → FUNCTION MAPPING:"
        ws['A3'].font = Font(size=14, bold=True, color="D63384")
        
        # Headers for mapping table
        headers = ["Test_Category", "Python Function", "Status", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Mapping data
        mappings = [
            ["SETUP", "test_environment_setup()", "✅ Available", "Validates environment setup"],
            ["CONFIGURATION", "test_dummy_config_availability()", "✅ Available", "Checks configuration availability"],
            ["SECURITY", "test_environment_credentials()", "✅ Available", "Validates credentials"],
            ["CONNECTION", "test_postgresql_connection()", "✅ Available", "Tests database connection"],
            ["QUERIES", "test_postgresql_basic_queries()", "✅ Available", "Tests SQL query execution"],
            ["PERFORMANCE", "test_postgresql_connection_performance()", "✅ Available", "Tests connection performance"],
            ["TABLE_EXISTS", "smoke_test_table_exists()", "✅ Available", "Validates table existence (requires table_name parameter)"],
            ["TABLE_SELECT", "smoke_test_table_select_possible()", "✅ Available", "Tests table SELECT access (requires table_name parameter)"],
            ["TABLE_ROWS", "smoke_test_table_has_rows()", "✅ Available", "Validates table has data (requires table_name, optional min_rows parameter)"],
            ["TABLE_STRUCTURE", "smoke_test_table_structure()", "✅ Available", "Analyzes table structure (requires table_name parameter)"],
            ["COMPATIBILITY", "test_compatibility()", "⚠️ Not implemented", "Backwards compatibility testing"],
            ["MONITORING", "test_monitoring()", "🔜 Future", "System monitoring tests"],
            ["BACKUP", "test_backup_restore()", "🔜 Future", "Backup and restore tests"],
        ]
        
        for row, mapping in enumerate(mappings, 6):
            for col, value in enumerate(mapping, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Valid Values Section
        ws['A16'] = "📊 VALID VALUES FOR DROPDOWN FIELDS:"
        ws['A16'].font = Font(size=14, bold=True, color="D63384")
        
        row = 18
        for field, values in self.dropdown_options.items():
            ws[f'A{row}'] = f"{field}:"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = ", ".join(values)
            row += 1
        
        # Field Requirements
        ws[f'A{row + 2}'] = "⚠️ FIELD REQUIREMENTS:"
        ws[f'A{row + 2}'].font = Font(size=14, bold=True, color="D63384")
        
        requirements = [
            "Test_Case_ID: Must be unique, format: SMOKE_PG_001",
            "Test_Case_Name: Required, descriptive name",
            "Timeout_Seconds: Number between 5-3600",
            "Description: Max 500 characters",
            "Prerequisites: Max 1000 characters",
            "Tags: Comma-separated, no spaces in individual tags",
            "Parameters: Optional, format: key=value or key1=value1,key2=value2",
            "          • For table tests: table_name=schema.tablename",
            "          • TABLE_ROWS also supports: min_rows=number",
        ]
        
        for i, req in enumerate(requirements, row + 4):
            ws[f'A{i}'] = f"• {req}"
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 35
    
    def _create_controller_sheet(self, wb):
        """Create CONTROLLER sheet for managing test sheet execution"""
        ws = wb.create_sheet("CONTROLLER")
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="D63384", end_color="D63384", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Controller headers
        controller_headers = [
            {"name": "Enable", "width": 12},
            {"name": "Sheet_Name", "width": 20},
            {"name": "Description", "width": 40},
            {"name": "Priority", "width": 15}
        ]
        
        # Add headers
        for col_idx, col_def in enumerate(controller_headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_def["name"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            
            # Set column width
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = col_def["width"]
        
        # Add sample controller data
        controller_data = [
            ["TRUE", "SMOKE", "PostgreSQL smoke tests - basic connectivity and functionality", "HIGH"],
            ["FALSE", "INTEGRATION", "Integration tests with external systems", "MEDIUM"],
            ["FALSE", "PERFORMANCE", "Performance and load testing suite", "LOW"],
            ["FALSE", "SECURITY", "Security and penetration testing", "HIGH"],
            ["FALSE", "REGRESSION", "Full regression test suite", "MEDIUM"],
            ["TRUE", "DATAVALIDATIONS", "Advanced data validation tests - schema, row count, and data quality validation", "HIGH"]
        ]
        
        # Add sample data
        for row_idx, row_data in enumerate(controller_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                # Add subtle styling to data rows
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Add data validation for Enable column
        enable_dv = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=False
        )
        enable_dv.error = "Must be TRUE or FALSE"
        enable_dv.errorTitle = "Invalid Enable Value"
        enable_dv.prompt = "Select TRUE to enable sheet execution, FALSE to disable"
        enable_dv.promptTitle = "Enable Sheet Execution"
        
        # Apply to Enable column (A2:A1000)
        enable_dv.add("A2:A1000")
        ws.add_data_validation(enable_dv)
        
        # Add data validation for Priority column
        priority_dv = DataValidation(
            type="list",
            formula1='"HIGH,MEDIUM,LOW"',
            allow_blank=True
        )
        priority_dv.error = "Must be HIGH, MEDIUM, or LOW"
        priority_dv.errorTitle = "Invalid Priority"
        priority_dv.prompt = "Select execution priority for this sheet"
        priority_dv.promptTitle = "Sheet Priority"
        
        # Apply to Priority column (D2:D1000)
        priority_dv.add("D2:D1000")
        ws.add_data_validation(priority_dv)
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Don't insert rows - keep headers in row 1 for MultiSheetTestController compatibility
        # The MultiSheetTestController expects headers in row 1
        
        # Add a note in cell A1 of INSTRUCTIONS sheet instead of modifying CONTROLLER structure
        # The CONTROLLER sheet should maintain its simple structure for programmatic access
    
    def _create_data_validations_sheet(self, wb):
        """Create DATAVALIDATIONS sheet for advanced src vs target testing"""
        ws = wb.create_sheet("DATAVALIDATIONS")
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Data validation headers
        headers = [
            "Enable", "Test_Case_ID", "Test_Case_Name", "Application_Name",
            "Environment_Name", "Priority", "Test_Category", "Expected_Result",
            "Timeout_Seconds", "Description", "Prerequisites", "Tags", 
            "Source_Table", "Target_Table", "Validation_Type", "Source_Schema", 
            "Target_Schema", "Column_Mapping", "Parameters"
        ]
        
        # Create headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        column_widths = {
            1: 12,   # Enable
            2: 18,   # Test_Case_ID
            3: 30,   # Test_Case_Name
            4: 18,   # Application_Name
            5: 18,   # Environment_Name
            6: 12,   # Priority
            7: 20,   # Test_Category
            8: 16,   # Expected_Result
            9: 16,   # Timeout_Seconds
            10: 40,  # Description
            11: 30,  # Prerequisites
            12: 20,  # Tags
            13: 25,  # Source_Table
            14: 25,  # Target_Table
            15: 20,  # Validation_Type
            16: 20,  # Source_Schema
            17: 20,  # Target_Schema
            18: 30,  # Column_Mapping
            19: 25   # Parameters
        }
        
        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Add sample data validation test cases
        sample_data = [
            # Schema Validation Tests
            [
                "TRUE", "DVAL_001", "Products Schema Validation", "POSTGRES", "DEV", "HIGH",
                "SCHEMA_VALIDATION", "PASS", "60", 
                "Compare schema structure between source products and target new_products tables",
                "Both tables must exist", "data_validation,schema", 
                "public.products", "public.new_products", "SCHEMA_COMPARE", 
                "public", "public", "ALL_COLUMNS", "ignore_sequence=true"
            ],
            [
                "TRUE", "DVAL_002", "Employees Schema Validation", "POSTGRES", "DEV", "HIGH",
                "SCHEMA_VALIDATION", "PASS", "60",
                "Compare schema structure between source employees and target new_employees tables",
                "Both tables must exist", "data_validation,schema",
                "public.employees", "public.new_employees", "SCHEMA_COMPARE",
                "public", "public", "ALL_COLUMNS", "ignore_sequence=true"
            ],
            [
                "TRUE", "DVAL_003", "Orders Schema Validation", "POSTGRES", "DEV", "HIGH",
                "SCHEMA_VALIDATION", "PASS", "60",
                "Compare schema structure between source orders and target new_orders tables",
                "Both tables must exist", "data_validation,schema",
                "public.orders", "public.new_orders", "SCHEMA_COMPARE",
                "public", "public", "ALL_COLUMNS", "ignore_sequence=true"
            ],
            
            # Row Count Validation Tests
            [
                "TRUE", "DVAL_004", "Products Row Count Validation", "POSTGRES", "DEV", "MEDIUM",
                "ROW_COUNT_VALIDATION", "PASS", "30",
                "Compare row counts between source products and target new_products tables",
                "Both tables must exist and have data", "data_validation,row_count",
                "public.products", "public.new_products", "ROW_COUNT_COMPARE",
                "public", "public", "COUNT(*)", "tolerance_percent=5"
            ],
            [
                "TRUE", "DVAL_005", "Employees Row Count Validation", "POSTGRES", "DEV", "MEDIUM",
                "ROW_COUNT_VALIDATION", "PASS", "30",
                "Compare row counts between source employees and target new_employees tables",
                "Both tables must exist and have data", "data_validation,row_count",
                "public.employees", "public.new_employees", "ROW_COUNT_COMPARE",
                "public", "public", "COUNT(*)", "tolerance_percent=5"
            ],
            [
                "TRUE", "DVAL_006", "Orders Row Count Validation", "POSTGRES", "DEV", "MEDIUM",
                "ROW_COUNT_VALIDATION", "PASS", "30",
                "Compare row counts between source orders and target new_orders tables",
                "Both tables must exist and have data", "data_validation,row_count",
                "public.orders", "public.new_orders", "ROW_COUNT_COMPARE",
                "public", "public", "COUNT(*)", "tolerance_percent=5"
            ],
            
            # Column-to-Column NULL Validation Tests
            [
                "TRUE", "DVAL_007", "Products NULL Value Validation", "POSTGRES", "DEV", "HIGH",
                "NULL_VALUE_VALIDATION", "PASS", "45",
                "Compare NULL patterns between source and target product tables",
                "Both tables must exist", "data_validation,null_check",
                "public.products", "public.new_products", "NULL_COMPARE",
                "public", "public", "product_id,product_name,category_id,price,stock_quantity",
                "null_match_required=true"
            ],
            [
                "TRUE", "DVAL_008", "Employees NULL Value Validation", "POSTGRES", "DEV", "HIGH",
                "NULL_VALUE_VALIDATION", "PASS", "45",
                "Compare NULL patterns between source and target employee tables",
                "Both tables must exist", "data_validation,null_check",
                "public.employees", "public.new_employees", "NULL_COMPARE",
                "public", "public", "employee_id,first_name,last_name,email,hire_date,salary",
                "null_match_required=true"
            ],
            [
                "TRUE", "DVAL_009", "Orders NULL Value Validation", "POSTGRES", "DEV", "HIGH",
                "NULL_VALUE_VALIDATION", "PASS", "45",
                "Compare NULL patterns between source and target order tables",
                "Both tables must exist", "data_validation,null_check",
                "public.orders", "public.new_orders", "NULL_COMPARE",
                "public", "public", "order_id,customer_id,employee_id,order_date,required_date,shipped_date",
                "null_match_required=true"
            ],
            
            # Data Quality Validation Tests
            [
                "TRUE", "DVAL_010", "Products Data Quality Check", "POSTGRES", "DEV", "MEDIUM",
                "DATA_QUALITY_VALIDATION", "PASS", "60",
                "Validate data quality metrics between source and target product tables",
                "Both tables must exist", "data_validation,data_quality",
                "public.products", "public.new_products", "DATA_QUALITY_COMPARE",
                "public", "public", "price,stock_quantity,category_id",
                "check_ranges=true,check_patterns=true"
            ]
        ]
        
        # Add sample data
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                # Add subtle styling to data rows
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Add data validation dropdowns for key columns
        self._add_data_validations_dropdowns(ws)
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def _add_data_validations_dropdowns(self, ws):
        """Add dropdown validations for DATAVALIDATIONS sheet"""
        
        # Enable dropdown (Column A)
        enable_dv = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=False
        )
        enable_dv.error = "Must be TRUE or FALSE"
        enable_dv.errorTitle = "Invalid Enable Value"
        enable_dv.add("A2:A1000")
        ws.add_data_validation(enable_dv)
        
        # Priority dropdown (Column F)
        priority_dv = DataValidation(
            type="list",
            formula1='"HIGH,MEDIUM,LOW"',
            allow_blank=False
        )
        priority_dv.error = "Must be HIGH, MEDIUM, or LOW"
        priority_dv.errorTitle = "Invalid Priority"
        priority_dv.add("F2:F1000")
        ws.add_data_validation(priority_dv)
        
        # Test Category dropdown (Column G)
        test_category_dv = DataValidation(
            type="list",
            formula1='"SCHEMA_VALIDATION,ROW_COUNT_VALIDATION,NULL_VALUE_VALIDATION,DATA_QUALITY_VALIDATION,COLUMN_COMPARE_VALIDATION"',
            allow_blank=False
        )
        test_category_dv.error = "Must be a valid data validation test category"
        test_category_dv.errorTitle = "Invalid Test Category"
        test_category_dv.add("G2:G1000")
        ws.add_data_validation(test_category_dv)
        
        # Expected Result dropdown (Column H)
        expected_result_dv = DataValidation(
            type="list",
            formula1='"PASS,FAIL"',
            allow_blank=False
        )
        expected_result_dv.error = "Must be PASS or FAIL"
        expected_result_dv.errorTitle = "Invalid Expected Result"
        expected_result_dv.add("H2:H1000")
        ws.add_data_validation(expected_result_dv)
        
        # Validation Type dropdown (Column O)
        validation_type_dv = DataValidation(
            type="list",
            formula1='"SCHEMA_COMPARE,ROW_COUNT_COMPARE,NULL_COMPARE,DATA_QUALITY_COMPARE,COLUMN_COMPARE"',
            allow_blank=False
        )
        validation_type_dv.error = "Must be a valid validation type"
        validation_type_dv.errorTitle = "Invalid Validation Type"
        validation_type_dv.add("O2:O1000")
        ws.add_data_validation(validation_type_dv)

    def update_existing_file(self, filename: str) -> bool:
        """
        Update an existing Excel file to add data validation dropdowns
        
        Args:
            filename: Path to existing Excel file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            from openpyxl import load_workbook
            
            # Load existing workbook
            wb = load_workbook(filename)
            
            if "SMOKE" not in wb.sheetnames:
                print(f"❌ Worksheet 'SMOKE' not found in {filename}")
                return False
            
            ws = wb["SMOKE"]
            
            # Add data validations to existing file
            self._add_data_validations(ws)
            
            # Create additional worksheets if they don't exist
            if "INSTRUCTIONS" not in wb.sheetnames:
                self._create_instructions_worksheet(wb)
            
            if "REFERENCE" not in wb.sheetnames:
                self._create_reference_worksheet(wb)
            
            # Save updated file
            backup_name = f"{Path(filename).stem}_backup{Path(filename).suffix}"
            wb.save(backup_name)
            wb.save(filename)
            
            print(f"✅ Updated {filename} with data validation dropdowns")
            print(f"📄 Backup saved as: {backup_name}")
            return True
            
        except Exception as e:
            print(f"❌ Error updating Excel file: {e}")
            return False


def main():
    """Main function for testing"""
    generator = ExcelTemplateGenerator()
    
    # Create template with sample data and controller
    print("Creating Excel template with data validation dropdowns and CONTROLLER sheet...")
    success = generator.create_template("test_suite_template_with_controller.xlsx", True, True)
    
    if success:
        print("\n🎉 Template created successfully!")
        print("\n📋 What was created:")
        print("• SMOKE worksheet with data validation dropdowns")
        print("• CONTROLLER worksheet for managing test sheet execution")
        print("• INSTRUCTIONS worksheet with usage guide")
        print("• REFERENCE worksheet with valid values and function mappings")
        print("\n💡 Next steps:")
        print("1. Open the Excel file and try the dropdowns")
        print("2. Use the CONTROLLER sheet to enable/disable test sheets")
        print("3. Copy/modify the sample data for your tests")
        print("4. Validate with: python validate_excel.py test_suite_template_with_controller.xlsx")
        print("5. Execute with: python excel_test_driver.py --excel-file test_suite_template_with_controller.xlsx --multi-sheet --reports")


if __name__ == "__main__":
    main()