"""
Multi-Sheet Test Controller

This module manages execution of multiple test sheets based on a CONTROLLER sheet
that enables/disables different test suites dynamically.
"""
import sys
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

# Add project root to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from src.utils.excel_test_suite_reader import ExcelTestSuiteReader, TestCase
from src.core.test_executor import TestExecutor
from src.models.test_result import TestResult
from src.validation.excel_validator import ExcelTestSuiteValidator, ValidationSeverity


@dataclass
class SheetController:
    """Represents a sheet control entry"""
    enable: bool
    sheet_name: str
    description: Optional[str] = None
    priority: Optional[str] = None
    test_count: int = 0
    enabled_test_count: int = 0


class MultiSheetTestController:
    """Controls execution of multiple test sheets based on CONTROLLER sheet"""
    
    def __init__(self, excel_file: str):
        """Initialize the multi-sheet controller"""
        self.excel_file = Path(excel_file)
        self.workbook: Optional[Workbook] = None
        self.sheet_controllers: List[SheetController] = []
        self.validator = ExcelTestSuiteValidator()
        self.validation_passed = True
        self.validation_messages = []
    
    def load_workbook(self) -> bool:
        """Load the Excel workbook"""
        try:
            self.workbook = load_workbook(self.excel_file)
            return True
        except Exception as e:
            print(f"❌ Error loading Excel file: {e}")
            return False
    
    def validate_controller_sheet(self) -> bool:
        """Validate the CONTROLLER sheet structure and data"""
        if "CONTROLLER" not in self.workbook.sheetnames:
            print("❌ CONTROLLER sheet not found in Excel file!")
            print("💡 Use 'python create_excel_template.py --add-controller' to add CONTROLLER sheet")
            return False
        
        ws = self.workbook["CONTROLLER"]
        
        # Validate headers
        expected_headers = ["Enable", "Sheet_Name", "Description", "Priority"]
        actual_headers = []
        
        for col in range(1, 5):
            cell_value = ws.cell(row=1, column=col).value
            actual_headers.append(str(cell_value).strip() if cell_value else "")
        
        for i, expected in enumerate(expected_headers):
            if i < len(actual_headers) and actual_headers[i] != expected:
                print(f"❌ Header mismatch in CONTROLLER sheet column {i+1}")
                print(f"   Expected: '{expected}', Found: '{actual_headers[i]}'")
                return False
        
        print("✅ CONTROLLER sheet structure is valid")
        return True
    
    def load_controller_data(self) -> bool:
        """Load controller data from CONTROLLER sheet"""
        if not self.validate_controller_sheet():
            return False
        
        ws = self.workbook["CONTROLLER"]
        self.sheet_controllers = []
        
        row_num = 2
        while True:
            # Check if we've reached the end
            sheet_name = ws.cell(row=row_num, column=2).value  # Sheet_Name column
            if not sheet_name:
                break
            
            try:
                enable_value = ws.cell(row=row_num, column=1).value  # Enable column
                description = ws.cell(row=row_num, column=3).value  # Description column
                priority = ws.cell(row=row_num, column=4).value  # Priority column
                
                # Convert enable value to boolean
                enable = self._convert_bool(enable_value)
                
                # Validate sheet exists
                sheet_name_str = str(sheet_name).strip()
                if sheet_name_str not in self.workbook.sheetnames:
                    print(f"⚠️  Warning: Sheet '{sheet_name_str}' referenced in CONTROLLER but not found in workbook")
                    print(f"   Available sheets: {', '.join(self.workbook.sheetnames)}")
                
                controller = SheetController(
                    enable=enable,
                    sheet_name=sheet_name_str,
                    description=str(description) if description else None,
                    priority=str(priority) if priority else None
                )
                
                # Get test counts if sheet exists and is enabled
                if controller.enable and sheet_name_str in self.workbook.sheetnames:
                    test_counts = self._get_sheet_test_counts(sheet_name_str)
                    controller.test_count = test_counts[0]
                    controller.enabled_test_count = test_counts[1]
                
                self.sheet_controllers.append(controller)
                
            except Exception as e:
                print(f"❌ Error processing CONTROLLER row {row_num}: {e}")
                return False
            
            row_num += 1
        
        return True
    
    def _convert_bool(self, value) -> bool:
        """Convert various boolean representations to bool"""
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        if isinstance(value, str):
            return value.upper() in ['TRUE', 'YES', 'Y', '1']
        return False
    
    def _get_sheet_test_counts(self, sheet_name: str) -> Tuple[int, int]:
        """Get total and enabled test counts for a sheet"""
        try:
            reader = ExcelTestSuiteReader(str(self.excel_file))
            reader.workbook = self.workbook
            
            # Read test cases from the specific sheet
            test_cases = reader.read_test_cases(sheet_name)
            
            total_count = len(test_cases)
            enabled_count = sum(1 for tc in test_cases if tc.enable)
            
            return total_count, enabled_count
        except Exception:
            return 0, 0
    
    def get_enabled_sheets(self) -> List[SheetController]:
        """Get list of enabled sheets from controller"""
        return [controller for controller in self.sheet_controllers if controller.enable]
    
    def get_disabled_sheets(self) -> List[SheetController]:
        """Get list of disabled sheets from controller"""
        return [controller for controller in self.sheet_controllers if not controller.enable]
    
    def print_controller_summary(self):
        """Print summary of controller configuration"""
        enabled_sheets = self.get_enabled_sheets()
        disabled_sheets = self.get_disabled_sheets()
        
        print("📋 CONTROLLER SHEET SUMMARY")
        print("=" * 50)
        print(f"📊 Total sheets configured: {len(self.sheet_controllers)}")
        print(f"✅ Enabled sheets: {len(enabled_sheets)}")
        print(f"❌ Disabled sheets: {len(disabled_sheets)}")
        
        if enabled_sheets:
            print(f"\n🚀 ENABLED SHEETS (will be executed):")
            print("-" * 40)
            total_tests = 0
            total_enabled_tests = 0
            
            for controller in enabled_sheets:
                status = "✅" if controller.sheet_name in self.workbook.sheetnames else "❌ Missing"
                priority_str = f" [{controller.priority}]" if controller.priority else ""
                test_info = f" ({controller.enabled_test_count}/{controller.test_count} tests)" if controller.test_count > 0 else ""
                
                print(f"   {status} {controller.sheet_name}{priority_str}{test_info}")
                if controller.description:
                    print(f"      📝 {controller.description}")
                
                total_tests += controller.test_count
                total_enabled_tests += controller.enabled_test_count
            
            print(f"\n📊 Total tests to execute: {total_enabled_tests} (from {total_tests} total)")
        
        if disabled_sheets:
            print(f"\n⏭️  DISABLED SHEETS (will be skipped):")
            print("-" * 40)
            for controller in disabled_sheets:
                status = "📄" if controller.sheet_name in self.workbook.sheetnames else "❌ Missing"
                priority_str = f" [{controller.priority}]" if controller.priority else ""
                
                print(f"   {status} {controller.sheet_name}{priority_str}")
                if controller.description:
                    print(f"      📝 {controller.description}")
    
    def execute_controlled_tests(self, **filters) -> Dict[str, List[TestResult]]:
        """
        Execute tests from all enabled sheets
        
        Args:
            **filters: Additional filters to apply (priority, category, etc.)
            
        Returns:
            Dictionary mapping sheet names to their test results
        """
        if not self.load_controller_data():
            print("❌ Failed to load controller data")
            return {}
        
        self.print_controller_summary()
        
        enabled_sheets = self.get_enabled_sheets()
        if not enabled_sheets:
            print("\n⚠️  No sheets are enabled in CONTROLLER. Nothing to execute.")
            return {}
        
        print(f"\n🚀 Starting multi-sheet test execution")
        print("=" * 60)
        
        all_results = {}
        executor = TestExecutor()
        
        for sheet_idx, controller in enumerate(enabled_sheets, 1):
            if controller.sheet_name not in self.workbook.sheetnames:
                print(f"\n[{sheet_idx}/{len(enabled_sheets)}] ❌ Skipping '{controller.sheet_name}' - Sheet not found")
                continue
            
            print(f"\n[{sheet_idx}/{len(enabled_sheets)}] 📊 Executing sheet: {controller.sheet_name}")
            if controller.description:
                print(f"📝 Description: {controller.description}")
            
            try:
                # Read test cases from this sheet
                reader = ExcelTestSuiteReader(str(self.excel_file))
                reader.workbook = self.workbook
                test_cases = reader.read_test_cases(controller.sheet_name)
                
                # Apply filters
                filtered_cases = self._apply_filters(test_cases, **filters)
                
                if not filtered_cases:
                    print(f"   ⚠️  No tests match the filters in sheet '{controller.sheet_name}'")
                    all_results[controller.sheet_name] = []
                    continue
                
                print(f"   📊 Tests to execute: {len(filtered_cases)}")
                
                # Execute tests
                sheet_results = []
                for test_idx, test_case in enumerate(filtered_cases, 1):
                    print(f"\n   [{test_idx}/{len(filtered_cases)}] 🧪 Executing: {test_case.test_case_id}")
                    
                    result = executor.execute_test_case(test_case)
                    sheet_results.append(result)
                    
                    # Print result
                    status_icon = "✅" if result.status == "PASS" else "❌" if result.status == "FAIL" else "⏭️"
                    print(f"      {status_icon} {result.status} ({result.duration_seconds:.2f}s)")
                    if result.error_message:
                        print(f"      💬 {result.error_message}")
                
                all_results[controller.sheet_name] = sheet_results
                
                # Sheet summary
                passed = sum(1 for r in sheet_results if r.status == "PASS")
                failed = sum(1 for r in sheet_results if r.status == "FAIL")
                skipped = sum(1 for r in sheet_results if r.status in ["SKIP", "SKIPPED"])
                
                print(f"\n   📋 Sheet '{controller.sheet_name}' Summary:")
                print(f"      ✅ Passed: {passed}")
                print(f"      ❌ Failed: {failed}")
                print(f"      ⏭️  Skipped: {skipped}")
                print(f"      📈 Success Rate: {(passed / len(sheet_results) * 100):.1f}%")
                
            except Exception as e:
                print(f"   ❌ Error executing sheet '{controller.sheet_name}': {e}")
                all_results[controller.sheet_name] = []
        
        # Overall summary
        self._print_overall_summary(all_results)
        
        return all_results
    
    def _apply_filters(self, test_cases: List[TestCase], **filters) -> List[TestCase]:
        """Apply filters to test cases"""
        filtered_cases = [tc for tc in test_cases if tc.enable]  # Only enabled tests
        
        if filters.get('priority'):
            filtered_cases = [tc for tc in filtered_cases if tc.priority.upper() == filters['priority'].upper()]
        
        if filters.get('category'):
            filtered_cases = [tc for tc in filtered_cases if tc.test_category.upper() == filters['category'].upper()]
        
        if filters.get('environment'):
            filtered_cases = [tc for tc in filtered_cases if tc.environment_name.upper() == filters['environment'].upper()]
        
        if filters.get('application'):
            filtered_cases = [tc for tc in filtered_cases if tc.application_name.upper() == filters['application'].upper()]
        
        if filters.get('test_ids'):
            test_ids = [tid.strip() for tid in filters['test_ids'].split(',')]
            filtered_cases = [tc for tc in filtered_cases if tc.test_case_id in test_ids]
        
        return filtered_cases
    
    def _print_overall_summary(self, all_results: Dict[str, List[TestResult]]):
        """Print overall execution summary"""
        print(f"\n{'='*80}")
        print(f"📋 OVERALL MULTI-SHEET EXECUTION SUMMARY")
        print(f"{'='*80}")
        
        total_tests = sum(len(results) for results in all_results.values())
        total_passed = sum(1 for results in all_results.values() for r in results if r.status == "PASS")
        total_failed = sum(1 for results in all_results.values() for r in results if r.status == "FAIL")
        total_skipped = sum(1 for results in all_results.values() for r in results if r.status in ["SKIP", "SKIPPED"])
        
        total_duration = sum(r.duration_seconds for results in all_results.values() for r in results)
        
        print(f"📊 Sheets executed: {len([k for k, v in all_results.items() if v])}")
        print(f"📊 Total tests: {total_tests}")
        print(f"✅ Passed: {total_passed}")
        print(f"❌ Failed: {total_failed}")
        print(f"⏭️  Skipped: {total_skipped}")
        print(f"⏱️  Total duration: {total_duration:.2f}s")
        if total_tests > 0:
            print(f"📈 Overall success rate: {(total_passed / total_tests * 100):.1f}%")
        
        # Per-sheet breakdown
        print(f"\n📋 PER-SHEET BREAKDOWN:")
        print("-" * 50)
        for sheet_name, results in all_results.items():
            if results:
                passed = sum(1 for r in results if r.status == "PASS")
                failed = sum(1 for r in results if r.status == "FAIL")
                skipped = sum(1 for r in results if r.status in ["SKIP", "SKIPPED"])
                duration = sum(r.duration_seconds for r in results)
                success_rate = (passed / len(results) * 100) if results else 0
                
                print(f"📄 {sheet_name}:")
                print(f"   Tests: {len(results)} | Passed: {passed} | Failed: {failed} | Skipped: {skipped}")
                print(f"   Duration: {duration:.2f}s | Success: {success_rate:.1f}%")
            else:
                print(f"📄 {sheet_name}: No tests executed")


# Add method to ExcelTestSuiteReader to read from specific sheet
def read_test_cases_from_sheet(self, sheet_name: str = "SMOKE") -> List[TestCase]:
    """Read test cases from a specific worksheet"""
    if not self.workbook:
        if not self.load_workbook():
            return []
    
    if sheet_name not in self.workbook.sheetnames:
        print(f"❌ Worksheet '{sheet_name}' not found")
        return []
    
    ws = self.workbook[sheet_name]
    
    # Run validation if validator is available
    if hasattr(self, 'validator') and self.validator:
        is_valid, validation_messages = self.validator.validate_test_suite(self.workbook, sheet_name)
        if not is_valid:
            errors = [msg for msg in validation_messages if msg.severity == ValidationSeverity.ERROR]
            print(f"❌ Validation failed for sheet '{sheet_name}': {len(errors)} errors found")
            return []
    
    self.test_cases = []
    
    # Get headers mapping
    headers = {}
    for col in range(1, 13):  # 12 columns expected
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            headers[col] = str(cell_value).strip()
    
    # Read data rows
    row_num = 2
    while True:
        # Check if we've reached the end (empty Test_Case_ID)
        test_id_cell = None
        for col, header in headers.items():
            if header == "Test_Case_ID":
                test_id_cell = ws.cell(row=row_num, column=col).value
                break
        
        if not test_id_cell:
            break
        
        try:
            # Read row data
            row_data = {}
            for col, header in headers.items():
                cell_value = ws.cell(row=row_num, column=col).value
                row_data[header] = cell_value
            
            # Convert and validate data
            test_case = TestCase(
                enable=self._convert_bool(row_data.get("Enable", False)),
                test_case_id=str(row_data.get("Test_Case_ID", "")),
                test_case_name=str(row_data.get("Test_Case_Name", "")),
                application_name=str(row_data.get("Application_Name", "")),
                environment_name=str(row_data.get("Environment_Name", "")),
                priority=str(row_data.get("Priority", "MEDIUM")),
                test_category=str(row_data.get("Test_Category", "")),
                expected_result=str(row_data.get("Expected_Result", "PASS")),
                timeout_seconds=self._convert_int(row_data.get("Timeout_Seconds", 60)),
                description=str(row_data.get("Description", "")),
                prerequisites=str(row_data.get("Prerequisites", "")),
                tags=str(row_data.get("Tags", "")),
            )
            
            # Basic validation
            if test_case.test_case_id and test_case.test_case_name:
                self.test_cases.append(test_case)
            
        except Exception as e:
            print(f"❌ Error processing row {row_num} in sheet '{sheet_name}': {e}")
        
        row_num += 1
    
    return self.test_cases


# Monkey patch the method to ExcelTestSuiteReader
ExcelTestSuiteReader.read_test_cases = read_test_cases_from_sheet