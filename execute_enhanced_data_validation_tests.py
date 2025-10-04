#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Enhanced Data Validation Test Execution Script
===============================================
Executes data validation tests from the enhanced Excel test suite with proper table parameter support.
This version bypasses header validation issues and directly reads the test data.

Author: Multi-Database Data Validation Framework
Date: October 4, 2025
"""

import sys
import os
import time
from pathlib import Path

# Set UTF-8 encoding for Windows console output
if os.name == 'nt':  # Windows
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
from datetime import datetime
from typing import List, Dict, Any
from dataclasses import dataclass
from openpyxl import load_workbook

# Add src directory to path for imports
sys.path.insert(0, str(Path(__file__).parent / "src"))

from src.core.test_executor import TestExecutor


@dataclass
class SimpleTestCase:
    """Simple test case data structure"""
    enable: bool
    test_case_id: str
    test_case_name: str
    application_name: str
    environment_name: str
    priority: str
    test_category: str
    expected_result: str
    timeout_seconds: int
    description: str
    prerequisites: str
    tags: str
    parameters: str


def load_test_cases_from_excel(excel_file: str, sheet_name: str = "DATAVALIDATIONS") -> List[SimpleTestCase]:
    """Load test cases directly from Excel without strict validation"""
    test_cases = []
    
    try:
        workbook = load_workbook(excel_file)
        
        if sheet_name not in workbook.sheetnames:
            print(f"❌ Sheet '{sheet_name}' not found in workbook")
            print(f"   Available sheets: {', '.join(workbook.sheetnames)}")
            return []
        
        ws = workbook[sheet_name]
        
        # Read data rows (skip header row)
        for row in range(2, ws.max_row + 1):
            # Check if row has data
            if not ws.cell(row=row, column=2).value:  # TEST_CASE_ID column
                continue
                
            try:
                test_case = SimpleTestCase(
                    enable=str(ws.cell(row=row, column=1).value or "").upper() in ["TRUE", "YES", "1"],
                    test_case_id=str(ws.cell(row=row, column=2).value or ""),
                    test_case_name=str(ws.cell(row=row, column=3).value or ""),
                    application_name=str(ws.cell(row=row, column=4).value or ""),
                    environment_name=str(ws.cell(row=row, column=5).value or ""),
                    priority=str(ws.cell(row=row, column=6).value or ""),
                    test_category=str(ws.cell(row=row, column=7).value or ""),
                    expected_result=str(ws.cell(row=row, column=8).value or "PASS"),
                    timeout_seconds=int(ws.cell(row=row, column=9).value or 60),
                    description=str(ws.cell(row=row, column=10).value or ""),
                    prerequisites=str(ws.cell(row=row, column=11).value or ""),
                    tags=str(ws.cell(row=row, column=12).value or ""),
                    parameters=str(ws.cell(row=row, column=13).value or "")
                )
                
                if test_case.enable:
                    test_cases.append(test_case)
                    
            except Exception as e:
                print(f"⚠️  Warning: Error reading row {row}: {e}")
                continue
    
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return []
    
    return test_cases


def execute_enhanced_data_validation_tests(excel_file: str):
    """Execute data validation tests with enhanced table parameter support"""
    
    print("🔍 EXECUTING ENHANCED DATA VALIDATION TESTS")
    print("=" * 64)
    print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Test Suite: {excel_file}")
    print(f"Target Sheet: DATAVALIDATIONS")
    print()
    
    # Load test cases
    test_cases = load_test_cases_from_excel(excel_file)
    
    if not test_cases:
        print("❌ No test cases loaded from Excel file")
        return
    
    # Filter for data validation tests
    data_validation_tests = [tc for tc in test_cases if tc.test_category in [
        'SCHEMA_VALIDATION', 'ROW_COUNT_VALIDATION', 
        'NULL_VALUE_VALIDATION'
    ]]
    
    print(f"📊 Successfully loaded {len(data_validation_tests)} data validation test cases")
    print()
    
    # Initialize test executor
    executor = TestExecutor()
    
    # Execute tests
    results = []
    passed = 0
    failed = 0
    
    for i, test_case in enumerate(data_validation_tests, 1):
        print(f"[{i}/{len(data_validation_tests)}] 🧪 Executing: {test_case.test_case_id}")
        print(f"   Name: {test_case.test_case_name}")
        print(f"   Category: {test_case.test_category}")
        print(f"   Parameters: {test_case.parameters}")
        
        # Parse parameters
        params = {}
        if test_case.parameters:
            for param in test_case.parameters.split(';'):
                if '=' in param:
                    key, value = param.strip().split('=', 1)
                    params[key.strip()] = value.strip()
        
        source_table = params.get('source_table', 'products')
        target_table = params.get('target_table', 'new_products')
        
        print(f"   Source Table: {source_table}")
        print(f"   Target Table: {target_table}")
        
        start_time = time.time()
        
        try:
            # Execute based on category
            if test_case.test_category == "SCHEMA_VALIDATION":
                result = executor.data_validator.schema_validation_compare(source_table, target_table)
            elif test_case.test_category == "ROW_COUNT_VALIDATION":
                result = executor.data_validator.row_count_validation_compare(source_table, target_table)
            elif test_case.test_category == "NULL_VALUE_VALIDATION":
                result = executor.data_validator.null_value_validation_compare(source_table, target_table)
            else:
                result = type('Result', (), {'passed': False, 'message': f'Unknown category: {test_case.test_category}'})()
            
            duration = time.time() - start_time
            
            if result.passed:
                print(f"   ✅ PASS ({duration:.3f}s)")
                if test_case.test_category == "SCHEMA_VALIDATION" and hasattr(result, 'details'):
                    print(f"   📊 Schema Match: {result.details.get('source_columns', 0)} columns validated")
                passed += 1
                status = "PASS"
            else:
                print(f"   ❌ FAIL ({duration:.3f}s)")
                print(f"   💬 {result.message}")
                
                # Display detailed schema comparison for SCHEMA_VALIDATION
                if test_case.test_category == "SCHEMA_VALIDATION" and hasattr(result, 'details'):
                    details = result.details
                    if 'detailed_report' in details:
                        print(f"   ")
                        print(f"   📋 DETAILED SCHEMA COMPARISON:")
                        print(f"   {'='*60}")
                        print(f"   SOURCE TABLE: {details.get('source_table', source_table)}")
                        print(f"   TARGET TABLE: {details.get('target_table', target_table)}")
                        print(f"   ")
                        
                        for i, diff in enumerate(details['detailed_report'], 1):
                            print(f"   [{i}] COLUMN: {diff['column']}")
                            print(f"       ISSUE: {diff['issue']}")
                            print(f"       SOURCE: {diff['source_type']}")
                            print(f"       TARGET: {diff['target_type']}")
                            print(f"       DESC: {diff['description']}")
                            print(f"   ")
                        
                        print(f"   SUMMARY: {len(details['detailed_report'])} schema difference(s) found")
                        print(f"   {'='*60}")
                
                # Display detailed NULL value comparison for NULL_VALUE_VALIDATION
                elif test_case.test_category == "NULL_VALUE_VALIDATION" and hasattr(result, 'details'):
                    details = result.details
                    if 'null_differences' in details:
                        print(f"   ")
                        print(f"   📋 DETAILED NULL VALUE COMPARISON:")
                        print(f"   {'='*60}")
                        print(f"   SOURCE TABLE: {details.get('source_table', source_table)} ({details.get('source_total_rows', 0):,} rows)")
                        print(f"   TARGET TABLE: {details.get('target_table', target_table)} ({details.get('target_total_rows', 0):,} rows)")
                        print(f"   ")
                        
                        for i, diff in enumerate(details['null_differences'], 1):
                            issue_icon = "🚨" if diff['issue_type'] == "CONSTRAINT_VIOLATION" else "⚠️"
                            print(f"   [{i}] {issue_icon} COLUMN: {diff['column']} ({diff['data_type']})")
                            print(f"       ISSUE: {diff['issue_type']}")
                            print(f"       CONSTRAINT: SRC {'NOT NULL' if not diff['source_nullable'] else 'NULLABLE'} | TGT {'NOT NULL' if not diff['target_nullable'] else 'NULLABLE'}")
                            print(f"       NULL COUNT: SRC {diff['source_nulls']:,} ({diff['source_null_percentage']}%) | TGT {diff['target_nulls']:,} ({diff['target_null_percentage']}%)")
                            print(f"       DIFFERENCE: {diff['difference']:,} null value(s)")
                            if diff['issue_type'] == "CONSTRAINT_VIOLATION":
                                print(f"       🚨 CRITICAL: NOT NULL constraint violated!")
                            print(f"   ")
                        
                        print(f"   SUMMARY: {len(details['null_differences'])} column(s) with NULL value issues")
                        print(f"   {'='*60}")
                
                # Display detailed data quality comparison for DATA_QUALITY_VALIDATION
                elif test_case.test_category == "DATA_QUALITY_VALIDATION" and hasattr(result, 'details'):
                    details = result.details
                    if 'quality_issues' in details:
                        print(f"   ")
                        print(f"   📋 DETAILED DATA QUALITY ANALYSIS:")
                        print(f"   {'='*60}")
                        print(f"   SOURCE TABLE: {details.get('source_table', source_table)} ({details.get('source_total_rows', 0):,} rows)")
                        print(f"   TARGET TABLE: {details.get('target_table', target_table)} ({details.get('target_total_rows', 0):,} rows)")
                        print(f"   ")
                        
                        for i, issue in enumerate(details['quality_issues'], 1):
                            severity_icon = "🚨" if issue['severity'] == "HIGH" else "⚠️" if issue['severity'] == "MEDIUM" else "ℹ️"
                            print(f"   [{i}] {severity_icon} ISSUE: {issue['issue_type']} ({issue['severity']} SEVERITY)")
                            print(f"       TABLE: {issue['table']}")
                            
                            if issue['issue_type'] == "DUPLICATE_RECORDS":
                                print(f"       COLUMN: {issue['column']}")
                                print(f"       AFFECTED: {issue['affected_values']} unique values with duplicates")
                                print(f"       EXTRA ROWS: {issue['total_duplicate_rows']} duplicate rows found")
                                if issue['sample_duplicates']:
                                    print(f"       SAMPLES: {', '.join([f'{dup[0]}({dup[1]}x)' for dup in issue['sample_duplicates'][:3]])}")
                            
                            elif issue['issue_type'] == "ORPHANED_RECORDS":
                                print(f"       FOREIGN KEY: {issue['foreign_key']} → {issue['reference_table']}")
                                print(f"       AFFECTED: {issue['affected_records']} records with invalid references")
                                print(f"       INVALID VALUES: {issue['unique_invalid_values']} unique invalid IDs")
                                if issue['sample_invalid_ids']:
                                    print(f"       SAMPLES: {', '.join([f'ID {inv[0]}({inv[1]} records)' for inv in issue['sample_invalid_ids'][:3]])}")
                            
                            elif issue['issue_type'] == "INVALID_DATA_VALUES":
                                print(f"       COLUMN: {issue['column']}")
                                print(f"       AFFECTED: {issue['affected_records']} records ({issue['percentage']}%)")
                                print(f"       RULE: {issue['validation_rule']}")
                            
                            elif issue['issue_type'] == "MISSING_CRITICAL_DATA":
                                print(f"       COLUMN: {issue['column']}")
                                print(f"       AFFECTED: {issue['affected_records']} records ({issue['percentage']}%)")
                                print(f"       IMPACT: {issue['impact']}")
                            
                            print(f"       DESC: {issue['description']}")
                            print(f"   ")
                        
                        total_issues = details.get('total_issues', 0)
                        high_severity = details.get('high_severity_issues', 0)
                        print(f"   SUMMARY: {total_issues} data quality issue(s) found ({high_severity} high severity)")
                        print(f"   CHECKS: {', '.join(details.get('checks_performed', []))}")
                        print(f"   {'='*60}")
                
                failed += 1
                status = "FAIL"
            
            results.append({
                'test_id': test_case.test_case_id,
                'test_name': test_case.test_case_name,
                'category': test_case.test_category,
                'source_table': source_table,
                'target_table': target_table,
                'status': status,
                'duration': duration,
                'message': result.message if hasattr(result, 'message') else ""
            })
        
        except Exception as e:
            duration = time.time() - start_time
            print(f"   ❌ ERROR ({duration:.3f}s)")
            print(f"   💬 {str(e)}")
            failed += 1
            
            results.append({
                'test_id': test_case.test_case_id,
                'test_name': test_case.test_case_name,
                'category': test_case.test_category,
                'source_table': source_table,
                'target_table': target_table,
                'status': "ERROR",
                'duration': duration,
                'message': str(e)
            })
        
        print()
    
    # Print summary
    total_tests = len(data_validation_tests)
    success_rate = (passed / total_tests * 100) if total_tests > 0 else 0
    total_duration = sum(r['duration'] for r in results)
    
    print("📋 ENHANCED DATA VALIDATION TEST SUMMARY:")
    print("=" * 64)
    print(f"   ✅ Passed: {passed}")
    print(f"   ❌ Failed: {failed}")
    print(f"   📈 Success Rate: {success_rate:.1f}%")
    print(f"   ⏱️  Total duration: {total_duration:.2f}s")
    print()
    
    # Detailed results
    print("📋 DETAILED TEST RESULTS:")
    print("-" * 40)
    print()
    
    if passed > 0:
        print("✅ PASSED TESTS:")
        for i, result in enumerate([r for r in results if r['status'] == 'PASS'], 1):
            print(f"    {i}. {result['test_id']} - {result['test_name']}")
            print(f"       Tables: {result['source_table']} → {result['target_table']}")
            print(f"       Category: {result['category']} | Duration: {result['duration']:.3f}s")
        print()
    
    if failed > 0:
        print("❌ FAILED TESTS:")
        for i, result in enumerate([r for r in results if r['status'] in ['FAIL', 'ERROR']], 1):
            print(f"    {i}. {result['test_id']} - {result['test_name']}")
            print(f"       Tables: {result['source_table']} → {result['target_table']}")
            print(f"       Category: {result['category']} | Duration: {result['duration']:.3f}s")
            print(f"       Error: {result['message']}")
        print()
    
    print("✅ ENHANCED DATA VALIDATION TEST EXECUTION COMPLETE!")
    return success_rate > 0


def main():
    """Main execution function"""
    excel_file = sys.argv[1] if len(sys.argv) > 1 else "enhanced_sdm_test_suite.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return False
    
    return execute_enhanced_data_validation_tests(excel_file)


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)