#!/usr/bin/env python3
"""
Enhanced Excel Test Suite Template Generator
============================================
Generates a comprehensive Excel test suite with proper table configurations
and enhanced validation capabilities for data validation testing.

Author: Multi-Database Data Validation Framework
Date: October 4, 2025
"""

import sys
import os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Add src directory to path for imports
sys.path.insert(0, str(Path(__file__).parent / "src"))


class EnhancedExcelTemplateGenerator:
    """Enhanced Excel template generator with proper table configurations"""
    
    def __init__(self):
        self.workbook = Workbook()
        self.setup_styles()
        
    def setup_styles(self):
        """Set up Excel styles"""
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.data_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        
    def create_controller_sheet(self):
        """Create enhanced CONTROLLER sheet with validation rules"""
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
            
        ws = self.workbook.create_sheet("CONTROLLER", 0)
        
        # Headers
        headers = [
            "SHEET_NAME", "ENABLED", "PRIORITY", "DESCRIPTION", 
            "TEST_COUNT", "VALIDATION_STATUS", "EXECUTION_ORDER"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Data rows with proper configurations
        sheet_configs = [
            ["SMOKE", "YES", "HIGH", "PostgreSQL smoke tests - basic connectivity and functionality", 30, "VALID", 1],
            ["DATAVALIDATIONS", "YES", "HIGH", "Enhanced data validation tests with proper table mappings", 12, "VALID", 2],
            ["INTEGRATION", "NO", "MEDIUM", "Integration tests with external systems", 15, "MISSING", 3],
            ["PERFORMANCE", "NO", "LOW", "Performance and load testing suite", 20, "MISSING", 4],
            ["SECURITY", "NO", "HIGH", "Security and penetration testing", 10, "MISSING", 5],
            ["REGRESSION", "NO", "MEDIUM", "Full regression test suite", 50, "MISSING", 6]
        ]
        
        for row_idx, config in enumerate(sheet_configs, 2):
            for col_idx, value in enumerate(config, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                cell.font = self.data_font
                if col_idx in [2, 6]:  # ENABLED and VALIDATION_STATUS columns
                    cell.alignment = self.center_alignment
        
        # Add data validation for ENABLED column
        enabled_validation = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
        enabled_validation.error = "Please select YES or NO"
        enabled_validation.errorTitle = "Invalid Entry"
        ws.add_data_validation(enabled_validation)
        enabled_validation.add(f"B2:B{len(sheet_configs) + 1}")
        
        # Add data validation for PRIORITY column
        priority_validation = DataValidation(type="list", formula1='"HIGH,MEDIUM,LOW"', allow_blank=False)
        priority_validation.error = "Please select HIGH, MEDIUM, or LOW"
        priority_validation.errorTitle = "Invalid Priority"
        ws.add_data_validation(priority_validation)
        priority_validation.add(f"C2:C{len(sheet_configs) + 1}")
        
        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['D'].width = 50  # Description column wider
        
        return ws
    
    def create_enhanced_datavalidations_sheet(self):
        """Create enhanced DATAVALIDATIONS sheet with proper table configurations"""
        ws = self.workbook.create_sheet("DATAVALIDATIONS")
        
        # Headers
        headers = [
            "ENABLE", "TEST_CASE_ID", "TEST_CASE_NAME", "APPLICATION_NAME",
            "ENVIRONMENT_NAME", "PRIORITY", "TEST_CATEGORY", "EXPECTED_RESULT",
            "TIMEOUT_SECONDS", "DESCRIPTION", "PREREQUISITES", "TAGS", "PARAMETERS"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Enhanced test cases with proper table configurations
        test_cases = [
            # Schema Validation Tests
            [
                "YES", "DVAL_001", "Products Schema Validation", "POSTGRES", "DEV", "HIGH",
                "SCHEMA_VALIDATION", "PASS", 60,
                "Compare schema structure between source products and target new_products tables",
                "Both tables must exist with data", "schema,validation,products",
                "source_table=products;target_table=new_products;ignore_sequence=true;check_constraints=true"
            ],
            [
                "YES", "DVAL_002", "Employees Schema Validation", "POSTGRES", "DEV", "HIGH",
                "SCHEMA_VALIDATION", "PASS", 60,
                "Compare schema structure between source employees and target new_employees tables",
                "Both tables must exist with data", "schema,validation,employees",
                "source_table=employees;target_table=new_employees;ignore_sequence=true;check_constraints=true"
            ],
            [
                "YES", "DVAL_003", "Orders Schema Validation", "POSTGRES", "DEV", "HIGH",
                "SCHEMA_VALIDATION", "PASS", 60,
                "Compare schema structure between source orders and target new_orders tables",
                "Both tables must exist with data", "schema,validation,orders",
                "source_table=orders;target_table=new_orders;ignore_sequence=true;check_constraints=true"
            ],
            
            # Row Count Validation Tests
            [
                "YES", "DVAL_004", "Products Row Count Validation", "POSTGRES", "DEV", "HIGH",
                "ROW_COUNT_VALIDATION", "FAIL", 30,
                "Compare row counts between source products and target new_products tables",
                "Both tables must exist", "rowcount,validation,products",
                "source_table=products;target_table=new_products;tolerance_percent=5;expected_variance=large"
            ],
            [
                "YES", "DVAL_005", "Employees Row Count Validation", "POSTGRES", "DEV", "HIGH",
                "ROW_COUNT_VALIDATION", "FAIL", 30,
                "Compare row counts between source employees and target new_employees tables",
                "Both tables must exist", "rowcount,validation,employees",
                "source_table=employees;target_table=new_employees;tolerance_percent=5;expected_variance=large"
            ],
            [
                "YES", "DVAL_006", "Orders Row Count Validation", "POSTGRES", "DEV", "HIGH",
                "ROW_COUNT_VALIDATION", "FAIL", 30,
                "Compare row counts between source orders and target new_orders tables",
                "Both tables must exist", "rowcount,validation,orders",
                "source_table=orders;target_table=new_orders;tolerance_percent=5;expected_variance=large"
            ],
            
            # NULL Value Validation Tests
            [
                "YES", "DVAL_007", "Products NULL Value Validation", "POSTGRES", "DEV", "MEDIUM",
                "NULL_VALUE_VALIDATION", "FAIL", 45,
                "Compare NULL patterns between source products and target new_products tables",
                "Both tables must exist with data", "null,validation,products",
                "source_table=products;target_table=new_products;null_match_required=false;report_differences=true"
            ],
            [
                "YES", "DVAL_008", "Employees NULL Value Validation", "POSTGRES", "DEV", "MEDIUM",
                "NULL_VALUE_VALIDATION", "FAIL", 45,
                "Compare NULL patterns between source employees and target new_employees tables",
                "Both tables must exist with data", "null,validation,employees",
                "source_table=employees;target_table=new_employees;null_match_required=false;report_differences=true"
            ],
            [
                "YES", "DVAL_009", "Orders NULL Value Validation", "POSTGRES", "DEV", "MEDIUM",
                "NULL_VALUE_VALIDATION", "FAIL", 45,
                "Compare NULL patterns between source orders and target new_orders tables",
                "Both tables must exist with data", "null,validation,orders",
                "source_table=orders;target_table=new_orders;null_match_required=false;report_differences=true"
            ],
            
            # Data Quality Validation Tests
            [
                "YES", "DVAL_010", "Products Data Quality Check", "POSTGRES", "DEV", "MEDIUM",
                "DATA_QUALITY_VALIDATION", "PASS", 60,
                "Validate data quality metrics for products table",
                "Products table must exist with data", "quality,validation,products",
                "source_table=products;target_table=new_products;check_ranges=true;check_patterns=true;check_duplicates=true"
            ],
            [
                "YES", "DVAL_011", "Cross-Table Referential Integrity", "POSTGRES", "DEV", "HIGH",
                "DATA_QUALITY_VALIDATION", "PASS", 90,
                "Validate referential integrity between products, orders, and employees",
                "All tables must exist with proper relationships", "integrity,validation,cross-table",
                "source_table=products;target_table=orders;check_foreign_keys=true;validate_relationships=true"
            ],
            [
                "YES", "DVAL_012", "Data Completeness Validation", "POSTGRES", "DEV", "MEDIUM",
                "DATA_QUALITY_VALIDATION", "PASS", 45,
                "Validate data completeness across all source tables",
                "All source tables must exist", "completeness,validation,all-tables",
                "source_table=products,employees,orders;check_required_fields=true;missing_data_threshold=5"
            ]
        ]
        
        # Add test cases to worksheet
        for row_idx, test_case in enumerate(test_cases, 2):
            for col_idx, value in enumerate(test_case, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                cell.font = self.data_font
                if col_idx in [1, 7]:  # ENABLE and EXPECTED_RESULT columns
                    cell.alignment = self.center_alignment
        
        # Add data validations
        enable_validation = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
        ws.add_data_validation(enable_validation)
        enable_validation.add(f"A2:A{len(test_cases) + 1}")
        
        priority_validation = DataValidation(type="list", formula1='"HIGH,MEDIUM,LOW"', allow_blank=False)
        ws.add_data_validation(priority_validation)
        priority_validation.add(f"F2:F{len(test_cases) + 1}")
        
        category_validation = DataValidation(type="list", 
            formula1='"SCHEMA_VALIDATION,ROW_COUNT_VALIDATION,NULL_VALUE_VALIDATION,DATA_QUALITY_VALIDATION"', 
            allow_blank=False)
        ws.add_data_validation(category_validation)
        category_validation.add(f"G2:G{len(test_cases) + 1}")
        
        expected_validation = DataValidation(type="list", formula1='"PASS,FAIL"', allow_blank=False)
        ws.add_data_validation(expected_validation)
        expected_validation.add(f"H2:H{len(test_cases) + 1}")
        
        # Auto-fit columns
        column_widths = [8, 12, 25, 12, 12, 10, 20, 12, 12, 50, 30, 20, 60]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        return ws
    
    def create_instructions_sheet(self):
        """Create enhanced INSTRUCTIONS sheet"""
        ws = self.workbook.create_sheet("INSTRUCTIONS")
        
        instructions = [
            ["📋 ENHANCED DATA VALIDATION TEST SUITE INSTRUCTIONS", ""],
            ["", ""],
            ["🎯 PURPOSE:", ""],
            ["This Excel file contains enhanced test cases for comprehensive data validation", ""],
            ["testing with proper source/target table configurations.", ""],
            ["", ""],
            ["📊 SHEET DESCRIPTIONS:", ""],
            ["", ""],
            ["🔧 CONTROLLER:", ""],
            ["- Controls which test sheets are executed", ""],
            ["- Set ENABLED to YES/NO for each sheet", ""],
            ["- PRIORITY determines execution order (HIGH > MEDIUM > LOW)", ""],
            ["- VALIDATION_STATUS shows sheet readiness", ""],
            ["", ""],
            ["🔍 DATAVALIDATIONS:", ""],
            ["- Enhanced data validation tests with proper table mappings", ""],
            ["- Each test specifies source_table and target_table parameters", ""],
            ["- Includes schema, row count, NULL value, and data quality validations", ""],
            ["- Expected results reflect realistic test scenarios", ""],
            ["", ""],
            ["📋 PARAMETER FORMATS:", ""],
            ["", ""],
            ["Schema Validation:", "source_table=products;target_table=new_products;ignore_sequence=true"],
            ["Row Count Validation:", "source_table=employees;target_table=new_employees;tolerance_percent=5"],
            ["NULL Value Validation:", "source_table=orders;target_table=new_orders;null_match_required=false"],
            ["Data Quality Validation:", "source_table=products;check_ranges=true;check_patterns=true"],
            ["", ""],
            ["🎯 TABLE MAPPINGS:", ""],
            ["", ""],
            ["Source Tables → Target Tables:", ""],
            ["products (1200 rows) → new_products (8 rows)", ""],
            ["employees (1000 rows) → new_employees (7 rows)", ""],
            ["orders (800 rows) → new_orders (7 rows)", ""],
            ["", ""],
            ["⚠️ EXPECTED RESULTS:", ""],
            ["", ""],
            ["- Schema tests: PASS (if structures match)", ""],
            ["- Row count tests: FAIL (due to significant count differences)", ""],
            ["- NULL value tests: FAIL (due to pattern differences)", ""],
            ["- Data quality tests: PASS (single table validation)", ""],
            ["", ""],
            ["🚀 EXECUTION:", ""],
            ["", ""],
            ["1. Run: python execute_data_validation_tests.py", ""],
            ["2. Review detailed results and analysis", ""],
            ["3. Check validation reports for anomalies", ""],
            ["", ""],
            ["📞 SUPPORT:", ""],
            ["For issues or questions, check the validation report generated", ""],
            ["during test execution for detailed diagnostics.", ""]
        ]
        
        for row_idx, (col1, col2) in enumerate(instructions, 1):
            ws.cell(row=row_idx, column=1, value=col1).font = Font(bold=True) if col1.startswith(("📋", "🎯", "📊", "🔧", "🔍", "⚠️", "🚀", "📞")) else Font()
            ws.cell(row=row_idx, column=2, value=col2)
        
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 80
        
        return ws
    
    def create_reference_sheet(self):
        """Create REFERENCE sheet with table and parameter information"""
        ws = self.workbook.create_sheet("REFERENCE")
        
        reference_data = [
            ["📊 DATABASE REFERENCE INFORMATION", ""],
            ["", ""],
            ["🗃️ AVAILABLE TABLES:", ""],
            ["", ""],
            ["Table Name", "Row Count", "Description"],
            ["products", "1200", "Source product catalog data"],
            ["employees", "1000", "Source employee information"],
            ["orders", "800", "Source order transactions"],
            ["new_products", "8", "Target product catalog (test data)"],
            ["new_employees", "7", "Target employee information (test data)"],
            ["new_orders", "7", "Target order transactions (test data)"],
            ["validation_results", "0", "System table for storing validation results"],
            ["", ""],
            ["🔧 PARAMETER REFERENCE:", ""],
            ["", ""],
            ["Parameter", "Description", "Example Values"],
            ["source_table", "Source table name for comparison", "products, employees, orders"],
            ["target_table", "Target table name for comparison", "new_products, new_employees, new_orders"],
            ["tolerance_percent", "Acceptable percentage difference for row counts", "5, 10, 15"],
            ["ignore_sequence", "Ignore sequence differences in schema validation", "true, false"],
            ["null_match_required", "Require exact NULL pattern matching", "true, false"],
            ["check_constraints", "Include constraint validation in schema checks", "true, false"],
            ["check_ranges", "Validate data ranges in quality checks", "true, false"],
            ["check_patterns", "Validate data patterns in quality checks", "true, false"],
            ["check_duplicates", "Check for duplicate records", "true, false"],
            ["expected_variance", "Expected level of data variance", "none, small, medium, large"],
            ["report_differences", "Generate detailed difference reports", "true, false"],
            ["check_foreign_keys", "Validate foreign key relationships", "true, false"],
            ["validate_relationships", "Check referential integrity", "true, false"],
            ["check_required_fields", "Validate required field completeness", "true, false"],
            ["missing_data_threshold", "Acceptable percentage of missing data", "0, 5, 10"],
            ["", ""],
            ["🎯 TEST CATEGORIES:", ""],
            ["", ""],
            ["Category", "Purpose", "Expected Results"],
            ["SCHEMA_VALIDATION", "Compare table structures", "PASS if schemas match"],
            ["ROW_COUNT_VALIDATION", "Compare record counts", "FAIL due to count differences"],
            ["NULL_VALUE_VALIDATION", "Compare NULL patterns", "FAIL due to pattern differences"],
            ["DATA_QUALITY_VALIDATION", "Validate data quality metrics", "PASS for single table checks"]
        ]
        
        # Add headers with styling
        for row_idx, (col1, col2, *rest) in enumerate(reference_data, 1):
            cell1 = ws.cell(row=row_idx, column=1, value=col1)
            cell2 = ws.cell(row=row_idx, column=2, value=col2)
            
            if col1.startswith(("📊", "🗃️", "🔧", "🎯")):
                cell1.font = Font(bold=True, size=12)
            elif row_idx == 5 or row_idx == 17 or row_idx == 35:  # Header rows
                cell1.font = Font(bold=True)
                cell2.font = Font(bold=True)
                if rest:
                    ws.cell(row=row_idx, column=3, value=rest[0]).font = Font(bold=True)
            
            if rest:
                ws.cell(row=row_idx, column=3, value=rest[0])
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        
        return ws
    
    def save_template(self, filename="enhanced_sdm_test_suite.xlsx"):
        """Save the enhanced Excel template"""
        try:
            # Create all sheets
            print("📊 Creating CONTROLLER sheet...")
            self.create_controller_sheet()
            
            print("🔍 Creating enhanced DATAVALIDATIONS sheet...")
            self.create_enhanced_datavalidations_sheet()
            
            print("📋 Creating INSTRUCTIONS sheet...")
            self.create_instructions_sheet()
            
            print("📖 Creating REFERENCE sheet...")
            self.create_reference_sheet()
            
            # Save the workbook
            self.workbook.save(filename)
            print(f"✅ Enhanced Excel template saved as: {filename}")
            
            # Generate summary
            self.print_template_summary(filename)
            
            return True
            
        except Exception as e:
            print(f"❌ Error saving template: {str(e)}")
            return False
    
    def print_template_summary(self, filename):
        """Print summary of the generated template"""
        print(f"\n📋 ENHANCED EXCEL TEMPLATE SUMMARY")
        print("=" * 50)
        print(f"📁 File: {filename}")
        print(f"📊 Sheets: {len(self.workbook.sheetnames)}")
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            print(f"   • {sheet_name}: {sheet.max_row} rows")
        
        print(f"\n🎯 KEY ENHANCEMENTS:")
        print(f"   ✅ Proper source/target table configurations")
        print(f"   ✅ Realistic expected results (PASS/FAIL)")
        print(f"   ✅ Enhanced parameter specifications")
        print(f"   ✅ Data validation rules for dropdowns")
        print(f"   ✅ Comprehensive documentation")
        print(f"   ✅ Table reference information")
        print(f"   ✅ 12 data validation test cases")
        
        print(f"\n🚀 USAGE:")
        print(f"   1. Replace existing 'sdm_test_suite.xlsx' with this enhanced version")
        print(f"   2. Run: python execute_data_validation_tests.py")
        print(f"   3. Review enhanced validation results")


def main():
    """Generate enhanced Excel template"""
    print("🚀 ENHANCED EXCEL TEMPLATE GENERATOR")
    print("=" * 50)
    print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    generator = EnhancedExcelTemplateGenerator()
    
    # Generate the template
    success = generator.save_template()
    
    if success:
        print(f"\n✅ TEMPLATE GENERATION COMPLETE!")
        print(f"   Ready for enhanced data validation testing")
    else:
        print(f"\n❌ TEMPLATE GENERATION FAILED!")
        
    return success


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)